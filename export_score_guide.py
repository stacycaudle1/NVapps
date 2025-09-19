import openpyxl
from openpyxl.styles import Alignment

# Score Guide details for all factors (as in gui.py, September 2025)
GUIDE_DETAILS = [
    {
        "Factor": "Score",
        "Guiding Question": "How would you rate the overall performance and reliability of this system today?",
        "Scoring Help": "1–3: Poor/unstable performance or frequent issues\n4–6: Adequate for current needs but with notable limitations\n7–8: Strong performance/reliability with minor gaps\n9–10: Excellent/mission-grade performance and reliability"
    },
    {
        "Factor": "Need",
        "Guiding Question": "Is this system required to operate the business, or is it optional/nice-to-have?",
        "Scoring Help": "1–3: Optional; workarounds readily available\n4–6: Helpful but not strictly required; alternatives exist\n7–8: Required for multiple teams or core workflows\n9–10: Absolutely required to operate or meet obligations"
    },
    {
        "Factor": "Criticality",
        "Guiding Question": "What is the business impact if this system is unavailable or fails?",
        "Scoring Help": "1–3: Minor inconvenience; small/localized impact\n4–6: Moderate productivity loss; delays but manageable\n7–8: Major disruption to critical processes\n9–10: Immediate stoppage/safety/regulatory impact"
    },
    {
        "Factor": "Installed",
        "Guiding Question": "How entrenched and widely adopted is this system across the organization?",
        "Scoring Help": "1–3: Pilot/limited use; easy to replace\n4–6: Moderate adoption; switching cost exists but manageable\n7–8: Widely adopted; high dependency across teams\n9–10: Deeply entrenched; very high switching cost/risk"
    },
    {
        "Factor": "Disaster Recovery",
        "Guiding Question": "How quickly must this system be recovered after an outage, and how much data loss is tolerable?",
        "Scoring Help": "1–3: Long RTO/RPO acceptable; daily backups sufficient\n4–6: Recovery within hours required; some data loss acceptable\n7–8: Near real-time recovery/replication expected\n9–10: Continuous availability; near-zero downtime/data loss"
    },
    {
        "Factor": "Safety",
        "Guiding Question": "Could failure of this system impact human safety or regulatory health/safety obligations?",
        "Scoring Help": "1–2: No relation to safety\n3–4: Indirect/low safety relevance\n5–6: Some safety implication depending on context\n7–8: High safety implications in certain scenarios\n9–10: Life/safety critical"
    },
    {
        "Factor": "Security",
        "Guiding Question": "What is the security exposure if this system is compromised or misused?",
        "Scoring Help": "1–3: Isolated/low-sensitivity data; limited access\n4–6: Internal data with moderate sensitivity\n7–8: Handles sensitive/regulated data or external integrations\n9–10: High-risk target; critical/regulated data and broad access"
    },
    {
        "Factor": "Monetary",
        "Guiding Question": "What is the direct financial impact of downtime, failure, or poor performance?",
        "Scoring Help": "1–3: Minimal/indirect cost impact\n4–6: Moderate costs or revenue impact\n7–8: High revenue impact or material cost exposure\n9–10: Immediate material revenue loss/penalties"
    },
    {
        "Factor": "Customer Service",
        "Guiding Question": "How directly does this system affect the customer experience or service quality?",
        "Scoring Help": "1–3: Internal-only; no direct customer effect\n4–6: Indirect impact on customer interactions\n7–8: Customer-facing; noticeable impact on service quality\n9–10: Primary channel or critical to customer experience"
    },
]

def export_to_xlsx(filename):
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    ws.title = "Score Guide"
    ws.append(["Factor", "Guiding Question", "Scoring Help"])
    for col in range(1, 4):
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for detail in GUIDE_DETAILS:
        ws.append([
            detail["Factor"],
            detail["Guiding Question"],
            detail["Scoring Help"]
        ])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 48
    wb.save(filename)
    print(f"Exported Score Guide details to {filename}")

if __name__ == "__main__":
    export_to_xlsx("Score_Guide_Details.xlsx")
