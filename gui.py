import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
from datetime import datetime
import database

# Theme and config constants
ACCENT = '#0078d4'       # Primary accent blue
HEADER_BG = '#1f4e79'    # Darker header blue
HEADER_FG = '#ffffff'    # Header text color
WIN_BG = '#f5f5f5'       # Window background
VERBOSE_LOG = False

def get_risk_color(score):
    """Return tag name for risk color based on score.

    Thresholds:
    - 70–100: 'red' (High)
    - 40–69: 'yellow' (Medium)
    - 0–39: 'green' (Low)
    Any non-numeric or missing value defaults to 'green'.
    """
    try:
        if score is None:
            return 'green'
        val = float(score)
        if val >= 70:
            return 'red'
        if val >= 40:
            return 'yellow'
        return 'green'
    except Exception:
        return 'green'


class AppTracker(tk.Tk):
    def __init__(self):
        super().__init__()
        try:
            self.title('NV Applications')
        except Exception:
            pass
        try:
            self.configure(bg=WIN_BG)
        except Exception:
            pass

        # Ensure supporting tables exist (best-effort)
        try:
            conn_schema = database.connect_db()
            cur_schema = conn_schema.cursor()
            cur_schema.execute('''CREATE TABLE IF NOT EXISTS integration_categories (
                integration_id INTEGER NOT NULL,
                category_id INTEGER NOT NULL,
                PRIMARY KEY (integration_id, category_id),
                FOREIGN KEY (integration_id) REFERENCES system_integrations(id) ON DELETE CASCADE,
                FOREIGN KEY (category_id) REFERENCES categories(id) ON DELETE CASCADE
            )''')
            conn_schema.commit()
            conn_schema.close()
        except Exception as e:
            if VERBOSE_LOG:
                print(f"DEBUG: Failed to ensure integration_categories table exists: {e}")

        # Build UI
        try:
            self.setup_style()
        except Exception:
            pass
        try:
            self.create_widgets()
        except Exception:
            pass

        # Initial data population (best-effort; UI may not have these yet depending on layout)
        try:
            if VERBOSE_LOG:
                print("DEBUG: Initializing departments")
            departments = self.get_departments()
            if VERBOSE_LOG:
                print(f"DEBUG: Got departments: {departments}")
            if hasattr(self, 'department_listbox') and self.department_listbox is not None:
                self.department_listbox.delete(0, 'end')
                for _dept_id, dept_name in departments:
                    if VERBOSE_LOG:
                        print(f"DEBUG: Inserting department: {dept_name}")
                    self.department_listbox.insert('end', dept_name)
        except Exception:
            pass

        try:
            self.populate_category_listbox()
        except Exception:
            pass

        try:
            self.refresh_table()
        except Exception:
            pass

        # Ensure submit button starts disabled & primary style
        try:
            if hasattr(self, 'submit_button') and self.submit_button is not None:
                self.submit_button.state(['disabled'])
                self.submit_button.configure(style='Primary.TButton')
        except Exception:
            pass


    def setup_style(self):
        style = ttk.Style(self)
        # prefer 'clam' for consistent Treeview heading styling; fall back to native themes
        try:
            if 'clam' in style.theme_names():
                style.theme_use('clam')
            else:
                for theme in ('vista', 'xpnative', 'clam'):
                    try:
                        style.theme_use(theme)
                        break
                    except Exception:
                        pass
        except Exception:
            pass
            
        # Configure notebook (tabs) styling with blue colors
        style.configure('TNotebook', background=ACCENT)
        style.configure('TNotebook.Tab', background=HEADER_BG, foreground='white', padding=[10, 4])
        # Configure active tab
        style.map('TNotebook.Tab', background=[('selected', ACCENT)], foreground=[('selected', 'white')])
        default_font = ('Segoe UI', 10)
        # Denser font and smaller row height for tighter table layout
        tree_font = ('Segoe UI', 8)
        heading_font = ('Segoe UI', 10, 'bold')
        style.configure('.', font=default_font)
        # Reduce rowheight to make rows closer together
        style.configure('Treeview', rowheight=22, font=tree_font, background='white', fieldbackground='white')
        # Tighten heading padding so columns sit closer together
        # Slightly more horizontal padding so heading text doesn't crowd the borders
        style.configure('Treeview.Heading', font=heading_font, background=HEADER_BG, foreground=HEADER_FG, relief='flat', borderwidth=0, padding=(8,4))
        # Ensure mapping works across themes
        try:
            style.map('Treeview.Heading', background=[('active', HEADER_BG), ('!disabled', HEADER_BG)], foreground=[('active', HEADER_FG), ('!disabled', HEADER_FG)])
        except Exception:
            pass
        style.configure('TButton', padding=8, font=('Segoe UI', 10))
        style.configure('TEntry', padding=6, font=('Segoe UI', 10))
        style.map('TButton', foreground=[('active', '!disabled', 'black')])
        # frame and window backgrounds
        style.configure('TFrame', background=WIN_BG)
        # Primary, Secondary and Danger button styles (uniform mapping)
        style.configure('Primary.TButton', background=ACCENT, foreground='white', borderwidth=0)
        style.map('Primary.TButton', background=[('active', '!disabled', '#005a9e')], foreground=[('disabled', '#d0d0d0')])
        style.configure('Secondary.TButton', background='#e1e1e1', foreground='black', borderwidth=0)
        style.map('Secondary.TButton', background=[('active', '!disabled', '#cfcfcf')], foreground=[('disabled', '#a0a0a0')])
        style.configure('Danger.TButton', background='#f8d7da', foreground='#8b0000', borderwidth=0)
        style.map('Danger.TButton', background=[('active', '!disabled', '#f5c6cb')])
        # Green button style for note functions
        style.configure('Success.TButton', background='#90EE90', foreground='black', borderwidth=0)  # Light green
        style.map('Success.TButton', background=[('active', '!disabled', '#7CCD7C')], foreground=[('disabled', '#a0a0a0')])

        # LabelFrame and label styles for dialogs and guides
        style.configure('Guide.TLabelframe', background=WIN_BG)
        style.configure('Guide.TLabelframe.Label', background=WIN_BG, foreground=ACCENT, font=('Segoe UI', 10, 'bold'))
        style.configure('GuideNumber.TLabel', background=WIN_BG, foreground='#333333', font=('Segoe UI', 10, 'bold'))
        style.configure('GuideText.TLabel', background=WIN_BG, foreground='#333333', font=('Segoe UI', 10))
        style.configure('FormLabel.TLabel', background=WIN_BG, foreground=ACCENT, font=('Segoe UI', 10, 'bold'))

    def get_departments(self):
        """Get all business units (departments) from the database"""
        conn = database.connect_db()
        c = conn.cursor()
        c.execute('SELECT id, name FROM business_units ORDER BY name ASC')
        departments = c.fetchall()
        conn.close()
        if VERBOSE_LOG:
            print(f"DEBUG: Retrieved business units: {departments}")  # Debugging log
        return departments

    def get_categories(self):
        """Get all categories from the database"""
        try:
            cats = database.get_categories()
            # Normalize to list of tuples (id, name)
            result = []
            for row in cats:
                try:
                    if hasattr(row, 'keys'):
                        result.append((row['id'], row['name']))
                    else:
                        result.append((row[0], row[1]))
                except Exception:
                    pass
            return result
        except Exception as e:
            if VERBOSE_LOG:
                print(f"DEBUG: Failed to get categories: {e}")
            return []

    def populate_category_listbox(self):
        """Refresh the category listbox items from DB."""
        if not hasattr(self, 'category_listbox') or self.category_listbox is None:
            return
        try:
            self.category_listbox.delete(0, 'end')
            for cid, cname in self.get_categories():
                if VERBOSE_LOG:
                    print(f"DEBUG: Inserting category: {cname}")
                self.category_listbox.insert('end', cname)
        except Exception:
            pass

    def indicate_duplicate_category(self, name: str):
        """Visually indicate a duplicate category attempt using the main 'Add Category' button.
        Fades button from its primary blue to red, shows warning, selects existing category, then fades back."""
        try:
            # Locate the main Add Category button if we stored it earlier
            # Button is created as tk.Button in create_widgets; provide type hint for linters
            btn = getattr(self, 'main_add_category_button', None)  # type: ignore[attr-defined]
            if btn is None:
                # Attempt to discover by text (last resort); this may fail silently
                for child in self.winfo_children():
                    try:
                        if isinstance(child, ttk.Frame):
                            for sub in child.winfo_children():
                                try:
                                    if hasattr(sub, 'cget') and sub.cget('text') == 'Add Category':
                                        btn = sub
                                        break
                                except Exception:
                                    pass
                        if btn:
                            break
                    except Exception:
                        pass
            if btn is None:
                # Fallback: just warn
                messagebox.showwarning('Duplicate Category', f'Categories must be unique. "{name}" is already in the list.')
                return

            # Fade using direct background updates (tk.Button supports bg)
            start_rgb = (0, 120, 212)   # #0078d4
            end_rgb = (200, 30, 30)     # red tone
            steps = 16
            interval = 40

            def lerp(a,b,t):
                return int(a + (b-a)*t)

            def fade_forward(i=0):
                try:
                    t = i/float(steps)
                    r = lerp(start_rgb[0], end_rgb[0], t)
                    g = lerp(start_rgb[1], end_rgb[1], t)
                    b = lerp(start_rgb[2], end_rgb[2], t)
                    try:
                        btn.config(bg=f'#{r:02x}{g:02x}{b:02x}')  # type: ignore[attr-defined]
                    except Exception:
                        pass
                except Exception:
                    pass
                if i < steps:
                    self.after(interval, lambda: fade_forward(i+1))
                else:
                    show_warning_and_select()

            def show_warning_and_select():
                try:
                    messagebox.showwarning('Duplicate Category', f'Categories must be unique. "{name}" is already in the list.')
                except Exception:
                    pass
                # Select the existing entry
                try:
                    target = name.strip().lower()
                    for idx in range(self.category_listbox.size()):
                        if self.category_listbox.get(idx).strip().lower() == target:
                            self.category_listbox.selection_clear(0,'end')
                            self.category_listbox.selection_set(idx)
                            self.category_listbox.see(idx)
                            break
                except Exception:
                    pass
                fade_back()

            def fade_back(i=0):
                try:
                    t = i/float(steps)
                    r = lerp(end_rgb[0], start_rgb[0], t)
                    g = lerp(end_rgb[1], start_rgb[1], t)
                    b = lerp(end_rgb[2], start_rgb[2], t)
                    try:
                        btn.config(bg=f'#{r:02x}{g:02x}{b:02x}')  # type: ignore[attr-defined]
                    except Exception:
                        pass
                except Exception:
                    pass
                if i < steps:
                    self.after(interval, lambda: fade_back(i+1))
                else:
                    try:
                        btn.config(bg='#0078d4')  # type: ignore[attr-defined]
                    except Exception:
                        pass

            fade_forward()
        except Exception:
            pass

    def add_category_popup(self):
        popup = tk.Toplevel(self)
        popup.title('Add Category')
        ttk.Label(popup, text='Category:').pack(padx=10, pady=5)
        entry = ttk.Entry(popup)
        entry.pack(padx=10, pady=5)

        def show_dup_and_flash(name):
            # Trigger fade on main Add Category button
            try:
                self.indicate_duplicate_category(name)
            except Exception:
                pass
            # After fade-to-red completes the helper will show warning & fade back

        def add_cat():
            name = entry.get().strip()
            if not name:
                return
            existing = [self.category_listbox.get(i).strip().lower() for i in range(self.category_listbox.size())]
            if name.lower() in existing:
                show_dup_and_flash(name)
                return
            try:
                database.ensure_category(name)
                self.populate_category_listbox()
                popup.destroy()
            except Exception as e:
                if VERBOSE_LOG:
                    print(f"DEBUG: Failed to add category: {e}")
                messagebox.showerror('Error', f'Failed to add category: {e}')

        ttk.Button(popup, text='Add', command=add_cat, style='Primary.TButton').pack(padx=10, pady=10)

    def manage_categories_popup(self):
        popup = tk.Toplevel(self)
        popup.title('Manage Categories')
        listbox = tk.Listbox(popup, selectmode='browse', exportselection=0)
        cats = self.get_categories()
        for _, cname in cats:
            listbox.insert('end', cname)
        listbox.pack(padx=10, pady=10)

        def get_selected_category_id():
            sel = listbox.curselection()
            if not sel:
                return None
            idx = sel[0]
            try:
                cid = cats[idx][0]
                return cid
            except Exception:
                return None

        def delete_selected():
            cid = get_selected_category_id()
            if not cid:
                return
            if not messagebox.askyesno('Confirm Delete', 'Delete selected category? Applications will have no category set.'):
                return
            try:
                database.delete_category(cid)
                # refresh both listboxes
                self.populate_category_listbox()
                popup.destroy()
            except Exception as e:
                messagebox.showerror('Error', f'Failed to delete category: {e}')

        def rename_selected():
            cid = get_selected_category_id()
            if not cid:
                return
            new = simpledialog.askstring('Rename Category', 'New name:')
            if not new:
                return
            try:
                database.update_category(cid, new.strip())
                self.populate_category_listbox()
                popup.destroy()
            except Exception as e:
                messagebox.showerror('Error', f'Failed to rename category: {e}')

        btns = ttk.Frame(popup)
        btns.pack(padx=10, pady=10, fill='x')
        ttk.Button(btns, text='Rename', command=rename_selected, style='Primary.TButton').pack(side='left', padx=5)
        ttk.Button(btns, text='Delete', command=delete_selected, style='Danger.TButton').pack(side='left', padx=5)

    def add_department_popup(self):
        popup = tk.Toplevel(self)
        popup.title('Add Department')
        ttk.Label(popup, text='Business Unit:').pack(padx=10, pady=5)
        dept_entry = ttk.Entry(popup)
        dept_entry.pack(padx=10, pady=5)

        def add_dept():
            dept_name = dept_entry.get().strip()
            if not dept_name:
                return
            conn = database.connect_db()
            c = conn.cursor()
            try:
                # Upsert business unit name and update last_modified when it already exists
                c.execute('''INSERT INTO business_units (name, last_modified) VALUES (?, CURRENT_TIMESTAMP)
                             ON CONFLICT(name) DO UPDATE SET last_modified = CURRENT_TIMESTAMP''', (dept_name,))
                conn.commit()
            finally:
                conn.close()
            # Update department_listbox with new departments
            self.department_listbox.delete(0, 'end')
            for dept_id, dept_name in self.get_departments():
                self.department_listbox.insert('end', dept_name)
            # No need to update filter_combo anymore as we're using search functionality
            popup.destroy()

        ttk.Button(popup, text='Add', command=add_dept, style='Primary.TButton').pack(padx=10, pady=10)

    def add_division_popup(self):
        """Popup to add a new Division (stored in applications.division when new apps are created)."""
        popup = tk.Toplevel(self)
        popup.title('Add Division')
        ttk.Label(popup, text='Division:').pack(padx=10, pady=5)
        div_entry = ttk.Entry(popup)
        div_entry.pack(padx=10, pady=5)

        def add_div():
            div_name = div_entry.get().strip()
            if not div_name:
                return
            # We just insert a placeholder application to register division? Avoid that; instead we
            # keep a lightweight memory list by inserting into a helper table or simply rely on
            # existing apps. For simplicity, create an inert row only if division not present via an app.
            db_success = False
            try:
                # If a division already exists as an application, just touch its last_modified
                conn = database.connect_db()
                c = conn.cursor()
                c.execute('SELECT id FROM applications WHERE TRIM(division) = ? LIMIT 1', (div_name,))
                row = c.fetchone()
                conn.close()
                if row and row[0]:
                    try:
                        database.touch_application(int(row[0]))
                        db_success = True
                    except Exception:
                        db_success = False
                else:
                    # Create a new application row for this division
                    database.add_application(div_name, '', {}, [], notes='')
                    db_success = True
            except Exception as e:
                # Common case may be 'database is locked' or other DB errors; inform the user and fall back to local update
                try:
                    messagebox.showwarning('Database Busy', f'Could not add division to database right now: {e}. The division will be shown temporarily until DB becomes available.')
                except Exception:
                    pass
            # Refresh division list — if DB write failed, still insert the division locally so it appears immediately
            try:
                if db_success:
                    self.refresh_division_listbox()
                else:
                    # Insert locally at end if not present
                    try:
                        existing = [self.division_listbox.get(i) for i in range(self.division_listbox.size())]
                        if div_name not in existing:
                            self.division_listbox.insert(tk.END, div_name)
                    except Exception:
                        pass
            except Exception:
                pass
            popup.destroy()

        ttk.Button(popup, text='Add', command=add_div, style='Primary.TButton').pack(padx=10, pady=10)

    def manage_divisions_popup(self):
        """Popup to rename or delete existing divisions (by modifying application rows)."""
        popup = tk.Toplevel(self)
        popup.title('Manage Divisions')
        popup.geometry('300x350')

        listbox = tk.Listbox(popup, selectmode='single')
        listbox.pack(fill='both', expand=True, padx=10, pady=10)
        try:
            import sqlite3
            conn = sqlite3.connect('business_apps.db')
            c = conn.cursor()
            c.execute('SELECT DISTINCT division FROM applications WHERE division IS NOT NULL AND TRIM(division) != "" ORDER BY division')
            for (d,) in c.fetchall():
                listbox.insert(tk.END, d)
            conn.close()
        except Exception:
            pass

        def get_selected_division():
            sel = listbox.curselection()
            if not sel:
                return None
            return listbox.get(sel[0])

        def rename_division():
            old = get_selected_division()
            if not old:
                return
            new = simpledialog.askstring('Rename Division', 'New name:')
            if not new:
                return
            try:
                import sqlite3
                conn = sqlite3.connect('business_apps.db')
                c = conn.cursor()
                c.execute('UPDATE applications SET division = ? WHERE division = ?', (new.strip(), old))
                conn.commit()
                conn.close()
                self.refresh_division_listbox()
                popup.destroy()
            except Exception as e:
                messagebox.showerror('Error', f'Failed to rename division: {e}')

        def delete_division():
            d = get_selected_division()
            if not d:
                return
            if not messagebox.askyesno('Delete Division', f'Delete division "{d}"? All related application rows will retain this division unless manually changed.'):
                return
            # Deleting a division without modifying apps would leave them; skipping destructive change.
            # Instead, we warn user. Could implement clearing division; for now we just abort.
            messagebox.showinfo('Notice', 'Division deletion does not remove existing application rows. Implement custom logic if needed.')

        btn_frame = ttk.Frame(popup)
        btn_frame.pack(fill='x', padx=10, pady=5)
        ttk.Button(btn_frame, text='Rename', command=rename_division, style='Primary.TButton').pack(side='left', padx=5)
        ttk.Button(btn_frame, text='Delete', command=delete_division, style='Danger.TButton').pack(side='left', padx=5)

    def create_table_with_scrollbars(self, parent, columns, widths=None, sort_handler=None, pack_opts=None,
                                     anchors=None, formatters=None, zebra=True, zebra_colors=("#ffffff", "#f7f7f7")):
        """Create a Treeview inside a scrollable container with V/H scrollbars.
        Returns (container_frame, tree).
        - parent: parent widget
        - columns: sequence of column names
        - widths: optional dict of {col: width}
        - sort_handler: callable(tree, col, reverse) used by header commands; defaults to self.report_sort_table
        - pack_opts: optional dict passed to container.pack
        - anchors: optional dict of {col: 'w'|'center'|'e'} to override alignment
        - formatters: optional dict of {col: callable(value) -> formatted_value}
        - zebra: whether to enable zebra helper methods on the tree
        - zebra_colors: tuple (even_bg, odd_bg)
        """
        container = ttk.Frame(parent)
        # Default pack options
        opts = {'fill': 'both', 'expand': True}
        if pack_opts:
            opts.update(pack_opts)
        container.pack(**opts)

        tree = ttk.Treeview(container, columns=columns, show='headings')
        # Configure columns
        widths = widths or {}
        anchors = anchors or {}
        for col in columns:
            handler = sort_handler or (lambda tr, c, r: self.report_sort_table(tr, c, r))
            tree.heading(col, text=col, anchor='w', command=lambda c=col: handler(tree, c, False))
            # Guess alignment: text left, numeric centered
            left_cols = {'Business Unit', 'Division', 'Integration Name', 'Vendor', 'Status', 'Notes', 'System Name'}
            anchor = anchors.get(col, ('w' if col in left_cols else 'center'))
            w = widths.get(col, 120)
            tree.column(col, width=w, anchor=anchor, stretch=True if col in left_cols else False)

        # Scrollbars
        vsb = ttk.Scrollbar(container, orient='vertical', command=tree.yview)
        hsb = ttk.Scrollbar(container, orient='horizontal', command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree.pack(side='left', fill='both', expand=True)
        vsb.pack(side='right', fill='y')
        hsb.pack(side='bottom', fill='x')

        # Auto-hide horizontal scrollbar logic
        def _update_hsb_visibility(event=None):
            try:
                total_width = 0
                for c in columns:
                    total_width += int(tree.column(c, 'width'))
                visible = tree.winfo_width() or container.winfo_width()
                if total_width > max(visible, 1):
                    if not hsb.winfo_manager():
                        hsb.pack(side='bottom', fill='x')
                else:
                    if hsb.winfo_manager():
                        hsb.pack_forget()
            except Exception:
                # If anything goes wrong, keep the scrollbar visible
                if not hsb.winfo_manager():
                    hsb.pack(side='bottom', fill='x')

        tree.bind('<Configure>', _update_hsb_visibility)
        container.bind('<Configure>', _update_hsb_visibility)
        container.after(50, _update_hsb_visibility)

        # Attach helpers (returned to caller)
        col_order = list(columns)
        col_formatters = formatters or {}

        def _apply_format(values):
            try:
                if isinstance(values, dict):
                    seq = [values.get(col) for col in col_order]
                else:
                    seq = list(values)
                out = []
                for i, col in enumerate(col_order):
                    val = seq[i] if i < len(seq) else None
                    fmt = col_formatters.get(col)
                    try:
                        out.append(fmt(val) if fmt else val)
                    except Exception:
                        out.append(val)
                return tuple(out)
            except Exception:
                return values

        if zebra:
            def _apply_zebra():
                try:
                    even_bg, odd_bg = zebra_colors
                    tree.tag_configure('zebra_even', background=even_bg)
                    tree.tag_configure('zebra_odd', background=odd_bg)
                    for idx, item in enumerate(tree.get_children()):
                        tag = 'zebra_odd' if idx % 2 else 'zebra_even'
                        current = tree.item(item, 'tags') or ()
                        # Preserve existing tags (e.g., risk colors) and append zebra tag
                        if tag not in current:
                            tree.item(item, tags=(*current, tag))
                except Exception:
                    pass
            apply_zebra = _apply_zebra
        else:
            apply_zebra = lambda: None

        api = {'format': _apply_format, 'apply_zebra': apply_zebra}
        return container, tree, api

    def report_sort_table(self, tree, col, reverse):
        # Sort function specifically for the report window
        try:
            # Get data for sorting
            data = [(tree.set(k, col), k) for k in tree.get_children('')]
            
            # Sort the data based on appropriate type conversion
            data.sort(reverse=reverse, key=lambda t: self._convert_report_value(t[0]))
            
            # Rearrange items in the tree
            for index, (val, k) in enumerate(data):
                tree.move(k, '', index)
                
            # Update the heading to show sort direction and set command for next sort
            sort_direction = "▼" if reverse else "▲"  # Down arrow for descending, up for ascending
            
            # Get original heading text without any arrows
            heading_text = tree.heading(col, 'text').replace("▲", "").replace("▼", "").strip()
            
            # Reset all column headings to remove any previous sort indicators
            for column in tree['columns']:
                if column != col:  # Skip the column we're currently sorting
                    current_text = tree.heading(column, 'text').replace("▲", "").replace("▼", "").strip()
                    tree.heading(column, text=current_text)
            
            # Set new heading with sort indicator
            tree.heading(col, text=f"{heading_text} {sort_direction}", 
                        command=lambda: self.report_sort_table(tree, col, not reverse))
                
        except Exception as e:
            messagebox.showerror('Error', f'Failed to sort report table: {e}')
    
    def _convert_report_value(self, value):
        # Helper to convert values for sorting in reports
        if not value:
            return ""  # Empty values sort first
        
        # Try to convert to number
        try:
            return float(value)
        except ValueError:
            pass
        
        # Default: return lowercase string for case-insensitive sorting
        return value.lower() if isinstance(value, str) else value
    
    def update_criticality_chart(self, chart_frame, sort_by):
        """Update the criticality chart based on the selected sorting option"""
        # Clear existing chart
        for widget in chart_frame.winfo_children():
            widget.destroy()
            
        # Get division data from database
        conn = database.connect_db()
        c = conn.cursor()
        c.execute('''SELECT division, criticality, risk_score 
                     FROM applications 
                     ORDER BY division''')
        data = c.fetchall()
        conn.close()
        
        if not data:
            # Show a message if no data is available
            msg = tk.Label(chart_frame, text="No system data available", 
                         font=('Segoe UI', 12), fg='gray')
            msg.pack(expand=True)
            return
            
        # Chart functionality temporarily disabled
        msg = tk.Label(chart_frame, text="Chart functionality temporarily disabled", 
                     font=('Segoe UI', 12), fg='gray')
        msg.pack(expand=True)
    
    def debug_show_table_data(self):
        """Debug helper to check data in tables"""
        conn = database.connect_db()
        conn.row_factory = database.sqlite3.Row
        c = conn.cursor()
        
        # Check system_integrations
        print("\nChecking system_integrations table:")
        c.execute("SELECT COUNT(*) as count FROM system_integrations")
        count = c.fetchone()['count']
        print(f"Total integrations: {count}")
        if count > 0:
            c.execute("SELECT * FROM system_integrations LIMIT 3")
            print("Sample rows:")
            for row in c.fetchall():
                print(dict(row))
        
        # Check applications
        print("\nChecking applications table:")
        c.execute("SELECT COUNT(*) as count FROM applications")
        count = c.fetchone()['count']
        print(f"Total applications: {count}")
        if count > 0:
            c.execute("SELECT * FROM applications LIMIT 3")
            print("Sample rows:")
            for row in c.fetchall():
                print(dict(row))
        
        # Check business_units
        print("\nChecking business_units table:")
        c.execute("SELECT COUNT(*) as count FROM business_units")
        count = c.fetchone()['count']
        print(f"Total business units: {count}")
        if count > 0:
            c.execute("SELECT * FROM business_units LIMIT 3")
            print("Sample rows:")
            for row in c.fetchall():
                print(dict(row))
        
        # Check relationships
        print("\nChecking relationships:")
        c.execute("""
            SELECT 
                i.id as integration_id,
                i.name as integration_name,
                a.id as app_id,
                a.name as app_name,
                bu.id as bu_id,
                bu.name as bu_name
            FROM system_integrations i
            LEFT JOIN applications a ON i.parent_app_id = a.id
            LEFT JOIN application_business_units abu ON a.id = abu.app_id
            LEFT JOIN business_units bu ON abu.unit_id = bu.id
            LIMIT 5
        """)
        print("Sample joined rows:")
        for row in c.fetchall():
            print(dict(row))
        
        conn.close()

    def show_report(self):
        # Debug: check table data first
        self.debug_show_table_data()
        
        report_win = tk.Toplevel(self)
        report_win.title('System Reports')
        report_win.configure(bg=WIN_BG)
        report_win.geometry('1200x800')
        # Ensure we have an in-memory store for report refresh timestamps and labels
        if not hasattr(self, '_report_refresh_state'):
            # keys: 'risk', 'bu', 'crit' -> {'ts': datetime or None, 'label': widget}
            self._report_refresh_state = {}

        # Helper to format elapsed time
        def _format_elapsed(ts):
            if not ts:
                return ''
            delta = datetime.now() - ts
            seconds = int(delta.total_seconds())
            if seconds < 60:
                return f"{seconds}s ago"
            minutes, sec = divmod(seconds, 60)
            if minutes < 60:
                return f"{minutes}m {sec}s ago"
            hours, minutes = divmod(minutes, 60)
            if hours < 24:
                return f"{hours}h {minutes}m ago"
            days, hours = divmod(hours, 24)
            return f"{days}d {hours}h ago"

        # Update a specific label from the stored timestamp
        def _update_report_label(key):
            state = self._report_refresh_state.get(key)
            if not state or 'label' not in state:
                return
            lbl = state['label']
            ts = state.get('ts')
            try:
                if ts:
                    lbl.config(text=f"Last refresh: {ts.strftime('%Y-%m-%d %H:%M:%S')} ({_format_elapsed(ts)})")
                else:
                    lbl.config(text='Last refresh: N/A')
            except Exception:
                pass

        # Periodically update the relative elapsed text for all report labels while the window exists
        def _refresh_elapsed_labels():
            try:
                if not report_win.winfo_exists():
                    return
                for k in list(self._report_refresh_state.keys()):
                    _update_report_label(k)
            except Exception:
                pass
            try:
                report_win.after(1000, _refresh_elapsed_labels)
            except Exception:
                # If scheduling fails, silently stop
                pass
        
        # Create a notebook for different report tabs
        notebook = ttk.Notebook(report_win)
        notebook.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Risk Range Integrations Tab
        risk_frame = ttk.Frame(notebook)
        notebook.add(risk_frame, text='Risk Range Report')
        
        # Add header with title and legend
        risk_header = tk.Frame(risk_frame, bg=WIN_BG)
        risk_header.pack(fill='x', pady=(10, 15))
        risk_label = tk.Label(risk_header, text='Integrations by Risk Range', 
                          font=('Segoe UI', 14, 'bold'), fg=ACCENT, bg=WIN_BG)
        risk_label.pack(side='left', padx=15)
        # Color legend on the right
        risk_legend = tk.Frame(risk_header, bg=WIN_BG)
        risk_legend.pack(side='right', padx=15)
        def _legend_chip_rf(parent, text, bg):
            tk.Label(parent, text=text, bg=bg, fg='black', padx=6, pady=2).pack(side='left', padx=3)
        _legend_chip_rf(risk_legend, 'High (≥70)', '#ffcccc')
        _legend_chip_rf(risk_legend, 'Med (40–69)', '#fff2cc')
        _legend_chip_rf(risk_legend, 'Low (1–39)', '#ccffcc')
        
        # Controls frame for risk range selection
        controls_frame = ttk.Frame(risk_frame)
        controls_frame.pack(fill='x', padx=15, pady=5)

        # Risk range selector
        ttk.Label(controls_frame, text="Risk Range:").pack(side='left', padx=5)
        risk_var = tk.StringVar(value="All")
        risk_options = ttk.Combobox(controls_frame, textvariable=risk_var, 
                                   values=["All", "Low (1-39)", "Med (40-69)", "High (70+)"],
                                   width=15, state='readonly')
        risk_options.pack(side='left', padx=5)

        # Create the risk-filtered table with shared helper
        columns = ('Business Unit', 'Division', 'Integration Name', 'Vendor', 
                  'Score', 'Criticality', 'Risk', 'Notes')
        widths = {
            'Business Unit': 200, 'Division': 180, 'Integration Name': 200,
            'Vendor': 150, 'Score': 80, 'Criticality': 80, 'Risk': 80, 'Notes': 200
        }
        # Define consistent numeric formatters for this table
        risk_formatters = {
            'Score': lambda v: '' if v in (None, '') else f"{int(float(v))}",
            'Criticality': lambda v: '' if v in (None, '') else f"{int(float(v))}",
            'Risk': lambda v: '' if v in (None, '') else f"{float(v):.1f}",
        }
        _, risk_tree, risk_api = self.create_table_with_scrollbars(
            risk_frame,
            columns,
            widths=widths,
            sort_handler=self.report_sort_table,
            formatters=risk_formatters,
            pack_opts={'fill': 'both', 'expand': True, 'padx': 15, 'pady': 5}
        )

        # Small label to show last refresh time for Risk Range
        risk_last_lbl = ttk.Label(controls_frame, text='Last refresh: N/A')
        risk_last_lbl.pack(side='left', padx=(10,0))
        # Register label widget in in-memory report refresh state
        try:
            self._report_refresh_state['risk'] = {'label': risk_last_lbl, 'ts': None}
        except Exception:
            pass

        def update_risk_table(*args):
            risk_tree.delete(*risk_tree.get_children())

            conn = database.connect_db()
            conn.row_factory = database.sqlite3.Row
            c = conn.cursor()

            # Build the risk range condition based on integration risk scores
            risk_selection = risk_var.get()
            if risk_selection == "Low (1-39)":
                risk_condition = "AND i.risk_score < 40 AND i.risk_score >= 1"
            elif risk_selection == "Med (40-69)":
                risk_condition = "AND i.risk_score >= 40 AND i.risk_score < 70"
            elif risk_selection == "High (70+)":
                risk_condition = "AND i.risk_score >= 70"
            else:
                risk_condition = ""
            print(f"Selected risk range: {risk_selection}, condition: {risk_condition}")

            # Debug: print the query being executed
            query = f'''
                SELECT DISTINCT
                    bu.name as business_unit,
                    a.name as division,
                    i.name as integration_name,
                    i.vendor,
                    i.score,
                    i.criticality,
                    i.risk_score as risk,
                    i.notes
                FROM system_integrations i
                JOIN applications a ON i.parent_app_id = a.id
                JOIN application_business_units abu ON a.id = abu.app_id
                JOIN business_units bu ON abu.unit_id = bu.id
                WHERE 1=1 {risk_condition}
                ORDER BY bu.name, a.name, i.name
            '''
            print(f"Executing query: {query}")
            try:
                c.execute(query)
                results = c.fetchall()
                print(f"Found {len(results)} rows")
            except Exception as e:
                print(f"Query error: {str(e)}")
                messagebox.showerror("Error", f"Failed to fetch data: {str(e)}")
                return

            for row in results:
                # Build raw values with numeric fields unformatted; let the helper apply formatting
                raw_values = (
                    row['business_unit'],
                    row['division'],
                    row['integration_name'],
                    row['vendor'],
                    row['score'],  # int-like
                    row['criticality'],  # int-like
                    (row['risk'] if row['risk'] is not None else 0.0),  # float-like
                    row['notes'] or ""
                )
                values = risk_api['format'](raw_values)
                tag = get_risk_color(row['risk'])
                # Only pass tag if it's not None
                if tag:
                    risk_tree.insert('', 'end', values=values, tags=(tag,))
                else:
                    risk_tree.insert('', 'end', values=values)

            # Apply zebra striping for readability (preserves risk color tags)
            try:
                risk_api['apply_zebra']()
            except Exception:
                pass

            conn.close()

            # Configure row colors based on risk
            risk_tree.tag_configure('red', background='#ffcccc')
            risk_tree.tag_configure('yellow', background='#fff2cc')
            risk_tree.tag_configure('green', background='#ccffcc')
            # Update last refresh timestamp in in-memory state and refresh label text
            try:
                state = self._report_refresh_state.get('risk')
                if state is None:
                    self._report_refresh_state['risk'] = {'label': risk_last_lbl, 'ts': datetime.now()}
                else:
                    state['ts'] = datetime.now()
                _update_report_label('risk')
            except Exception:
                pass

        # Add Refresh button now that update_risk_table is defined
        ttk.Button(controls_frame, text='Refresh', command=update_risk_table, style='Primary.TButton').pack(side='left', padx=5)

        # Export to XLSX functionality
        # Simple toast notification at bottom of report window
        def _show_toast(msg: str):
            try:
                toast = tk.Toplevel(report_win)
                toast.overrideredirect(True)
                toast.configure(bg='#222222')
                tk.Label(toast, text=msg, bg='#222222', fg='white', font=('Segoe UI', 9)).pack(padx=14, pady=8)
                # Position bottom-right of report window
                report_win.update_idletasks()
                x = report_win.winfo_rootx() + report_win.winfo_width() - 260
                y = report_win.winfo_rooty() + report_win.winfo_height() - 90
                toast.geometry(f"250x40+{x}+{y}")
                # Fade out after delay
                def _fade(step=0):
                    try:
                        if step > 10:
                            toast.destroy()
                            return
                        toast.attributes('-alpha', max(0, 1 - step*0.1))
                        toast.after(60, _fade, step+1)
                    except Exception:
                        pass
                toast.after(1800, _fade)
            except Exception:
                pass

        def export_risk_report():
            try:
                # Collect current rows from treeview in appearance order
                rows = []
                for item_id in risk_tree.get_children():
                    vals = risk_tree.item(item_id, 'values')
                    # Expect order: Business Unit, Division, Integration Name, Vendor, Score, Criticality, Risk, Notes
                    rows.append(vals)
                if not rows:
                    messagebox.showinfo("Export", "No data to export.")
                    return
                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                except ImportError:
                    messagebox.showerror("Missing Dependency", "openpyxl is required. Please install with pip install openpyxl")
                    return

                wb = Workbook()
                # Annotate worksheet as Any to satisfy static analysis if openpyxl types unavailable
                from typing import Any
                ws: Any = wb.active  # type: ignore
                ws.title = "Risk Range"

                # Styles
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                bu_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                div_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                high_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # red
                med_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")   # yellow
                low_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # green
                header_font = Font(bold=True, color="FFFFFF")
                bold_font = Font(bold=True)
                center = Alignment(horizontal="center")
                thin = Side(style='thin', color='888888')
                border = Border(left=thin, right=thin, top=thin, bottom=thin)

                # Write report metadata
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
                title_cell = ws.cell(row=1, column=1, value=f"Integrations by Risk Range Export - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                title_cell.font = Font(bold=True, size=14)
                title_cell.alignment = Alignment(horizontal='center')

                # Header row
                headers = ["Business Unit", "Division", "Integration Name", "Vendor", "Score", "Criticality", "Risk", "Notes"]
                for col, h in enumerate(headers, start=1):
                    cell = ws.cell(row=3, column=col, value=h)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center
                    cell.border = border

                # Group rows by Business Unit then Division
                from collections import defaultdict
                grouped = defaultdict(lambda: defaultdict(list))
                for (bu, division, name, vendor, score, crit, risk, notes) in rows:
                    grouped[bu][division].append((name, vendor, score, crit, risk, notes))

                # Removed Score Guide legend code (not relevant in export)

                # Write data rows
                current_row = 4
                for bu, divisions in grouped.items():
                    for division, items in divisions.items():
                        for (name, vendor, score, crit, risk, notes) in items:
                            for col, val in enumerate([bu, division, name, vendor, score, crit, risk, notes], start=1):
                                cell = ws.cell(row=current_row, column=col, value=val)
                                cell.border = border
                                cell.alignment = center
                                if col == 7:  # risk column used for color selection
                                    try:
                                        rv = float(val) if val not in (None, "") else 0
                                    except Exception:
                                        rv = 0
                                    if rv >= 70:
                                        fill = high_fill
                                    elif rv >= 40:
                                        fill = med_fill
                                    else:
                                        fill = low_fill
                                    # Apply fill to all cells in this row
                                    for cc in range(1, 9):
                                        ws.cell(row=current_row, column=cc).fill = fill
                            current_row += 1
                        # Blank row after division
                        current_row += 1
                    # Extra blank after BU
                    current_row += 1

                # Auto-size columns (basic heuristic)
                for col in range(1, 9):
                    max_len = 0
                    for r in range(1, current_row):
                        val = ws.cell(row=r, column=col).value
                        if val is None:
                            continue
                        vlen = len(str(val))
                        if vlen > max_len:
                            max_len = vlen
                    ws.column_dimensions[chr(64+col)].width = min(max_len + 2, 60)

                # Ask user for destination path
                from tkinter import filedialog
                default_name = f"risk_range_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=default_name, filetypes=[("Excel Workbook", "*.xlsx")])
                if not file_path:
                    return
                try:
                    wb.save(file_path)
                    _show_toast("Risk Range export saved")
                except Exception as e:
                    messagebox.showerror("Export Failed", f"Could not save file: {e}")
            except Exception as e:
                print(f"Export error: {e}")
                messagebox.showerror("Export Failed", f"Unexpected error: {e}")

        ttk.Button(controls_frame, text='Export XLSX', command=export_risk_report).pack(side='left', padx=5)

        def export_all_reports():
            """Create one workbook with three sheets: Risk Range (current filter), Business Unit Risk, Division Risk."""
            try:
                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                except ImportError:
                    messagebox.showerror("Missing Dependency", "openpyxl is required. Please install with pip install openpyxl")
                    return
                wb = Workbook()
                from typing import Any
                # Common styles
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                center = Alignment(horizontal='center')
                thin = Side(style='thin', color='888888')
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                high_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                med_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                low_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

                # --- Risk Range Sheet (with grouping) ---
                ws_risk: Any = wb.active  # type: ignore
                ws_risk.title = "Risk Range"
                rows_risk = []
                for iid in risk_tree.get_children():
                    rows_risk.append(risk_tree.item(iid, 'values'))
                from collections import defaultdict
                bu_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                div_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                ws_risk.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
                tcell = ws_risk.cell(row=1, column=1, value=f"Risk Range (Filter: {risk_var.get()}) - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                tcell.font = Font(bold=True, size=14)
                tcell.alignment = Alignment(horizontal='center')
                risk_headers = ["Business Unit", "Division", "Integration Name", "Vendor", "Score", "Criticality", "Risk", "Notes"]
                for c, h in enumerate(risk_headers, start=1):
                    hc = ws_risk.cell(row=3, column=c, value=h)
                    hc.fill = header_fill
                    hc.font = header_font
                    hc.alignment = center
                    hc.border = border
                grouped = defaultdict(lambda: defaultdict(list))
                for (bu, division, name, vendor, score, crit, risk, notes) in rows_risk:
                    grouped[bu][division].append((name, vendor, score, crit, risk, notes))
                r = 4
                bold_font = Font(bold=True)
                for bu in sorted(grouped.keys()):
                    ws_risk.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
                    bc = ws_risk.cell(row=r, column=1, value=f"Business Unit: {bu}")
                    bc.font = bold_font
                    bc.fill = bu_fill
                    bc.border = border
                    r += 1
                    for division in sorted(grouped[bu].keys()):
                        ws_risk.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
                        dc = ws_risk.cell(row=r, column=1, value=f"Division: {division}")
                        dc.font = bold_font
                        dc.fill = div_fill
                        dc.border = border
                        r += 1
                        for (name, vendor, score, crit, risk, notes) in grouped[bu][division]:
                            rowvals = [bu, division, name, vendor, score, crit, risk, notes]
                            for c, val in enumerate(rowvals, start=1):
                                cell = ws_risk.cell(row=r, column=c, value=val)
                                cell.border = border
                            try:
                                rv = float(risk) if risk not in (None, "") else 0
                            except Exception:
                                rv = 0
                            fill = high_fill if rv >= 70 else med_fill if rv >= 40 else low_fill
                            for cc in range(1, 9):
                                ws_risk.cell(row=r, column=cc).fill = fill
                            r += 1
                        r += 1  # blank after division
                    r += 1      # extra after BU
                # Autosize
                for c in range(1, 9):
                    mlen = 0
                    for rr in range(1, r):
                        v = ws_risk.cell(row=rr, column=c).value
                        if v is None:
                            continue
                        l = len(str(v))
                        if l > mlen:
                            mlen = l
                    ws_risk.column_dimensions[chr(64+c)].width = min(mlen + 2, 60)

                # --- Business Unit Sheet ---
                ws_bu = wb.create_sheet("Business Units")
                ws_bu.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
                tb = ws_bu.cell(row=1, column=1, value=f"Business Unit Risk - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                tb.font = Font(bold=True, size=14)
                tb.alignment = Alignment(horizontal='center')
                bu_headers = ["Business Unit", "App Count", "Avg Risk", "Status"]
                for c, h in enumerate(bu_headers, start=1):
                    hc = ws_bu.cell(row=3, column=c, value=h)
                    hc.fill = header_fill
                    hc.font = header_font
                    hc.alignment = center
                    hc.border = border
                br = 4
                for iid in bu_tree.get_children():
                    bu_name, app_count, avg_risk, status = bu_tree.item(iid, 'values')
                    rowvals = [bu_name, app_count, avg_risk, status]
                    for c, val in enumerate(rowvals, start=1):
                        cell = ws_bu.cell(row=br, column=c, value=val)
                        cell.border = border
                    try:
                        rv = float(avg_risk) if avg_risk not in (None, '', '0') else 0
                    except Exception:
                        rv = 0
                    fill = None
                    if rv >= 70:
                        fill = high_fill
                    elif rv >= 40:
                        fill = med_fill
                    elif rv > 0:
                        fill = low_fill
                    if fill:
                        for cc in range(1, 5):
                            ws_bu.cell(row=br, column=cc).fill = fill
                    br += 1
                for c in range(1, 5):
                    mlen = 0
                    for rr in range(1, br):
                        v = ws_bu.cell(row=rr, column=c).value
                        if v is None:
                            continue
                        l = len(str(v))
                        if l > mlen:
                            mlen = l
                    ws_bu.column_dimensions[chr(64+c)].width = min(mlen + 2, 50)

                # --- Categories Sheet (if category tab exists in this session) ---
                ws_div = wb.create_sheet("Divisions")
                ws_div.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
                td = ws_div.cell(row=1, column=1, value=f"Division Risk - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                td.font = Font(bold=True, size=14)
                td.alignment = Alignment(horizontal='center')
                div_headers = ["Division", "App Count", "Avg Risk", "Status"]
                for c, h in enumerate(div_headers, start=1):
                    hc = ws_div.cell(row=3, column=c, value=h)
                    hc.fill = header_fill
                    hc.font = header_font
                    hc.alignment = center
                    hc.border = border
                dr = 4
                for iid in div_tree.get_children():
                    division, app_count, avg_risk, status = div_tree.item(iid, 'values')
                    rowvals = [division, app_count, avg_risk, status]
                    for c, val in enumerate(rowvals, start=1):
                        cell = ws_div.cell(row=dr, column=c, value=val)
                        cell.border = border
                    try:
                        rv = float(avg_risk) if avg_risk not in (None, '', '0') else 0
                    except Exception:
                        rv = 0
                    fill = None
                    if rv >= 70:
                        fill = high_fill
                    elif rv >= 40:
                        fill = med_fill
                    elif rv > 0:
                        fill = low_fill
                    if fill:
                        for cc in range(1, 5):
                            ws_div.cell(row=dr, column=cc).fill = fill
                    dr += 1
                for c in range(1, 5):
                    mlen = 0
                    for rr in range(1, dr):
                        v = ws_div.cell(row=rr, column=c).value
                        if v is None:
                            continue
                        l = len(str(v))
                        if l > mlen:
                            mlen = l
                    ws_div.column_dimensions[chr(64+c)].width = min(mlen + 2, 50)

                # --- Summary Sheet (moved to last, enhanced) ---
                # --- Categories Sheet (if category tab exists in this session) ---
                try:
                    ws_cat = wb.create_sheet("Categories")
                    ws_cat.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
                    tcat = ws_cat.cell(row=1, column=1, value=f"Category Risk - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                    tcat.font = Font(bold=True, size=14)
                    tcat.alignment = Alignment(horizontal='center')
                    cat_headers = ["Category", "App Count", "Avg Risk", "Status"]
                    for c, h in enumerate(cat_headers, start=1):
                        hc = ws_cat.cell(row=3, column=c, value=h)
                        hc.fill = header_fill
                        hc.font = header_font
                        hc.alignment = center
                        hc.border = border
                    cr = 4
                    try:
                        for iid in cat_tree.get_children():
                            category, app_count, avg_risk, status = cat_tree.item(iid, 'values')
                            rowvals = [category, app_count, avg_risk, status]
                            for c, val in enumerate(rowvals, start=1):
                                cell = ws_cat.cell(row=cr, column=c, value=val)
                                cell.border = border
                            try:
                                rv = float(avg_risk) if avg_risk not in (None, '', '0') else 0
                            except Exception:
                                rv = 0
                            fill = None
                            if rv >= 70:
                                fill = high_fill
                            elif rv >= 40:
                                fill = med_fill
                            elif rv > 0:
                                fill = low_fill
                            if fill:
                                for cc in range(1, 5):
                                    ws_cat.cell(row=cr, column=cc).fill = fill
                            cr += 1
                    except Exception:
                        pass
                    for c in range(1, 5):
                        mlen = 0
                        for rr in range(1, cr):
                            v = ws_cat.cell(row=rr, column=c).value
                            if v is None:
                                continue
                            l = len(str(v))
                            if l > mlen:
                                mlen = l
                        ws_cat.column_dimensions[chr(64+c)].width = min(mlen + 2, 50)
                except Exception:
                    pass

                # --- Summary Sheet (moved to last, enhanced) ---
                try:
                    ws_sum = wb.create_sheet("Summary")  # append at end
                    ws_sum.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
                    scell = ws_sum.cell(row=1, column=1, value=f"Reports Summary - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                    scell.font = Font(bold=True, size=16, color="FFFFFF")
                    scell.alignment = Alignment(horizontal='center')
                    scell.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")

                    # Gather metrics including missing/zero and distribution
                    total_integrations = 0
                    risk_values = []  # non-zero
                    zero_or_missing = 0
                    for (bu, divisions) in grouped.items():
                        for division, entries in divisions.items():
                            for (_n, _v, _s, _c, risk_val, _notes) in entries:
                                total_integrations += 1
                                try:
                                    rv = float(risk_val) if risk_val not in (None, '') else 0
                                except Exception:
                                    rv = 0
                                if rv <= 0:
                                    zero_or_missing += 1
                                else:
                                    risk_values.append(rv)
                    unique_bu = len(grouped.keys())
                    unique_divisions = sum(len(divisions.keys()) for divisions in grouped.values())
                    avg_risk_overall = round(sum(risk_values)/len(risk_values), 2) if risk_values else 0
                    high_count = len([v for v in risk_values if v >= 70])
                    med_count = len([v for v in risk_values if 40 <= v < 70])
                    low_count = len([v for v in risk_values if 1 <= v < 40])
                    dist_total = max(len(risk_values), 1)  # avoid div by zero (exclude zero/missing from distribution)
                    high_pct = f"{(high_count/dist_total)*100:.1f}%" if dist_total else '0.0%'
                    med_pct = f"{(med_count/dist_total)*100:.1f}%" if dist_total else '0.0%'
                    low_pct = f"{(low_count/dist_total)*100:.1f}%" if dist_total else '0.0%'
                    zero_pct = f"{(zero_or_missing/max(total_integrations,1))*100:.1f}%" if total_integrations else '0.0%'

                    summary_rows = [
                        ("Total Business Units", unique_bu, ''),
                        ("Total Divisions", unique_divisions, ''),
                        ("Total Integrations (visible)", total_integrations, ''),
                        ("Integrations With Missing/Zero Risk", zero_or_missing, zero_pct),
                        ("Avg Integration Risk (non-zero)", avg_risk_overall, ''),
                        ("High Risk Integrations (≥70)", high_count, high_pct),
                        ("Medium Risk Integrations (40–69)", med_count, med_pct),
                        ("Low Risk Integrations (1–39)", low_count, low_pct),
                    ]

                    # Headers
                    headers = ["Metric", "Value", "% of Non-Zero" ]
                    for c, h in enumerate(headers, start=1):
                        hc = ws_sum.cell(row=3, column=c, value=h)
                        hc.fill = header_fill
                        hc.font = header_font
                        hc.alignment = Alignment(horizontal='center')
                        hc.border = border

                    sr = 4
                    for metric, val, pct in summary_rows:
                        c1 = ws_sum.cell(row=sr, column=1, value=metric)
                        c2 = ws_sum.cell(row=sr, column=2, value=val)
                        c3 = ws_sum.cell(row=sr, column=3, value=pct)
                        for cc in (c1, c2, c3):
                            cc.border = border
                            # Center the Value and Percentage columns
                            if cc.column in (2, 3):
                                cc.alignment = Alignment(horizontal='center')
                        # Color bands for risk distribution rows (apply to entire row)
                        fill_to_apply = None
                        if 'High Risk' in metric:
                            fill_to_apply = high_fill
                        elif 'Medium Risk' in metric:
                            fill_to_apply = med_fill
                        elif 'Low Risk' in metric:
                            fill_to_apply = low_fill
                        elif 'Missing/Zero' in metric:
                            fill_to_apply = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                        if fill_to_apply:
                            for cc_idx in range(1, 4):
                                ws_sum.cell(row=sr, column=cc_idx).fill = fill_to_apply
                        sr += 1

                    # Totals emphasis (first three rows)
                    for r_emph in range(4, 7):
                        ws_sum.cell(row=r_emph, column=1).font = Font(bold=True)
                        ws_sum.cell(row=r_emph, column=2).font = Font(bold=True)

                    # Column widths
                    ws_sum.column_dimensions['A'].width = 50
                    ws_sum.column_dimensions['B'].width = 18
                    ws_sum.column_dimensions['C'].width = 18
                except Exception:
                    pass

                # Save file
                from tkinter import filedialog
                default_name = f"all_reports_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                path = filedialog.asksaveasfilename(defaultextension='.xlsx', initialfile=default_name, filetypes=[('Excel Workbook','*.xlsx')])
                if not path:
                    return
                wb.save(path)
                _show_toast("All reports exported")
            except Exception as e:
                messagebox.showerror("Export Failed", f"Multi-sheet export failed: {e}")

        ttk.Button(controls_frame, text='Export All (Multi-Sheet)', command=export_all_reports).pack(side='left', padx=5)

        # Bind the update function to combobox selection
        risk_var.trace('w', update_risk_table)

        # Initial population of the risk table
        update_risk_table()
        
        # Business Unit Risk Tab
        bu_frame = ttk.Frame(notebook)
        notebook.add(bu_frame, text='Business Unit Risk')
        
        # Add a modern header to the business unit report with legend
        bu_header = tk.Frame(bu_frame, bg=WIN_BG)
        bu_header.pack(fill='x', pady=(10, 15))
        bu_label = tk.Label(bu_header, text='Business Unit Risk Overview', 
                          font=('Segoe UI', 14, 'bold'), fg=ACCENT, bg=WIN_BG)
        bu_label.pack(side='left', padx=15)
        bu_legend = tk.Frame(bu_header, bg=WIN_BG)
        bu_legend.pack(side='right', padx=15)
        def _legend_chip_bu(parent, text, bg):
            tk.Label(parent, text=text, bg=bg, fg='black', padx=6, pady=2).pack(side='left', padx=3)
        _legend_chip_bu(bu_legend, 'High (≥70)', '#ffcccc')
        _legend_chip_bu(bu_legend, 'Med (40–69)', '#fff2cc')
        _legend_chip_bu(bu_legend, 'Low (1–39)', '#ccffcc')
        
        # Business Unit Risk Tree using shared helper
        bu_columns = ('Business Unit', 'App Count', 'Avg Risk', 'Status')
        bu_widths = {'Business Unit': 320, 'App Count': 110, 'Avg Risk': 110, 'Status': 120}
        bu_formatters = {
            'App Count': lambda v: '' if v in (None, '') else f"{int(v)}",
            'Avg Risk': lambda v: '' if v in (None, '') else f"{float(v):.1f}",
        }
        _, bu_tree, bu_api = self.create_table_with_scrollbars(
            bu_frame,
            bu_columns,
            widths=bu_widths,
            formatters=bu_formatters,
            sort_handler=self.report_sort_table,
            pack_opts={'fill': 'both', 'expand': True, 'padx': 15, 'pady': 5}
        )
        
        # Division Risk Tab (inserted between Business Unit Risk and System Criticality)
        div_frame = ttk.Frame(notebook)
        notebook.add(div_frame, text='Division Risk')
        
        # Header with legend for Division Risk
        div_header = tk.Frame(div_frame, bg=WIN_BG)
        div_header.pack(fill='x', pady=(10, 15))
        div_label = tk.Label(div_header, text='Division Risk Overview', 
                          font=('Segoe UI', 14, 'bold'), fg=ACCENT, bg=WIN_BG)
        div_label.pack(side='left', padx=15)
        div_legend = tk.Frame(div_header, bg=WIN_BG)
        div_legend.pack(side='right', padx=15)
        def _legend_chip_div(parent, text, bg):
            tk.Label(parent, text=text, bg=bg, fg='black', padx=6, pady=2).pack(side='left', padx=3)
        _legend_chip_div(div_legend, 'High (≥70)', '#ffcccc')
        _legend_chip_div(div_legend, 'Med (40–69)', '#fff2cc')
        _legend_chip_div(div_legend, 'Low (1–39)', '#ccffcc')
        
        # Division Risk Tree using shared helper (same structure as BU Risk)
        div_columns = ('Division', 'App Count', 'Avg Risk', 'Status')
        div_widths = {'Division': 320, 'App Count': 110, 'Avg Risk': 110, 'Status': 120}
        div_formatters = {
            'App Count': lambda v: '' if v in (None, '') else f"{int(v)}",
            'Avg Risk': lambda v: '' if v in (None, '') else f"{float(v):.1f}",
        }
        _, div_tree, div_api = self.create_table_with_scrollbars(
            div_frame,
            div_columns,
            widths=div_widths,
            formatters=div_formatters,
            sort_handler=self.report_sort_table,
            pack_opts={'fill': 'both', 'expand': True, 'padx': 15, 'pady': 5}
        )
        
        # Populate division data via a refreshable function (group by application name)
        def update_div_table():
            # Clear table
            div_tree.delete(*div_tree.get_children())
            conn = database.connect_db()
            c = conn.cursor()
            # Aggregate by division (application name) using average integration risk
            c.execute('''SELECT a.name as division, COUNT(a.id) as app_count, AVG(i.risk_score) as avg_integration_risk
                     FROM applications a
                     LEFT JOIN system_integrations i ON a.id = i.parent_app_id
                     GROUP BY a.name''')
            results = c.fetchall()
            for division, count, avg_risk in results:
                if avg_risk is None or avg_risk < 1:
                    status = 'No Data'
                    raw_values = (division, count, 0.0, status)
                    div_tree.insert('', 'end', values=div_api['format'](raw_values))
                else:
                    try:
                        avg = float(avg_risk)
                    except Exception:
                        avg = 0.0
                    tag = get_risk_color(avg)
                    if tag == 'red':
                        status = 'High'
                    elif tag == 'yellow':
                        status = 'Med'
                    else:
                        status = 'Low'
                    raw_values = (division, count, avg, status)
                    formatted = div_api['format'](raw_values)
                    if tag:
                        div_tree.insert('', 'end', values=formatted, tags=(tag,))
                    else:
                        div_tree.insert('', 'end', values=formatted)
            # Configure row colors based on risk (do this once per refresh)
            div_tree.tag_configure('red', background='#ffcccc')
            div_tree.tag_configure('yellow', background='#fff2cc')
            div_tree.tag_configure('green', background='#ccffcc')
            conn.close()
            # Zebra striping
            try:
                div_api['apply_zebra']()
            except Exception:
                pass
            # Update Division last-refresh timestamp in in-memory state and update label
            try:
                state = self._report_refresh_state.get('div')
                if state is None:
                    self._report_refresh_state['div'] = {'label': None, 'ts': datetime.now()}
                else:
                    state['ts'] = datetime.now()
                _update_report_label('div')
            except Exception:
                pass
        
        # Controls row for Division with Refresh button
        div_controls = ttk.Frame(div_frame)
        div_controls.pack(fill='x', padx=15, pady=(0, 5))
        ttk.Button(div_controls, text='Refresh', command=update_div_table, style='Primary.TButton').pack(side='left')
        def export_division_report():
            try:
                # Gather current division table rows
                rows = []
                for iid in div_tree.get_children():
                    rows.append(div_tree.item(iid, 'values'))  # Division, App Count, Avg Risk, Status
                if not rows:
                    messagebox.showinfo("Export", "No division data to export.")
                    return
                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                except ImportError:
                    messagebox.showerror("Missing Dependency", "openpyxl is required. Please install with pip install openpyxl")
                    return
                wb = Workbook()
                from typing import Any
                ws: Any = wb.active  # type: ignore
                ws.title = "Division Risk"
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                center = Alignment(horizontal='center')
                thin = Side(style='thin', color='888888')
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
                title = ws.cell(row=1, column=1, value=f"Division Risk Overview - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                title.font = Font(bold=True, size=14)
                title.alignment = Alignment(horizontal='center')
                headers = ["Division", "App Count", "Avg Risk", "Status"]
                for col, h in enumerate(headers, start=1):
                    ccell = ws.cell(row=3, column=col, value=h)
                    ccell.fill = header_fill
                    ccell.font = header_font
                    ccell.alignment = center
                    ccell.border = border
                row_idx = 4
                for (division, app_count, avg_risk, status) in rows:
                    data = [division, app_count, avg_risk, status]
                    for col, val in enumerate(data, start=1):
                        cell = ws.cell(row=row_idx, column=col, value=val)
                        cell.border = border
                    try:
                        rv = float(avg_risk) if avg_risk not in (None, '', '0') else 0
                    except Exception:
                        rv = 0
                    from openpyxl.styles import PatternFill
                    if rv >= 70:
                        fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    elif rv >= 40:
                        fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                    elif rv > 0:
                        fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    else:
                        fill = None
                    if fill:
                        for c in range(1, 5):
                            ws.cell(row=row_idx, column=c).fill = fill
                    row_idx += 1
                for col in range(1, 5):
                    max_len = 0
                    for r in range(1, row_idx):
                        val = ws.cell(row=r, column=col).value
                        if val is None:
                            continue
                        l = len(str(val))
                        if l > max_len:
                            max_len = l
                    ws.column_dimensions[chr(64+col)].width = min(max_len + 2, 50)
                from tkinter import filedialog
                default_name = f"division_risk_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                path = filedialog.asksaveasfilename(defaultextension='.xlsx', initialfile=default_name, filetypes=[('Excel Workbook','*.xlsx')])
                if not path:
                    return
                wb.save(path)
                _show_toast("Division export saved")
            except Exception as e:
                messagebox.showerror("Export Failed", f"Division export failed: {e}")
        ttk.Button(div_controls, text='Export XLSX', command=export_division_report).pack(side='left', padx=5)
        div_last_lbl = ttk.Label(div_controls, text='Last refresh: N/A')
        div_last_lbl.pack(side='left', padx=(8,0))
        # Register DIV label widget for elapsed updates
        try:
            st = self._report_refresh_state.get('div')
            if st is None:
                self._report_refresh_state['div'] = {'label': div_last_lbl, 'ts': None}
            else:
                st['label'] = div_last_lbl
        except Exception:
            pass

        # Initial population
        update_div_table()
        
        # Category Risk Tab (new)
        cat_frame = ttk.Frame(notebook)
        notebook.add(cat_frame, text='Category Risk')
        
        # Header with legend for Category Risk
        cat_header = tk.Frame(cat_frame, bg=WIN_BG)
        cat_header.pack(fill='x', pady=(10, 15))
        cat_label = tk.Label(cat_header, text='Category Risk Overview', 
                          font=('Segoe UI', 14, 'bold'), fg=ACCENT, bg=WIN_BG)
        cat_label.pack(side='left', padx=15)
        cat_legend = tk.Frame(cat_header, bg=WIN_BG)
        cat_legend.pack(side='right', padx=15)
        def _legend_chip_cat(parent, text, bg):
            tk.Label(parent, text=text, bg=bg, fg='black', padx=6, pady=2).pack(side='left', padx=3)
        _legend_chip_cat(cat_legend, 'High (≥70)', '#ffcccc')
        _legend_chip_cat(cat_legend, 'Med (40–69)', '#fff2cc')
        _legend_chip_cat(cat_legend, 'Low (1–39)', '#ccffcc')
        
        # Category Risk Tree using shared helper
        cat_columns = ('Category', 'App Count', 'Avg Risk', 'Status')
        cat_widths = {'Category': 320, 'App Count': 110, 'Avg Risk': 110, 'Status': 120}
        cat_formatters = {
            'App Count': lambda v: '' if v in (None, '') else f"{int(v)}",
            'Avg Risk': lambda v: '' if v in (None, '') else f"{float(v):.1f}",
        }
        _, cat_tree, cat_api = self.create_table_with_scrollbars(
            cat_frame,
            cat_columns,
            widths=cat_widths,
            formatters=cat_formatters,
            sort_handler=self.report_sort_table,
            pack_opts={'fill': 'both', 'expand': True, 'padx': 15, 'pady': 5}
        )
        
        # Populate category data via a refreshable function
        def update_cat_table():
            # Clear table
            cat_tree.delete(*cat_tree.get_children())
            conn = database.connect_db()
            c = conn.cursor()
            # Aggregate by category using average integration risk across apps in the category.
            # Count applications linked via either the many-to-many table (application_categories)
            # OR the legacy applications.category_id column (union to avoid undercounting).
            c.execute('''
                WITH cat_integrations AS (
                    SELECT ic.category_id,
                           si.id AS integration_id,
                           si.parent_app_id,
                           si.risk_score
                    FROM integration_categories ic
                    JOIN system_integrations si ON si.id = ic.integration_id
                )
                SELECT c.name AS category,
                       COUNT(DISTINCT ci.integration_id) AS app_count,
                       AVG(ci.risk_score) AS avg_integration_risk
                FROM categories c
                LEFT JOIN cat_integrations ci ON ci.category_id = c.id
                GROUP BY c.id, c.name
                ORDER BY c.name ASC
            ''')
            results = c.fetchall()
            for category, count, avg_risk in results:
                if avg_risk is None or avg_risk < 1:
                    status = 'No Data'
                    raw_values = (category, count, 0.0, status)
                    cat_tree.insert('', 'end', values=cat_api['format'](raw_values))
                else:
                    try:
                        avg = float(avg_risk)
                    except Exception:
                        avg = 0.0
                    tag = get_risk_color(avg)
                    if tag == 'red':
                        status = 'High'
                    elif tag == 'yellow':
                        status = 'Med'
                    else:
                        status = 'Low'
                    raw_values = (category, count, avg, status)
                    formatted = cat_api['format'](raw_values)
                    if tag:
                        cat_tree.insert('', 'end', values=formatted, tags=(tag,))
                    else:
                        cat_tree.insert('', 'end', values=formatted)
            # Configure row colors based on risk (do this once per refresh)
            cat_tree.tag_configure('red', background='#ffcccc')
            cat_tree.tag_configure('yellow', background='#fff2cc')
            cat_tree.tag_configure('green', background='#ccffcc')
            conn.close()
            # Zebra striping
            try:
                cat_api['apply_zebra']()
            except Exception:
                pass
            # Update Category last-refresh timestamp in in-memory state and update label
            try:
                state = self._report_refresh_state.get('cat')
                if state is None:
                    self._report_refresh_state['cat'] = {'label': None, 'ts': datetime.now()}
                else:
                    state['ts'] = datetime.now()
                _update_report_label('cat')
            except Exception:
                pass
        
        # Controls row for Category with Refresh button and Export
        cat_controls = ttk.Frame(cat_frame)
        cat_controls.pack(fill='x', padx=15, pady=(0, 5))
        ttk.Button(cat_controls, text='Refresh', command=update_cat_table, style='Primary.TButton').pack(side='left')
        def export_category_report():
            try:
                rows = []
                for iid in cat_tree.get_children():
                    rows.append(cat_tree.item(iid, 'values'))  # Category, App Count, Avg Risk, Status
                if not rows:
                    messagebox.showinfo("Export", "No category data to export.")
                    return
                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                except ImportError:
                    messagebox.showerror("Missing Dependency", "openpyxl is required. Please install with pip install openpyxl")
                    return
                wb = Workbook()
                from typing import Any
                ws: Any = wb.active  # type: ignore
                ws.title = "Category Risk"
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                center = Alignment(horizontal='center')
                thin = Side(style='thin', color='888888')
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
                title = ws.cell(row=1, column=1, value=f"Category Risk Overview - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                title.font = Font(bold=True, size=14)
                title.alignment = Alignment(horizontal='center')
                headers = ["Category", "App Count", "Avg Risk", "Status"]
                for col, h in enumerate(headers, start=1):
                    ccell = ws.cell(row=3, column=col, value=h)
                    ccell.fill = header_fill
                    ccell.font = header_font
                    ccell.alignment = center
                    ccell.border = border
                row_idx = 4
                for (category, app_count, avg_risk, status) in rows:
                    data = [category, app_count, avg_risk, status]
                    for col, val in enumerate(data, start=1):
                        cell = ws.cell(row=row_idx, column=col, value=val)
                        cell.border = border
                    try:
                        rv = float(avg_risk) if avg_risk not in (None, '', '0') else 0
                    except Exception:
                        rv = 0
                    from openpyxl.styles import PatternFill
                    if rv >= 70:
                        fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    elif rv >= 40:
                        fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                    elif rv > 0:
                        fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    else:
                        fill = None
                    if fill:
                        for c in range(1, 5):
                            ws.cell(row=row_idx, column=c).fill = fill
                    row_idx += 1
                for col in range(1, 5):
                    max_len = 0
                    for r in range(1, row_idx):
                        val = ws.cell(row=r, column=col).value
                        if val is None:
                            continue
                        l = len(str(val))
                        if l > max_len:
                            max_len = l
                    ws.column_dimensions[chr(64+col)].width = min(max_len + 2, 50)
                from tkinter import filedialog
                default_name = f"category_risk_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                path = filedialog.asksaveasfilename(defaultextension='.xlsx', initialfile=default_name, filetypes=[('Excel Workbook','*.xlsx')])
                if not path:
                    return
                wb.save(path)
                _show_toast("Category export saved")
            except Exception as e:
                messagebox.showerror("Export Failed", f"Category export failed: {e}")
        ttk.Button(cat_controls, text='Export XLSX', command=export_category_report).pack(side='left', padx=5)
        cat_last_lbl = ttk.Label(cat_controls, text='Last refresh: N/A')
        cat_last_lbl.pack(side='left', padx=(8,0))
        # Register CAT label widget for elapsed updates
        try:
            st = self._report_refresh_state.get('cat')
            if st is None:
                self._report_refresh_state['cat'] = {'label': cat_last_lbl, 'ts': None}
            else:
                st['label'] = cat_last_lbl
        except Exception:
            pass

        # Initial population
        update_cat_table()
        
        # Ensure Risk Range tab stays selected by default now that System Criticality is removed
        try:
            notebook.select(risk_frame)
        except Exception:
            pass
        
        # Populate business unit data via a refreshable function
        def update_bu_table():
            # Clear table
            bu_tree.delete(*bu_tree.get_children())
            conn = database.connect_db()
            c = conn.cursor()
            # Aggregate by business unit using average integration risk (system_integrations.risk_score)
            c.execute('''SELECT bu.name AS business_unit,
                                COUNT(DISTINCT i.id) AS app_count,
                                AVG(i.risk_score) AS avg_integration_risk
                         FROM business_units bu
                         LEFT JOIN application_business_units abu ON bu.id = abu.unit_id
                         LEFT JOIN applications a ON abu.app_id = a.id
                         LEFT JOIN system_integrations i ON a.id = i.parent_app_id
                         GROUP BY bu.id, bu.name
                         ORDER BY bu.name ASC''')
            results = c.fetchall()
            for bu_name, count, avg_risk in results:
                if avg_risk is None or avg_risk < 1:
                    status = 'No Data'
                    raw_values = (bu_name, count, 0.0, status)
                    bu_tree.insert('', 'end', values=bu_api['format'](raw_values))
                else:
                    try:
                        avg = float(avg_risk)
                    except Exception:
                        avg = 0.0
                    tag = get_risk_color(avg)
                    if tag == 'red':
                        status = 'High'
                    elif tag == 'yellow':
                        status = 'Med'
                    else:
                        status = 'Low'
                    raw_values = (bu_name, count, avg, status)
                    formatted = bu_api['format'](raw_values)
                    if tag:
                        bu_tree.insert('', 'end', values=formatted, tags=(tag,))
                    else:
                        bu_tree.insert('', 'end', values=formatted)
            # Configure row colors based on risk (do this once per refresh)
            bu_tree.tag_configure('red', background='#ffcccc')
            bu_tree.tag_configure('yellow', background='#fff2cc')
            bu_tree.tag_configure('green', background='#ccffcc')
            conn.close()
            # Zebra striping
            try:
                bu_api['apply_zebra']()
            except Exception:
                pass
            # Update BU last-refresh timestamp in in-memory state and update label
            try:
                state = self._report_refresh_state.get('bu')
                if state is None:
                    self._report_refresh_state['bu'] = {'label': None, 'ts': datetime.now()}
                else:
                    state['ts'] = datetime.now()
                _update_report_label('bu')
            except Exception:
                pass

        # Controls row for BU with Refresh button
        bu_controls = ttk.Frame(bu_frame)
        bu_controls.pack(fill='x', padx=15, pady=(0, 5))
        ttk.Button(bu_controls, text='Refresh', command=update_bu_table, style='Primary.TButton').pack(side='left')
        def export_bu_report():
            try:
                rows = []
                for iid in bu_tree.get_children():
                    rows.append(bu_tree.item(iid, 'values'))  # BU, App Count, Avg Risk, Status
                if not rows:
                    messagebox.showinfo("Export", "No business unit data to export.")
                    return
                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                except ImportError:
                    messagebox.showerror("Missing Dependency", "openpyxl is required. Please install with pip install openpyxl")
                    return
                wb = Workbook()
                from typing import Any
                ws: Any = wb.active  # type: ignore
                ws.title = "Business Unit Risk"
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                center = Alignment(horizontal='center')
                thin = Side(style='thin', color='888888')
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
                title = ws.cell(row=1, column=1, value=f"Business Unit Risk Overview - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                title.font = Font(bold=True, size=14)
                title.alignment = Alignment(horizontal='center')
                headers = ["Business Unit", "App Count", "Avg Risk", "Status"]
                for col, h in enumerate(headers, start=1):
                    ccell = ws.cell(row=3, column=col, value=h)
                    ccell.fill = header_fill
                    ccell.font = header_font
                    ccell.alignment = center
                    ccell.border = border
                row_idx = 4
                for (bu_name, app_count, avg_risk, status) in rows:
                    data = [bu_name, app_count, avg_risk, status]
                    for col, val in enumerate(data, start=1):
                        cell = ws.cell(row=row_idx, column=col, value=val)
                        cell.border = border
                    try:
                        rv = float(avg_risk) if avg_risk not in (None, '', '0') else 0
                    except Exception:
                        rv = 0
                    from openpyxl.styles import PatternFill
                    if rv >= 70:
                        fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    elif rv >= 40:
                        fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                    elif rv > 0:
                        fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    else:
                        fill = None
                    if fill:
                        for c in range(1, 5):
                            ws.cell(row=row_idx, column=c).fill = fill
                    row_idx += 1
                for col in range(1, 5):
                    max_len = 0
                    for r in range(1, row_idx):
                        val = ws.cell(row=r, column=col).value
                        if val is None:
                            continue
                        l = len(str(val))
                        if l > max_len:
                            max_len = l
                    ws.column_dimensions[chr(64+col)].width = min(max_len + 2, 50)
                from tkinter import filedialog
                default_name = f"business_unit_risk_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                path = filedialog.asksaveasfilename(defaultextension='.xlsx', initialfile=default_name, filetypes=[('Excel Workbook','*.xlsx')])
                if not path:
                    return
                wb.save(path)
                _show_toast("Business Unit export saved")
            except Exception as e:
                messagebox.showerror("Export Failed", f"BU export failed: {e}")
        ttk.Button(bu_controls, text='Export XLSX', command=export_bu_report).pack(side='left', padx=5)
        bu_last_lbl = ttk.Label(bu_controls, text='Last refresh: N/A')
        bu_last_lbl.pack(side='left', padx=(8,0))
        # Register BU label widget for elapsed updates
        try:
            st = self._report_refresh_state.get('bu')
            if st is None:
                self._report_refresh_state['bu'] = {'label': bu_last_lbl, 'ts': None}
            else:
                st['label'] = bu_last_lbl
        except Exception:
            pass

        # Initial population
        update_bu_table()

        # Add styling to the report tree
        tree_style = ttk.Style()
        tree_style.configure('Treeview', rowheight=28, font=('Segoe UI', 9))
        tree_style.configure('Treeview.Heading', font=('Segoe UI', 10, 'bold'))

    # System Criticality tab removed per requirement; associated chart not created

    # Remove stray close; connections are managed within update functions

    def manage_departments_popup(self):
        popup = tk.Toplevel(self)
        popup.title('Manage Departments')
        dept_listbox = tk.Listbox(popup, selectmode='multiple', exportselection=0)
        departments = self.get_departments()
        for dept_id, dept_name in departments:
            dept_listbox.insert('end', dept_name)
        dept_listbox.pack(padx=10, pady=10)

        def delete_selected():
            selected_indices = dept_listbox.curselection()
            if not selected_indices:
                return
            # Re-fetch departments to get ids in order
            departments_current = self.get_departments()
            to_delete = [departments_current[i][0] for i in selected_indices]
            conn = database.connect_db()
            c = conn.cursor()
            for dept_id in to_delete:
                c.execute('DELETE FROM application_business_units WHERE unit_id = ?', (dept_id,))
                c.execute('DELETE FROM business_units WHERE id = ?', (dept_id,))
            conn.commit()
            conn.close()
            # Update department_listbox after deletion
            self.department_listbox.delete(0, 'end')
            for dept_id, dept_name in self.get_departments():
                self.department_listbox.insert('end', dept_name)
            popup.destroy()
            self.refresh_table()

        ttk.Button(popup, text='Delete Selected', command=delete_selected, style='Danger.TButton').pack(padx=10, pady=10)

    def on_tab_changed(self, event):
        # Get the selected tab index
        tab_id = self.tab_control.select()
        tab_index = self.tab_control.index(tab_id)
        
        # Handle tab-specific actions (can be expanded in the future)
        if tab_index == 0:  # Application Risk Assessment tab
            # Refresh data when returning to the main tab
            self.refresh_table()
        elif tab_index == 1:  # Reports tab
            # Future functionality for reports tab can be added here
            pass

    # Placeholder methods are now directly in create_widgets
            
    def delayed_search(self):
        """Perform the search after a short delay to prevent excessive refreshes while typing"""
        # Cancel any existing delayed search
        if hasattr(self, 'search_after_id') and self.search_after_id is not None:
            self.after_cancel(self.search_after_id)
            self.search_after_id = None
            
        # Only search if not showing the placeholder text
        if hasattr(self, 'search_entry') and self.search_entry is not None:
            search_text = self.search_entry.get()
            if search_text != "Type to search...":
                # Update suggestions as user types
                self.update_search_suggestions()
                # Then refresh the table with current search
                self.refresh_table()
    
    def clear_search(self):
        """Clear the search entry and refresh the table to show all results"""
        if hasattr(self, 'search_entry') and self.search_entry is not None:
            self.search_entry.set("")  # For Combobox, use set instead of delete/insert
            self.search_entry.insert(0, "Type to search...")
            self.search_entry.config(foreground='gray')
        self.refresh_table()
            
    def purge_database_gui(self):
        if messagebox.askyesno('Confirm Purge', 'Are you sure you want to delete ALL data? This cannot be undone.'):
            database.purge_database()
            # Refresh various UI elements to reflect empty database
            try:
                # Refresh department and division lists
                if hasattr(self, 'refresh_departments'):
                    try:
                        self.refresh_departments()
                    except Exception:
                        # fallback to manual refresh
                        self.department_listbox.delete(0, 'end')
                        for dept_id, dept_name in self.get_departments():
                            self.department_listbox.insert('end', dept_name)
                if hasattr(self, 'refresh_division_listbox'):
                    try:
                        self.refresh_division_listbox()
                    except Exception:
                        pass

                # Refresh category listbox
                if hasattr(self, 'populate_category_listbox'):
                    try:
                        self.populate_category_listbox()
                    except Exception:
                        # Fallback: try a manual repopulate from get_categories()
                        try:
                            if hasattr(self, 'category_listbox') and self.category_listbox is not None:
                                self.category_listbox.delete(0, 'end')
                                for cid, cname in self.get_categories():
                                    self.category_listbox.insert('end', cname)
                        except Exception:
                            pass

                # Clear selected app/category state
                try:
                    self.selected_app_id = None
                except Exception:
                    pass
                try:
                    self.selected_category_name = None
                except Exception:
                    pass

                # Refresh main table and integrations table
                try:
                    self.refresh_table()
                except Exception:
                    pass
                try:
                    self.refresh_integration_table(None)
                except Exception:
                    pass

                # Clear details area
                try:
                    if hasattr(self, 'details_text') and self.details_text is not None:
                        self.details_text.configure(state='normal')
                        self.details_text.delete('1.0', 'end')
                        self.details_text.configure(state='disabled')
                except Exception:
                    pass
            except Exception:
                pass

            messagebox.showinfo('Purge Complete', 'All data has been deleted.')

    def create_widgets(self):
        # No need to initialize filter_combo early anymore as we're using search functionality
        
        # Create the tab control
        self.tab_control = ttk.Notebook(self)
        self.tab_control.pack(fill='both', expand=True, padx=8, pady=8)
        
        # Bind tab change event
        self.tab_control.bind("<<NotebookTabChanged>>", self.on_tab_changed)
        
        # Create the first tab - Application Risk Assessment
        self.tab_risk = ttk.Frame(self.tab_control, style='TFrame')
        self.tab_control.add(self.tab_risk, text="System Risk Assessment")
        
        # Create a second tab as a placeholder
        self.tab_reports = ttk.Frame(self.tab_control, style='TFrame')
        self.tab_control.add(self.tab_reports, text="Admin")
        
        # Create a container for the top section of the Reports tab
        reports_top_frame = ttk.Frame(self.tab_reports)
        reports_top_frame.pack(fill='x', pady=10, padx=15)

        # Previously had a 'Generate Smoke Test Data' helper here; removed per request.

        # Add some placeholder content to the Reports tab
        reports_label = tk.Label(self.tab_reports, text="Admin Dashboard", 
                                font=('Segoe UI', 16, 'bold'), fg=ACCENT, bg=WIN_BG)
        reports_label.pack(pady=6)
        # Centered Purge Database button
        try:
            purge_center = ttk.Frame(self.tab_reports)
            purge_center.pack(fill='x')
            purge_btn = ttk.Button(purge_center, text='Purge Database', command=self.purge_database_gui, style='Danger.TButton')
            purge_btn.pack(pady=6)
        except Exception:
            pass

        info_label = tk.Label(self.tab_reports, text="",
                            font=('Segoe UI', 12), fg='#555555', bg=WIN_BG)
        info_label.pack(pady=10)

        # Add a small header area to the System Risk Assessment tab for actions (top-right)
        risk_header_frame = ttk.Frame(self.tab_risk)
        risk_header_frame.pack(fill='x', padx=16, pady=(12, 0))
        # Right-aligned container for action buttons
        risk_header_right = ttk.Frame(risk_header_frame)
        risk_header_right.pack(side='right')
    # Show Report button moved under the Submit Selection controls

        # use a PanedWindow to separate form and table inside the first tab
        paned = ttk.Panedwindow(self.tab_risk, orient='horizontal')
        paned.pack(fill='both', expand=True, padx=16, pady=16)

        form_frame = ttk.Frame(paned, width=280, padding=10)
        paned.add(form_frame, weight=0)

        right_frame = ttk.Frame(paned, padding=12)
        paned.add(right_frame, weight=1)

        # Define label style for modern field labels
        self.form_label_style = {'font': ('Segoe UI', 10, 'bold'), 'fg': ACCENT, 'bg': WIN_BG, 'padx': 5}
        
        # Form content with consistent padding
        padx = 5
        pady = 6
        # Define category headers style for grouping
        self.category_style = {'font': ('Segoe UI', 11, 'bold'), 'fg': '#333333', 'bg': WIN_BG, 'pady': 5}

        # --- Business Units section (moved above Rating Factors) ---
        bu_header = tk.Label(form_frame, text="Select Business Unit:", **self.category_style)
        bu_header.grid(row=0, column=0, columnspan=2, sticky='nw', padx=padx, pady=(0, 2))

        ttk.Separator(form_frame, orient='horizontal').grid(row=1, column=0, columnspan=2, sticky='ew', pady=(8, 8))

        side_btn_width = 20 # Consistent button width
        buttons_frame = ttk.Frame(form_frame)
        buttons_frame.grid(row=2, column=0, sticky='nw', padx=padx, pady=pady)
        ttk.Button(buttons_frame, text='Add Business Unit', command=self.add_department_popup, style='Primary.TButton', width=side_btn_width).pack(pady=5, anchor='w', fill='x')
        ttk.Button(buttons_frame, text='Manage Business Units', command=self.manage_departments_popup, style='Primary.TButton', width=side_btn_width).pack(pady=5, fill='x')

        dept_frame = ttk.Frame(form_frame)
        dept_frame.grid(row=2, column=1, sticky='nsew', padx=padx, pady=pady)
        dept_frame.grid_propagate(False)
        dept_frame.configure(width=250, height=120)
        self.department_listbox = tk.Listbox(dept_frame, selectmode='multiple', exportselection=0,
                                             height=6, width=30, bg='white', font=('Segoe UI', 10),
                                             highlightthickness=1, highlightcolor=ACCENT, borderwidth=1)
        self.department_listbox.pack(side='left', fill='both', expand=True)
        dept_scrollbar = ttk.Scrollbar(dept_frame, orient='vertical', command=self.department_listbox.yview)
        dept_scrollbar.pack(side='right', fill='y')
        self.department_listbox.configure(yscrollcommand=dept_scrollbar.set)

        form_frame.grid_columnconfigure(0, weight=0)
        form_frame.grid_columnconfigure(1, weight=1)
        dept_frame.grid_configure(pady=(pady, pady*2))

        # Add a spacer row to separate Business Units and Division
        # (Row indices below shift the rating form down to avoid collisions.)

        # Create a category header for Division (shifted down)
        # ratings_header = tk.Label(form_frame, text="Division", **self.category_style)
        #ratings_header.grid(row=5, column=0, sticky='w', pady=(0, 10))

        # Division section header styled like Business Unit header
        div_header = tk.Label(form_frame, text="Select Division:", **self.category_style)
        div_header.grid(row=6, column=0, columnspan=2, sticky='nw', padx=padx, pady=(10, 2))

        div_buttons_frame = ttk.Frame(form_frame)
        div_buttons_frame.grid(row=7, column=0, sticky='nw', padx=padx, pady=(0, pady))
        ttk.Button(div_buttons_frame, text='Add Division', command=self.add_division_popup, style='Primary.TButton', width=side_btn_width).pack(pady=5, anchor='w', fill='x')
        ttk.Button(div_buttons_frame, text='Manage Divisions', command=self.manage_divisions_popup, style='Primary.TButton', width=side_btn_width).pack(pady=5, fill='x')

        div_frame = ttk.Frame(form_frame)
        div_frame.grid(row=7, column=1, sticky='nsew', padx=(0, padx), pady=(0, pady))
        div_frame.grid_propagate(False)
        div_frame.configure(width=250, height=120)
        self.division_listbox = tk.Listbox(div_frame, selectmode='single', exportselection=0,
                                           height=6, width=30, bg='white', font=('Segoe UI', 10),
                                           highlightthickness=1, highlightcolor=ACCENT, borderwidth=1)
        self.division_listbox.pack(side='left', fill='both', expand=True)
        div_scrollbar = ttk.Scrollbar(div_frame, orient='vertical', command=self.division_listbox.yview)
        div_scrollbar.pack(side='right', fill='y')
        self.division_listbox.configure(yscrollcommand=div_scrollbar.set)
        # Populate initial divisions
        try:
            self.refresh_division_listbox()
        except Exception:
            pass

        # Category selection section
        cat_header = tk.Label(form_frame, text="Select Category:", **self.category_style)
        cat_header.grid(row=8, column=0, columnspan=2, sticky='nw', padx=padx, pady=(10, 2))

        cat_buttons_frame = ttk.Frame(form_frame)
        cat_buttons_frame.grid(row=9, column=0, sticky='nw', padx=padx, pady=(0, pady))
        # Store a reference so duplicate indicator can target this button.
        # Use tk.Button so we can reliably manipulate background for duplicate flash.
        self.main_add_category_button = tk.Button(
            cat_buttons_frame,
            text='Add Category',
            command=self.add_category_popup,
            bg='#0078d4', fg='white', activebackground='#006cbe', relief='flat', padx=6, pady=4
        )
        self.main_add_category_button.pack(pady=5, anchor='w', fill='x')
        ttk.Button(cat_buttons_frame, text='Manage Categories', command=self.manage_categories_popup, style='Primary.TButton', width=side_btn_width).pack(pady=5, fill='x')

        cat_frame = ttk.Frame(form_frame)
        cat_frame.grid(row=9, column=1, sticky='nsew', padx=padx, pady=(0, pady))
        cat_frame.grid_propagate(False)
        cat_frame.configure(width=250, height=120)
        # Allow multiple categories to be linked
        self.category_listbox = tk.Listbox(cat_frame, selectmode='multiple', exportselection=0,
                                           height=6, width=30, bg='white', font=('Segoe UI', 10),
                                           highlightthickness=1, highlightcolor=ACCENT, borderwidth=1)
        self.category_listbox.pack(side='left', fill='both', expand=True)
        cat_scrollbar = ttk.Scrollbar(cat_frame, orient='vertical', command=self.category_listbox.yview)
        cat_scrollbar.pack(side='right', fill='y')
        self.category_listbox.configure(yscrollcommand=cat_scrollbar.set)

        # Submit button to create an application entry from selected Business Unit(s), Division and Category(s)
        submit_frame = ttk.Frame(form_frame)
        submit_frame.grid(row=10, column=0, columnspan=2, sticky='w', padx=padx, pady=(6, 12))
        # Keep reference to the button so we can enable/disable it based on selections
        self.submit_button = ttk.Button(submit_frame, text='Submit Selection', command=self.submit_selection, style='Primary.TButton', width=side_btn_width)
        self.submit_button.pack(side='left')

        # Add Import CSV button under the submit controls (left-aligned next to Submit Selection)
        # Show Report button moved to top_buttons_frame
        # Place Import CSV in its own row under the submit controls (left-aligned)
        try:
            import_row = ttk.Frame(form_frame)
            import_row.grid(row=11, column=0, columnspan=2, sticky='w', padx=padx, pady=(0, 12))
            import_btn = ttk.Button(import_row, text='Import CSV', command=self.import_csv_dialog, style='Primary.TButton', width=side_btn_width)
            import_btn.pack(side='left')
        except Exception:
            pass

        # Bind selection changes to update the submit button state
        try:
            self.department_listbox.bind('<<ListboxSelect>>', lambda e: self.update_submit_button_state())
            self.division_listbox.bind('<<ListboxSelect>>', lambda e: self.update_submit_button_state())
            self.category_listbox.bind('<<ListboxSelect>>', lambda e: self.update_submit_button_state())
        except Exception:
            pass

        # # factor entries with modern styling (shifted row indices)
        # factor_labels = [
        #     ('Score', 7),
        #     ('Need', 8),
        #     ('Criticality', 9),
        #     ('Installed', 10),
        #     ('DisasterRecovery', 11),
        #     ('Safety', 12),
        #     ('Security', 13),
        #     ('Monetary', 14),
        #     ('CustomerService', 15)
        # ]

        # ttk.Separator(form_frame, orient='horizontal').grid(row=5, column=0, columnspan=2, sticky='ew', pady=(30, 5))

        # self.factor_entries = {}
        # for label, row in factor_labels:
        #     factor_label = tk.Label(form_frame, text=label, **self.form_label_style)
        #     factor_label.grid(row=row, column=0, sticky='e', padx=(padx, 2), pady=pady)
        #     entry = ttk.Entry(form_frame, width=10)
        #     entry.grid(row=row, column=1, sticky='w', padx=(0, padx), pady=pady)
        #     self.factor_entries[label] = entry

        # vendor_label = tk.Label(form_frame, text='Related Vendor', **self.form_label_style)
        # vendor_label.grid(row=16, column=0, sticky='e', padx=(padx, 2), pady=pady)
        # self.vendor_entry = ttk.Entry(form_frame)
        # self.vendor_entry.grid(row=16, column=1, sticky='ew', padx=(0, padx), pady=pady)

    # Note: the Division Add button is provided in the Division buttons frame above
    # (div_buttons_frame) which correctly calls `self.add_division_popup`.
    # Removed duplicate/miswired Add Division button that called `self.add_application`.

        # Create a frame for top-right buttons
        top_buttons_frame = ttk.Frame(right_frame)
        top_buttons_frame.pack(side='top', fill='x', pady=6)
        
        # Edit Selected and Save Changes buttons at the top with Primary style
        ttk.Button(top_buttons_frame, text='Edit Selected', command=self.load_selected_for_edit, style='Primary.TButton').pack(side='left', padx=6)
        # Save Changes and Show Report buttons (Save Changes first)
        ttk.Button(top_buttons_frame, text='Save Changes', command=self.save_edit, style='Primary.TButton').pack(side='left', padx=6)
        try:
            top_show = ttk.Button(top_buttons_frame, text='Show Reports', command=self.show_report, style='Primary.TButton')
            top_show.pack(side='left', padx=6)
        except Exception:
            pass
        
    # 'Purge Database' button moved to Admin tab (centered under Admin Dashboard)
        # 'Generate Smoke Test Data' button removed.
        # Import CSV button (single-file import matching smoke-test format)
        # Import CSV button moved under Submit Selection controls
        # 'Diagnose CSV' button removed.
        
        # Create Note and Save Note buttons in a separate frame below
        edit_frame = ttk.Frame(right_frame)
        edit_frame.pack(fill='x', pady=(0,6))
        ttk.Button(edit_frame, text='Create Note', command=self.create_note, style='Success.TButton').pack(side='left', padx=6)
        ttk.Button(edit_frame, text='Save Note', command=self.save_notes, style='Success.TButton').pack(side='left', padx=6)
        # Add 'Delete Selected' button 
        ttk.Button(edit_frame, text='Delete Selected', command=self.delete_selected_app, style='Danger.TButton').pack(side='left', padx=6)

        # Right frame: search + filter + table with modern styling
        control_frame = ttk.Frame(right_frame)
        control_frame.pack(fill='x', pady=(0,6))
        
        # Create an improved search frame with better layout
        search_frame = ttk.Frame(control_frame)
        search_frame.pack(side='left', fill='x', expand=True)
        
        filter_label_style = {'font': ('Segoe UI', 10), 'fg': ACCENT, 'bg': WIN_BG}
        
        # Create search bar with icon-like prefix
        search_container = ttk.Frame(search_frame)
        search_container.pack(side='top', fill='x', pady=2)
        
        search_icon = tk.Label(search_container, text="🔍", font=('Segoe UI', 12), bg=WIN_BG)
        search_icon.pack(side='left', padx=(0,2))
        
        search_label = tk.Label(search_container, text='Quick Search:', **filter_label_style)
        search_label.pack(side='left', padx=(0,6))
        
        # Create search as a combobox with autocomplete
        self.search_entry = ttk.Combobox(search_container, width=20)
        self.search_entry.pack(side='left', padx=(0,6))
        self.search_entry.insert(0, "Type to search...")
        self.search_entry.config(foreground='gray')
        # Set up focus events for placeholder behavior for combobox
        def on_combobox_click(event):
            if self.search_entry is not None and self.search_entry.get() == "Type to search...":
                self.search_entry.delete(0, "end")
                self.search_entry.config(foreground='black')
            # When the combobox is clicked, update the suggestions
            self.update_search_suggestions()
                
        def on_combobox_leave(event):
            if self.search_entry is not None and self.search_entry.get() == "":
                self.search_entry.set("")  # Clear first
                self.search_entry.insert(0, "Type to search...")
                self.search_entry.config(foreground='gray')
                
        self.search_entry.bind("<FocusIn>", on_combobox_click)
        self.search_entry.bind("<FocusOut>", on_combobox_leave)
        
        # Bind selection event to trigger search immediately
        self.search_entry.bind("<<ComboboxSelected>>", lambda event: self.refresh_table())
        
        # Bind key release event to trigger search with small delay
        self.search_entry.bind("<KeyRelease>", lambda event: self.after(300, self.delayed_search))
        
        # Initialize the suggestions list
        self.update_search_suggestions()
        
        # Create search options in a separate row for clarity - radio buttons first
        options_frame = ttk.Frame(search_frame)
        options_frame.pack(side='top', fill='x', pady=2)

        # Add radio buttons for search type
        self.search_type_var = tk.StringVar(value="Division")
        ttk.Radiobutton(options_frame, text="Division", variable=self.search_type_var,
                        value="Division", command=self.update_search_selections).pack(side='left', padx=(5,5))
        ttk.Radiobutton(options_frame, text="Business Unit", variable=self.search_type_var,
                        value="Business Unit", command=self.update_search_selections).pack(side='left', padx=5)
        
        # Move search container below the radio buttons
        search_container.pack_forget()  # Remove the current packing of the search container
        search_container.pack(side='top', anchor='w', pady=(5, 0), padx=(20, 0))  # Position below and indented
        
        # Add search button and clear button in a nicer layout
        button_frame = ttk.Frame(search_container)
        button_frame.pack(side='right', padx=(6,0))
        
        ttk.Button(search_container, text='Clear', command=self.clear_search, 
                 style='Secondary.TButton').pack(side='right', padx=(6,0))

        # Create main tables container
        tables_container = ttk.Frame(right_frame)
        tables_container.pack(fill='both', expand=True)
        
        # Main systems table area (top half)
        table_frame = ttk.Frame(tables_container)
        table_frame.pack(fill='both', expand=True, side='top')
        
        # Header row for main table: title on left, color legend on right
        header_frame = ttk.Frame(table_frame)
        header_frame.grid(row=0, column=0, columnspan=2, sticky='ew')
        header_frame.columnconfigure(0, weight=1)
        systems_title = ttk.Label(header_frame, text="Business Unit / Division", font=('Segoe UI', 11, 'bold'))
        systems_title.grid(row=0, column=0, sticky='w', padx=5, pady=(0, 5))
        # Add a small legend explaining color bands
        legend_frame = ttk.Frame(header_frame)
        legend_frame.grid(row=0, column=1, sticky='e', padx=5, pady=(0, 5))
        def _legend_chip(parent, text, bg):
            lbl = tk.Label(parent, text=text, bg=bg, fg='black', padx=6, pady=2)
            lbl.pack(side='left', padx=3)
        _legend_chip(legend_frame, 'High (≥70)', '#ffcccc')
        _legend_chip(legend_frame, 'Med (40–69)', '#fff2cc')
        _legend_chip(legend_frame, 'Low (1–39)', '#ccffcc')
        
        # Reduce visible columns to only what's required by the user: Business Unit, Division, Category, Last Modified
        self.tree = ttk.Treeview(
            table_frame,
            columns=('Business Unit', 'Division', 'Category', 'Last Modified'),
            show='headings'
        )
        # tuned widths and anchors to align like the provided screenshot
        cols = list(self.tree['columns'])
        base_widths = {'Business Unit': 380, 'Division': 260, 'Category': 250, 'Last Modified': 200}
        for col in cols:
            w = base_widths.get(col, 120)
            if col == 'Business Unit':
                anchor = 'w'
            elif col == 'Division':
                anchor = 'w'
            elif col == 'Category':
                anchor = 'w'
            elif col == 'Last Modified':
                anchor = 'w'
            else:
                anchor = 'e'
            heading_text = col
            # Ensure heading text aligns with data cells
            self.tree.heading(col, text=heading_text, anchor='w', command=lambda c=col: self.sort_table(c, False))
            # Disable stretch to keep columns at the specified widths and closer together
            self.tree.column(col, width=w, anchor=anchor, stretch=False, minwidth=60)

        # vertical scrollbar for main systems table
        vsb = ttk.Scrollbar(table_frame, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.grid(row=1, column=0, sticky='nsew')
        vsb.grid(row=1, column=1, sticky='ns')
        table_frame.rowconfigure(1, weight=1)
        # reserve a grid row for the details area so it aligns with the table columns
        table_frame.rowconfigure(2, weight=0)
        table_frame.columnconfigure(0, weight=1)
        
        # Sub-systems/Integrations table area (bottom half)
        integrations_frame = ttk.Frame(tables_container)
        integrations_frame.pack(fill='both', expand=True, side='bottom', pady=(10, 0))
        
        # Create a frame for integration controls
        integration_controls = ttk.Frame(integrations_frame)
        integration_controls.grid(row=0, column=0, sticky='ew', padx=5, pady=(0, 5))
        
        # Integrations title and add button
        self.integrations_title = ttk.Label(integration_controls, text="System Sub/Integrations", font=('Segoe UI', 11, 'bold'))
        self.integrations_title.pack(side='left')
        
        # Add and Delete integration buttons
        delete_integration_btn = ttk.Button(integration_controls, text="Delete Integration", 
                                         command=self.delete_selected_integration, style='Danger.TButton')
        delete_integration_btn.pack(side='right', padx=5)
        
        add_integration_btn = ttk.Button(integration_controls, text="Add Integration", 
                                      command=self.add_system_integration, style='Primary.TButton')
        add_integration_btn.pack(side='right', padx=5)
        
        # Create the integrations table with same columns but adjusted for integrations
        self.integrations_tree = ttk.Treeview(
            integrations_frame,
            columns=(
                'Name', 'Vendor', 'Risk', 'Score', 'Criticality', 'Need', 'Installed',
                    'DR', 'Safety', 'Security', 'Monetary', 'CustomerService',
                'Last Modified'
            ),
            show='headings'
        )
        
        # Apply column settings to integrations table
        int_cols = list(self.integrations_tree['columns'])
        col_widths = {
            'Name': 220, 'Vendor': 160, 'Risk': 80, 'Score': 80, 'Criticality': 90, 'Need': 80,
            'Installed': 80, 'DR': 80, 'Safety': 80, 'Security': 80,
            'Monetary': 80, 'CustomerService': 140, 'Last Modified': 150
        }
        
        for col in int_cols:
            w = col_widths.get(col, 100)
            # Left-align text columns, center-align numeric columns
            anchor = 'w' if col in ('Name', 'Vendor') else 'center'
            # Use consistent column names
            heading_text = (
                'Cust Service' if col == 'CustomerService' else (
                    'System/Integration' if col == 'Name' else col
                )
            )
            # Align integration headings with left-anchored values where applicable
            self.integrations_tree.heading(col, text=heading_text, anchor='w',
                                        command=lambda c=col: self.sort_integration_table(c, False))
            self.integrations_tree.column(col, width=w, anchor=anchor, stretch=True, minwidth=w)
        
        # Vertical scrollbar for integrations table
        int_vsb = ttk.Scrollbar(integrations_frame, orient='vertical', command=self.integrations_tree.yview)
        # Horizontal scrollbar for integrations table
        int_hsb = ttk.Scrollbar(integrations_frame, orient='horizontal', command=self.integrations_tree.xview)
        self.integrations_tree.configure(yscrollcommand=int_vsb.set, xscrollcommand=int_hsb.set)
        self.integrations_tree.grid(row=1, column=0, sticky='nsew')
        int_vsb.grid(row=1, column=1, sticky='ns')
        # Place horizontal scrollbar below the tree, spanning the table column
        int_hsb.grid(row=2, column=0, sticky='ew')
        integrations_frame.rowconfigure(1, weight=1)
        integrations_frame.columnconfigure(0, weight=1)
        
        # Make both tables the same height
        tables_container.update_idletasks()
        
        # Details area below the tables: place inside table_frame so left/right edges align
        details_frame = ttk.Frame(table_frame)
        # Span both columns (tree + its vertical scrollbar) so left/right align exactly
        details_frame.grid(row=2, column=0, columnspan=2, sticky='ew', pady=(6, 0))
        details_frame.columnconfigure(0, weight=1)

        # Remove fixed width so the Text widget expands to container width
        details_text = tk.Text(details_frame, wrap='word', height=6)
        details_vsb = ttk.Scrollbar(details_frame, orient='vertical', command=details_text.yview)
        details_text.configure(yscrollcommand=details_vsb.set, state='disabled')
        # Use grid to ensure exact alignment with table_frame columns
        details_text.grid(row=0, column=0, sticky='ew')
        details_vsb.grid(row=0, column=1, sticky='ns')
        # expose as attribute so handlers can update it
        self.details_text = details_text
        # Configure text tags to highlight risk value
        try:
            self.details_text.tag_configure('risk_red', foreground='#b00020', font=('Segoe UI', 10, 'bold'))
            self.details_text.tag_configure('risk_yellow', foreground='#8a6d3b', font=('Segoe UI', 10, 'bold'))
            self.details_text.tag_configure('risk_green', foreground='#2e7d32', font=('Segoe UI', 10, 'bold'))
            self.details_text.tag_configure('risk_na', foreground='#6c757d', font=('Segoe UI', 10, 'italic'))
        except Exception:
            pass

        # Store the current parent system ID for integration operations
        self.current_parent_system_id = None

        # populate details area when a row is selected
        self.tree.bind('<<TreeviewSelect>>', self.on_tree_select)
        self.integrations_tree.bind('<<TreeviewSelect>>', self.on_integration_select)
        
        # Double click on integration to edit it
        self.integrations_tree.bind('<Double-1>', self.on_integration_double_click)
        


    def load_selected_for_edit(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo('Edit', 'Please select a row to edit.')
            return
        item = sel[0]
        try:
            # Clear all current selections
            self.department_listbox.selection_clear(0, tk.END)

            # Get the departments for this application
            # Support composite iid format "appId:category"
            try:
                app_id = int(item.split(':', 1)[0])
            except ValueError:
                raise ValueError(f"Invalid application id format: {item}")
            app_departments = database.get_app_departments(app_id)

            # Select departments in the listbox
            for i in range(self.department_listbox.size()):
                dept_name = self.department_listbox.get(i)
                if dept_name in app_departments:
                    self.department_listbox.selection_set(i)

            # Load application record from DB and populate division listbox selection
            try:
                app_row = database.get_application(app_id)
                division_name = app_row['name'] if hasattr(app_row, 'keys') and 'name' in app_row.keys() else app_row[1]
                # Ensure list is populated
                self.refresh_division_listbox()
                # Select matching division
                for i in range(self.division_listbox.size()):
                    if self.division_listbox.get(i) == division_name:
                        self.division_listbox.selection_clear(0, tk.END)
                        self.division_listbox.selection_set(i)
                        self.division_listbox.see(i)
                        break
            except Exception:
                pass
            # Preselect categories for this app
            try:
                self.category_listbox.selection_clear(0, tk.END)
                cat_names = set(database.get_app_categories(app_id))
                cats = self.get_categories()
                for i, (_, cname) in enumerate(cats):
                    if cname in cat_names:
                        self.category_listbox.selection_set(i)
            except Exception:
                pass
        except Exception as e:
            messagebox.showerror('Error', f'Failed to load selected row for editing: {e}')

    def save_edit(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo('Save', 'Please select a row to save.')
            return
        item = sel[0]
        # get app id from iid
        try:
            app_id = int(item.split(':', 1)[0])
        except Exception:
            messagebox.showerror('Save', 'Could not determine application id for selected row.')
            return
        # Determine selected division
        name = ''
        try:
            sel_idx = self.division_listbox.curselection()
            if sel_idx:
                name = self.division_listbox.get(sel_idx[0]).strip()
        except Exception:
            pass
        # vendor = self.vendor_entry.get().strip()
        fields = {}
        # try:
        #     for key, entry in self.factor_entries.items():
        #         fields[key.lower()] = int(entry.get())
        # except ValueError:
        #     messagebox.showerror('Error', 'All factor fields must be integers.')
        #     return
        fields['name'] = name
        # If categories are selected, set many-to-many links after update
        # fields['vendor'] = vendor
        # update db
        database.update_application(app_id, fields)
        # Update many-to-many category links
        try:
            sel = self.category_listbox.curselection()
            ids = []
            if sel:
                cats = self.get_categories()
                for idx in sel:
                    if idx < len(cats):
                        ids.append(cats[idx][0])
            database.set_app_categories(app_id, ids)
        except Exception:
            pass
        self.refresh_table()
        
        # Clear form fields after saving
        # Clear division selection (no direct text entry now)
        try:
            self.division_listbox.selection_clear(0, tk.END)
        except Exception:
            pass
        # self.vendor_entry.delete(0, 'end')
        # for entry in self.factor_entries.values():
        #     entry.delete(0, 'end')
        self.department_listbox.selection_clear(0, tk.END)
        
        messagebox.showinfo('Saved', 'Changes saved.')

    def add_application(self):
        # Placeholder: division now selected via listbox
        try:
            sel_idx = self.division_listbox.curselection()
            name = self.division_listbox.get(sel_idx[0]) if sel_idx else ''
        except Exception:
            name = ''
        vendor = ""  # Default to empty string if vendor entry does not exist
        # if hasattr(self, 'vendor_entry'):
        #     try:
        #         vendor = self.vendor_entry.get()
        #     except Exception:
        #         vendor = ""
        # Build a safe factors dict. The form may not expose rating entries (they were commented out),
        # so default missing ratings to 0.
        factors = {}
        rating_keys = ['Score', 'Need', 'Criticality', 'Installed', 'DisasterRecovery', 'Safety', 'Security', 'Monetary', 'CustomerService']
        fe = getattr(self, 'factor_entries', {}) or {}
        for key in rating_keys:
            try:
                if key in fe and fe[key] is not None:
                    val = fe[key].get().strip()
                    factors[key] = int(val) if val != '' else 0
                else:
                    factors[key] = 0
            except Exception:
                factors[key] = 0

        selected_indices = self.department_listbox.curselection()
        if not selected_indices:
            messagebox.showerror('Error', 'Please select at least one department.')
            return

        departments = self.get_departments()
        dept_ids = [departments[i][0] for i in selected_indices]
        last_mod = datetime.utcnow().isoformat()

        # Use database helper to create application rows and links
        try:
            app_ids = database.add_application(name, vendor, factors, dept_ids, notes='')
            # Update category for the newly added app id (single)
            try:
                if app_ids:
                    app_id = app_ids
                    sel = self.category_listbox.curselection()
                    ids = []
                    if sel:
                        cats = self.get_categories()
                        for idx in sel:
                            if idx < len(cats):
                                ids.append(cats[idx][0])
                    database.set_app_categories(app_id, ids)
            except Exception:
                pass
        except Exception as e:
            messagebox.showerror('Error', f'Failed to add application: {e}')
            return

        if not app_ids:
            messagebox.showerror('Error', 'Failed to create application entries.')

        try:
            self.division_listbox.selection_clear(0, tk.END)
        except Exception:
            pass
        # self.vendor_entry.delete(0, tk.END)
        # for entry in self.factor_entries.values():
        #     entry.delete(0, tk.END)
        self.department_listbox.selection_clear(0, tk.END)
        try:
            self.category_listbox.selection_clear(0, tk.END)
        except Exception:
            pass
        self.refresh_table()

    def submit_selection(self):
        """Create or update application from selected Business Unit(s), Division and Category(s)."""
        # Get selected departments (business units)
        dept_indices = self.department_listbox.curselection()
        if not dept_indices:
            messagebox.showerror('Error', 'Please select at least one Business Unit.')
            return
        departments = self.get_departments()
        dept_ids = [departments[i][0] for i in dept_indices]

        # Get selected division (single select)
        try:
            sel_div = self.division_listbox.curselection()
            if not sel_div:
                messagebox.showerror('Error', 'Please select a Division.')
                return
            division_name = self.division_listbox.get(sel_div[0])
        except Exception:
            messagebox.showerror('Error', 'Please select a Division.')
            return

        # Get selected categories (multiple)
        cat_sel = self.category_listbox.curselection()
        if not cat_sel:
            messagebox.showerror('Error', 'Please select at least one Category.')
            return
        cats = self.get_categories()
        cat_ids = [cats[i][0] for i in cat_sel if i < len(cats)]

        # Check if division already exists
        existing_id, has_links = database.get_division_application(division_name)
        action_desc = "Update" if existing_id and has_links else "Create"
        
        # Confirm with the user with appropriate message
        message = f'{action_desc} application for Division "{division_name}" linked to {len(dept_ids)} Business Unit(s) and {len(cat_ids)} Category(s)?'
        if existing_id and has_links:
            message += "\n\nThis division already exists and has links. The operation will update its relationships."
        
        confirm = messagebox.askyesno('Confirm ' + action_desc, message)
        if not confirm:
            return

        try:
            if existing_id:
                # Update existing application's relationships
                database.link_app_to_departments(existing_id, dept_ids)
                database.set_app_categories(existing_id, cat_ids)
                database.touch_application(existing_id)
                app_id = existing_id
            else:
                # Create new application
                app_id = database.add_application(division_name, '', {}, dept_ids, notes='')
                if app_id:
                    database.set_app_categories(app_id, cat_ids)
                    database.touch_application(app_id)

            if app_id:
                try:
                    if hasattr(self, 'submit_button') and self.submit_button is not None:
                        self.submit_button.configure(style='Success.TButton')
                        self.submit_button.state(['!disabled'])
                except Exception:
                    pass
                messagebox.showinfo('Success', f'{action_desc}d application for Division "{division_name}"')
            else:
                messagebox.showerror('Error', 'Failed to process application entry.')
        except Exception as e:
            messagebox.showerror('Error', f'Failed to {action_desc.lower()} application: {e}')
            return

        # Clear selections and refresh
        try:
            self.department_listbox.selection_clear(0, tk.END)
        except Exception:
            pass
        try:
            self.division_listbox.selection_clear(0, tk.END)
        except Exception:
            pass
        try:
            self.category_listbox.selection_clear(0, tk.END)
        except Exception:
            pass
        self.refresh_table()

        # Highlight the newly created application's row(s) (one per category) after table refresh
        try:
            if app_id:
                # Find all iids matching the new app id prefix
                new_iids = [iid for iid in self.tree.get_children('') if str(iid).startswith(f"{app_id}:")]
                if new_iids:
                    # Define highlight tag
                    try:
                        self.tree.tag_configure('newflash', background='#cfe8ff')  # light blue
                    except Exception:
                        pass
                    for iid in new_iids:
                        try:
                            current_tags = self.tree.item(iid, 'tags') or ()
                            if 'newflash' not in current_tags:
                                self.tree.item(iid, tags=current_tags + ('newflash',))
                        except Exception:
                            pass
                    # Select first new row
                    try:
                        self.tree.selection_set(new_iids[0])
                        self.tree.see(new_iids[0])
                    except Exception:
                        pass

                    # Fade out effect: gradually blend back to transparent by alternating shade then removing tag
                    fade_steps = ['#cfe8ff', '#d9edff', '#e3f2ff', '#ecf6ff', '#f5faff']
                    def fade_step(i=0):
                        try:
                            if i < len(fade_steps):
                                shade = fade_steps[i]
                                try:
                                    self.tree.tag_configure('newflash', background=shade)
                                except Exception:
                                    pass
                                self.after(120, lambda: fade_step(i+1))
                            else:
                                # Remove highlight by clearing tag background and removing tag from rows
                                try:
                                    self.tree.tag_configure('newflash', background='')
                                except Exception:
                                    pass
                                for iid in new_iids:
                                    try:
                                        tags_now = tuple(t for t in self.tree.item(iid, 'tags') if t != 'newflash')
                                        self.tree.item(iid, tags=tags_now)
                                    except Exception:
                                        pass
                        except Exception:
                            pass
                    # Kick off fade after slight delay so initial color is visible
                    self.after(150, fade_step)
        except Exception:
            pass

        # Start fading the button from green to blue (ACCENT) over ~1s, then disable
        try:
            if hasattr(self, 'submit_button') and self.submit_button is not None:
                self._submit_fading = True

                # Define start (green) and end (blue) colors
                start_rgb = (144, 238, 144)  # #90EE90
                # Parse ACCENT hex (#0078d4 typical) if defined
                try:
                    end_hex = ACCENT if ACCENT.startswith('#') and len(ACCENT) == 7 else '#0078d4'
                except Exception:
                    end_hex = '#0078d4'
                end_rgb = (int(end_hex[1:3], 16), int(end_hex[3:5], 16), int(end_hex[5:7], 16))

                steps = 20
                interval = 50  # ms (20 * 50 = 1000ms)

                def lerp(a, b, t):
                    return int(a + (b - a) * t)

                # Use a single dynamic style for flash instead of direct background configure
                dyn_style_name = 'SubmitFlashDynamic.TButton'
                try:
                    ttk.Style(self.submit_button).configure(dyn_style_name, background='#90EE90', foreground='white', borderwidth=0)
                except Exception:
                    pass

                def step(i=0):
                    try:
                        if not hasattr(self, 'submit_button') or self.submit_button is None:
                            return
                        t = i / float(steps)
                        r = lerp(start_rgb[0], end_rgb[0], t)
                        g = lerp(start_rgb[1], end_rgb[1], t)
                        b = lerp(start_rgb[2], end_rgb[2], t)
                        hex_color = f"#{r:02x}{g:02x}{b:02x}"
                        try:
                            style = ttk.Style(self.submit_button)
                            style.configure(dyn_style_name, background=hex_color)
                            self.submit_button.configure(style=dyn_style_name)
                        except Exception:
                            pass
                        if i < steps:
                            self.after(interval, lambda: step(i + 1))
                        else:
                            # Finalize: apply primary style & disable
                            try:
                                self.submit_button.configure(style='Primary.TButton')
                            except Exception:
                                pass
                            try:
                                self.submit_button.state(['disabled'])
                            except Exception:
                                pass
                            self._submit_fading = False
                    except Exception:
                        pass

                # Kick off fade after UI idle so messagebox close doesn't block initial frame
                self.after(50, step)
        except Exception:
            pass

    def update_submit_button_state(self):
        """Enable the submit button only when a dept, division and category are selected."""
        try:
            has_dept = bool(self.department_listbox.curselection())
            has_div = bool(self.division_listbox.curselection())
            has_cat = bool(self.category_listbox.curselection())
            if hasattr(self, 'submit_button') and self.submit_button is not None:
                if getattr(self, '_submit_fading', False):
                    # During fade, ignore attempts to restyle/enable
                    return
                if has_dept and has_div and has_cat:
                    # Enable and switch to green success style
                    try:
                        self.submit_button.state(['!disabled'])
                        self.submit_button.configure(style='Success.TButton')
                    except Exception:
                        # Fallback: at least enable the button
                        try:
                            self.submit_button.state(['!disabled'])
                        except Exception:
                            pass
                else:
                    # Disable and revert to primary style
                    try:
                        self.submit_button.state(['disabled'])
                        self.submit_button.configure(style='Primary.TButton')
                    except Exception:
                        try:
                            self.submit_button.state(['disabled'])
                        except Exception:
                            pass
        except Exception:
            pass

    def update_search_selections(self):
        """Handle search type change and update dropdown values"""
        self.update_search_suggestions()
        self.refresh_table()
        
    def update_search_suggestions(self, event=None):
        """Update the dropdown suggestions based on the search type selected"""
        if not hasattr(self, 'search_entry') or self.search_entry is None:
            return
            
        search_type = "Division"
        if hasattr(self, 'search_type_var') and self.search_type_var is not None:
            search_type = self.search_type_var.get()
            
        # Get appropriate values from the database
        conn = database.sqlite3.connect(database.DB_NAME)
        c = conn.cursor()
        
        if search_type == "Division":
            c.execute("SELECT DISTINCT division FROM applications ORDER BY division")
            values = [row[0] for row in c.fetchall()]
        else:  # Business Unit
            c.execute("SELECT DISTINCT name FROM business_units ORDER BY name")
            values = [row[0] for row in c.fetchall()]
            
        conn.close()
        
        # Update the combobox values
        self.search_entry['values'] = values
    
    def refresh_table(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        # Use database helper to get a connection with the right row factory
        conn = database.connect_db()
        c = conn.cursor()
        
        # Get search parameters if available
        search_text = ""
        search_type = "Division"
        if hasattr(self, 'search_entry') and self.search_entry is not None:
            entered_text = self.search_entry.get().strip().lower()
            # Don't search if the placeholder text is showing
            if entered_text != "type to search...":
                search_text = entered_text
        if hasattr(self, 'search_type_var') and self.search_type_var is not None:
            search_type = self.search_type_var.get()
            
        # Pre-compute average integration risk per application
        avg_map = {}
        try:
            c.execute('SELECT parent_app_id, AVG(risk_score) FROM system_integrations GROUP BY parent_app_id')
            for app_id_val, avg_val in c.fetchall():
                try:
                    avg_map[int(app_id_val)] = float(avg_val) if avg_val is not None else None
                except Exception:
                    avg_map[int(app_id_val)] = None
        except Exception:
            avg_map = {}

        c.execute('''SELECT id, division AS name, vendor, score, need, criticality, installed, disaster_recovery, safety, security, monetary, customer_service, notes, risk_score, last_modified, category_id FROM applications''')
        rows = []  # will hold tuples of ((dept, division, category, last_mod), app_id, risk_score, category_name)
        # cache categories per app to avoid extra connections per category
        app_categories_cache = {}
        for app_row in c.fetchall():
            # Determine app id reliably
            try:
                app_id = int(app_row['id']) if hasattr(app_row, 'keys') else int(app_row[0])
            except Exception:
                app_id = int(app_row[0])

            # Get business unit list for this app
            depts = database.get_app_departments(app_id)
            # We'll display the first Business Unit (or comma-join if multiple)
            dept_str = ', '.join(depts)

            # Compute color based on average integration risk for this app
            avg_integration_risk = avg_map.get(app_id)
            risk_score = avg_integration_risk if avg_integration_risk is not None else 0

            # Initialize default values to prevent unbound variable errors
            category_display = ''
            last_mod = ''
            division = ''

            # Division: use 'name' as the Division column (previously System column)
            try:
                division = app_row['name'] if hasattr(app_row, 'keys') and 'name' in app_row.keys() else app_row[1]
            except Exception:
                division = app_row[1]

            # Categories & last modified: single query using shared connection
            try:
                if app_id not in app_categories_cache:
                    c.execute('''
                        SELECT DISTINCT c.name
                        FROM categories c
                        JOIN application_categories ac ON c.id = ac.category_id
                        WHERE ac.app_id = ?
                        ORDER BY c.name
                    ''', (app_id,))
                    app_categories_cache[app_id] = [r[0] for r in c.fetchall()]
                category_list = app_categories_cache.get(app_id, [])
                # Get max last_modified from integrations for this app
                c.execute('SELECT MAX(last_modified) FROM system_integrations WHERE parent_app_id = ?', (app_id,))
                lm_row = c.fetchone()
                if lm_row and lm_row[0]:
                    last_mod = str(lm_row[0]).split('.')[0]
                    if 'T' in last_mod:
                        last_mod = last_mod.replace('T', ' ')
                else:
                    last_mod = ''
            except Exception as e:
                print(f"DEBUG: Error retrieving categories for app {app_id}: {e}")
                category_list = []
                last_mod = ''

            # If no categories, still show a row with blank category so app not hidden
            if not category_list:
                category_list = ['']

            for cat_name in category_list:
                row = (dept_str, division, cat_name, last_mod)
                # Apply search filtering per row
                if search_text:
                    if search_type == "Division" and search_text.lower() not in (division or '').lower():
                        continue
                    elif search_type == "Business Unit" and not any(search_text.lower() in dept.lower() for dept in depts):
                        continue
                rows.append((row, app_id, risk_score, cat_name))
        # Sort rows by the 'Business Unit' column (index 0)
        rows.sort(key=lambda x: ((x[0][0] or '').lower(), (x[0][2] or '').lower()))
        for row, app_id, risk_score, cat_name in rows:
            color = get_risk_color(risk_score)
            # Composite iid: appId:category (category may be empty string)
            composite_iid = f"{app_id}:{cat_name}" if cat_name is not None else f"{app_id}:"
            if color:
                self.tree.insert('', 'end', iid=composite_iid, values=row, tags=(color,))
            else:
                self.tree.insert('', 'end', iid=composite_iid, values=row)
        self.tree.tag_configure('red', background='#ffcccc')
        self.tree.tag_configure('yellow', background='#fff2cc')
        self.tree.tag_configure('green', background='#ccffcc')
        conn.close()

    def on_tree_select(self, event):
        # called when a row is selected
        try:
            sel = self.tree.selection()
            if not sel:
                # clear details and integrations
                self.details_text.configure(state='normal')
                self.details_text.delete('1.0', 'end')
                self.details_text.configure(state='disabled')
                self.current_parent_system_id = None
                self.selected_app_id = None
                self.selected_category_name = None
                self.refresh_integration_table()
                return
                
            item = sel[0]
            # Composite iid format: appId:category
            try:
                if ':' in item:
                    app_part, category_part = item.split(':', 1)
                else:
                    app_part, category_part = item, ''
                self.selected_app_id = int(app_part)
                self.selected_category_name = category_part if category_part else None
                # Set current parent system id for integrations
                self.current_parent_system_id = self.selected_app_id
                
                # Update integrations title; fetch authoritative values from DB
                app_row = database.get_application(self.selected_app_id)
                try:
                    system_name = app_row['name'] if hasattr(app_row, 'keys') and 'name' in app_row.keys() else app_row[1]
                except Exception:
                    system_name = ''
                # Business unit(s)
                depts = database.get_app_departments(self.selected_app_id)
                business_unit = ', '.join(depts)
                # Build dynamic title parts
                title_parts = ["System Sub/Integrations"]
                if system_name:
                    title_parts.append(system_name)
                if business_unit:
                    title_parts.append(business_unit)
                # Append category (distinct from business unit) if selected
                if self.selected_category_name:
                    title_parts.append(self.selected_category_name)
                self.integrations_title.configure(text=" - ".join(title_parts))
            except Exception:
                self.selected_app_id = None
                self.current_parent_system_id = None
                self.integrations_title.configure(text="System Sub/Integrations")
            
            # Compute averages
            avg_val = None                # system average (from displayed integrations)
            division_avg_val = None       # division average
            division_name = None
            try:
                # Determine division name
                app_row_for_div = database.get_application(self.selected_app_id)
                try:
                    division_name = app_row_for_div['name'] if hasattr(app_row_for_div, 'keys') and 'name' in app_row_for_div.keys() else app_row_for_div[1]
                except Exception:
                    division_name = None

                # Division average (DB query)
                if division_name:
                    try:
                        conn = database.connect_db()
                        cur = conn.cursor()
                        cur.execute('''SELECT AVG(si.risk_score)
                                       FROM system_integrations si
                                       JOIN applications a ON si.parent_app_id = a.id
                                       WHERE a.division = ?''', (division_name,))
                        row_div = cur.fetchone()
                        if row_div is not None:
                            division_avg_val = float(row_div[0]) if row_div[0] is not None else None
                        conn.close()
                    except Exception:
                        division_avg_val = None

                # System average derived from currently displayed integrations (ensures it matches table)
                try:
                    # Ensure table is refreshed first so values are current
                    self.refresh_integration_table()
                    risk_values = []
                    # Determine the index of the 'Risk' column dynamically so this works
                    # even if columns were reordered. Use the Treeview's configured
                    # columns; if not available, fall back to scanning header values.
                    risk_col_index = None
                    try:
                        cols = list(self.integrations_tree['columns'])
                        if 'Risk' in cols:
                            risk_col_index = cols.index('Risk')
                    except Exception:
                        cols = []

                    for iid in self.integrations_tree.get_children():
                        vals = self.integrations_tree.item(iid, 'values')
                        # If we found a 'Risk' column, use its index; otherwise
                        # attempt to find a value that looks like 'number (Band)'.
                        if risk_col_index is not None:
                            if len(vals) > risk_col_index:
                                risk_text = vals[risk_col_index]
                            else:
                                risk_text = None
                        else:
                            # Scan values for the first one matching the pattern
                            risk_text = None
                            for v in vals:
                                if isinstance(v, str) and '(' in v and any(ch.isdigit() for ch in v):
                                    risk_text = v
                                    break

                        if risk_text and '(' in risk_text:
                            num_part = risk_text.split('(')[0].strip()
                            try:
                                val_f = float(num_part)
                                risk_values.append(val_f)
                            except ValueError:
                                pass
                    if risk_values:
                        avg_val = sum(risk_values) / len(risk_values)
                except Exception:
                    pass
            except Exception:
                avg_val = None
                division_avg_val = None

            # Determine band and tag for highlighting
            band = database.dr_priority_band(avg_val if avg_val is not None else 0)
            tag = None
            if avg_val is None or avg_val <= 0:
                tag = 'risk_na'
            else:
                tag_name = get_risk_color(avg_val)
                tag = f"risk_{tag_name}" if tag_name else 'risk_na'

            # Division average highlighting
            div_band = database.dr_priority_band(division_avg_val if division_avg_val is not None else 0)
            div_tag = None
            if division_avg_val is None or division_avg_val <= 0:
                div_tag = 'risk_na'
            else:
                div_tag_name = get_risk_color(division_avg_val)
                div_tag = f"risk_{div_tag_name}" if div_tag_name else 'risk_na'

            # fetch notes from DB for selected app/category
            notes = None
            if self.selected_app_id is not None and self.selected_category_name:
                import sqlite3
                conn = sqlite3.connect('business_apps.db')
                c = conn.cursor()
                c.execute('SELECT notes FROM category_notes WHERE app_id = ? AND category = ?', (self.selected_app_id, self.selected_category_name))
                row = c.fetchone()
                if row:
                    notes = row[0]
                conn.close()

            # display division average and system average risk lines followed by notes/default text
            self.details_text.configure(state='normal')
            self.details_text.delete('1.0', 'end')
            # Division line
            self.details_text.insert('end', 'Avg Division Integration Risk: ')
            if division_avg_val is None or division_avg_val <= 0:
                self.details_text.insert('end', 'N/A', div_tag)
            else:
                self.details_text.insert('end', f"{division_avg_val:.1f}", div_tag)
            self.details_text.insert('end', f" ({div_band})\n")
            # System line (average of currently shown integrations)
            self.details_text.insert('end', 'Avg System Integration Risk: ')
            if avg_val is None or avg_val <= 0:
                self.details_text.insert('end', 'N/A', tag)
            else:
                self.details_text.insert('end', f"{avg_val:.1f}", tag)
            self.details_text.insert('end', f" ({band})\n\n")
            self.details_text.insert('end', notes if notes else 'No notes')
            self.details_text.configure(state='disabled')
            
            # refresh integrations table
            # Already refreshed earlier when computing system average; still call to keep behavior consistent
            self.refresh_integration_table()
            
        except Exception as e:
            messagebox.showerror('Error', f'Failed to load selected system: {e}')

    def create_note(self):
        # Allow appending a note for the currently selected app
        if not hasattr(self, 'selected_app_id') or self.selected_app_id is None or not self.selected_category_name:
            messagebox.showinfo('Create Note', 'Please select a system and category to create a note for.')
            return
        try:
            self.details_text.configure(state='normal')
            existing_notes = self.details_text.get('1.0', 'end').strip()
            if existing_notes == 'No notes':
                existing_notes = ''
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            new_note = f"[{timestamp}]\n"
            self.details_text.delete('1.0', 'end')  # Clear existing content
            self.details_text.insert('1.0', f"{new_note}\n{existing_notes}")  # Append new note at the top
            self.details_text.mark_set('insert', f'1.{len(new_note)}')  # Place cursor after the timestamp
            self.details_text.focus_set()
        except Exception:
            messagebox.showerror('Error', 'Failed to prepare note creation.')

    def save_notes(self):
        # Save notes text into DB for selected app
        if not hasattr(self, 'selected_app_id') or self.selected_app_id is None or not self.selected_category_name:
            messagebox.showinfo('Save Note', 'Please select a system and category to save notes for.')
            return
        try:
            raw_text = self.details_text.get('1.0', 'end').strip()
            # Remove the first two dynamic risk lines if present
            lines = raw_text.splitlines()
            cleaned_lines = []
            skip_prefixes = ['Avg Division Integration Risk:', 'Avg System Integration Risk:']
            skipped = 0
            for line in lines:
                if any(line.startswith(p) for p in skip_prefixes) and skipped < 2:
                    skipped += 1
                    continue
                cleaned_lines.append(line)
            notes_text = '\n'.join(cleaned_lines).strip()
            app_id = self.selected_app_id
            category = self.selected_category_name
            import sqlite3
            conn = sqlite3.connect('business_apps.db')
            c = conn.cursor()
            c.execute('''INSERT INTO category_notes (app_id, category, notes) VALUES (?, ?, ?)
                         ON CONFLICT(app_id, category) DO UPDATE SET notes=excluded.notes''', (app_id, category, notes_text))
            conn.commit()
            conn.close()
            
            # Make read-only again
            self.details_text.configure(state='disabled')
            
            # Refresh table to reflect last_modified change
            self.refresh_table()
            
            # Re-select the previously selected application
            for item in self.tree.get_children():
                try:
                    if int(item.split(':', 1)[0]) == app_id:
                        self.tree.selection_set(item)
                        self.tree.see(item)  # Ensure the selected item is visible
                        break
                except ValueError:
                    continue
                    
            messagebox.showinfo('Saved', 'Notes saved.')
        except Exception as e:
            messagebox.showerror('Error', f'Failed to save notes: {e}')

    def sort_table(self, col, reverse):
        # Sort the table by the given column
        try:
            # Store current selection to restore after sorting
            current_selection = self.tree.selection()
            selected_ids = []
            if current_selection:
                for item in current_selection:
                    try:
                        selected_ids.append(int(item.split(':', 1)[0]))
                    except ValueError:
                        continue
            
            # Get data for sorting
            data = [(self.tree.set(k, col), k) for k in self.tree.get_children('')]
            
            # Sort the data based on appropriate type conversion
            data.sort(reverse=reverse, key=lambda t: self._convert_sort_value(t[0]))
            
            # Rearrange items in the tree
            for index, (val, k) in enumerate(data):
                self.tree.move(k, '', index)
            
            # Update the heading to show sort direction and set command for next sort
            sort_direction = "▼" if reverse else "▲"  # Down arrow for descending, up for ascending
            
            # Get original heading text without any arrows
            heading_text = self.tree.heading(col, 'text').replace("▲", "").replace("▼", "").strip()
            
            # Reset all column headings to remove any previous sort indicators
            for column in self.tree['columns']:
                if column != col:  # Skip the column we're currently sorting
                    current_text = self.tree.heading(column, 'text').replace("▲", "").replace("▼", "").strip()
                    self.tree.heading(column, text=current_text)
            
            # Set new heading with sort indicator
            self.tree.heading(col, text=f"{heading_text} {sort_direction}", 
                              command=lambda: self.sort_table(col, not reverse))
            
            # Restore selection if there was one
            if selected_ids:
                for item in self.tree.get_children():
                    try:
                        if int(item.split(':', 1)[0]) in selected_ids:
                            self.tree.selection_add(item)
                            self.tree.see(item)  # Make the first selected item visible
                            break
                    except ValueError:
                        continue
                
        except Exception as e:
            messagebox.showerror('Error', f'Failed to sort table: {e}')

    def _convert_sort_value(self, value):
        # Helper to convert values for sorting (e.g., numeric, date, or string)
        if not value:
            return ""  # Empty values sort first
            
        # Handle risk score format like "8 (High)"
        if isinstance(value, str) and "(" in value and ")" in value:
            try:
                # Extract the numeric part before the parentheses
                numeric_part = value.split('(')[0].strip()
                return float(numeric_part)
            except ValueError:
                pass
        
        # Try to convert to number
        try:
            return float(value)
        except ValueError:
            pass
            
        # Try to convert to date
        try:
            if isinstance(value, str) and '-' in value:
                return datetime.fromisoformat(value)
        except ValueError:
            pass
            
        # Default: return lowercase string for case-insensitive sorting
        return value.lower() if isinstance(value, str) else value

    def refresh_departments(self):
        # Refresh the department listbox with current data
        self.department_listbox.delete(0, 'end')
        departments = self.get_departments()
        for dept_id, dept_name in departments:
            self.department_listbox.insert('end', dept_name)

    def refresh_division_listbox(self):
        """Populate the division listbox with distinct divisions from applications table."""
        if not hasattr(self, 'division_listbox'):
            return
        try:
            import sqlite3
            conn = sqlite3.connect('business_apps.db')
            c = conn.cursor()
            c.execute('SELECT DISTINCT division FROM applications WHERE division IS NOT NULL AND TRIM(division) != "" ORDER BY division')
            divisions = [row[0] for row in c.fetchall() if row and row[0]]
            conn.close()
        except Exception:
            divisions = []
        # Preserve current selection text if possible
        current_selection = None
        try:
            sel = self.division_listbox.curselection()
            if sel:
                current_selection = self.division_listbox.get(sel[0])
        except Exception:
            pass
        self.division_listbox.delete(0, tk.END)
        for d in divisions:
            self.division_listbox.insert(tk.END, d)
        if current_selection:
            try:
                for i in range(self.division_listbox.size()):
                    if self.division_listbox.get(i) == current_selection:
                        self.division_listbox.selection_set(i)
                        self.division_listbox.see(i)
                        break
            except Exception:
                pass

    def delete_selected_app(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo('Delete', 'Please select an application to delete.')
            return
        item = sel[0]
        # get app id from iid
        try:
            app_id = int(item.split(':', 1)[0])
        except Exception:
            messagebox.showerror('Delete', 'Could not determine application id for selected row.')
            return
        if messagebox.askyesno('Confirm Delete', 'Are you sure you want to delete the selected application? This cannot be undone.'):
            conn = database.sqlite3.connect(database.DB_NAME)
            c = conn.cursor()
            # Remove from application_departments first to avoid foreign key constraint
            c.execute('DELETE FROM application_departments WHERE app_id = ?', (app_id,))
            # Now remove the application
            c.execute('DELETE FROM applications WHERE id = ?', (app_id,))
            conn.commit()
            conn.close()
            self.refresh_table()
            messagebox.showinfo('Deleted', 'Application deleted.')

    def _calculate_integration_risk(self, factors: dict) -> float:
        """Calculate the weighted risk percentage (0–100) for an integration.

        Accepts a dict with any of these keys per factor (case-insensitive):
        - score | Score | integration_score
        - need | Need | integration_need
        - criticality | Criticality | integration_criticality
        - installed | Installed | integration_installed
        - disaster_recovery | DisasterRecovery | disasterrecovery | integration_dr | dr
        - safety | Safety | integration_safety
        - security | Security | integration_security
        - monetary | Monetary | integration_monetary
        - customer_service | CustomerService | customerservice | integration_customerservice
        """
        weights = {
            'score': 10.0,
            'need': 0.01,
            'criticality': 0.1,
            'installed': 0.01,
            'disaster_recovery': 0.05,
            'safety': 0.01,
            'security': 0.01,
            'monetary': 0.01,
            'customer_service': 0.01,
        }
        aliases = {
            'score': ['score', 'Score', 'integration_score'],
            'need': ['need', 'Need', 'integration_need'],
            'criticality': ['criticality', 'Criticality', 'integration_criticality'],
            'installed': ['installed', 'Installed', 'integration_installed'],
            'disaster_recovery': ['disaster_recovery', 'DisasterRecovery', 'disasterrecovery', 'integration_dr', 'dr'],
            'safety': ['safety', 'Safety', 'integration_safety'],
            'security': ['security', 'Security', 'integration_security'],
            'monetary': ['monetary', 'Monetary', 'integration_monetary'],
            'customer_service': ['customer_service', 'CustomerService', 'customerservice', 'integration_customerservice'],
        }

        def _get_int(val):
            try:
                return int(val)
            except Exception:
                try:
                    return int(float(val))
                except Exception:
                    return 0

        def _get_factor(key: str) -> int:
            names = aliases.get(key, [key])
            for nm in names:
                if nm in factors and factors.get(nm) is not None:
                    return _get_int(factors.get(nm))
            return 0

        # Exclude factors with value 0 from the calculation (treat 0 as N/A)
        weighted_total = 0.0
        max_total = 0.0
        for k, w in weights.items():
            v = _get_factor(k)
            if v > 0:
                if k == 'score':
                    weighted_total += (10 - v) * w  # Invert score: lower score = higher risk
                else:
                    weighted_total += v * w
                max_total += 10 * w
        return 0.0 if max_total == 0 else (weighted_total / max_total) * 100.0

    def add_system_integration(self):
        """Open dialog to add a new system integration to the selected parent system."""
        if self.current_parent_system_id is None:
            messagebox.showinfo('Add Integration', 'Please select a parent system first.')
            return

        # Create and configure dialog
        dialog = tk.Toplevel(self)
        dialog.title('Add Integration')
        dialog.geometry('650x700')
        dialog.minsize(650, 700)
        dialog.resizable(False, False)
        dialog.transient(self)
        dialog.grab_set()

        # Create main container with grid
        main_container = ttk.Frame(dialog)
        main_container.grid(row=0, column=0, sticky='nsew', padx=10, pady=10)

        # Configure dialog grid
        dialog.grid_rowconfigure(0, weight=1)
        dialog.grid_columnconfigure(0, weight=1)

        # Configure main container grid (two columns: form | score guide)
        main_container.grid_rowconfigure(0, weight=1)
        main_container.grid_columnconfigure(0, weight=1)  # form column grows
        # Slightly wider guide column for better readability
        main_container.grid_columnconfigure(1, weight=0, minsize=280)  # guide column fixed width

        # Form frame setup with grid
        form_frame = ttk.Frame(main_container, padding=15)
        form_frame.grid(row=0, column=0, sticky='nsew')

        # Configure form frame grid
        form_frame.grid_columnconfigure(1, weight=1)

        # Name and vendor fields
        ttk.Label(form_frame, text='Integration Name:', anchor='e', style='FormLabel.TLabel').grid(row=0, column=0, sticky='e', pady=5, padx=5)
        name_entry = ttk.Entry(form_frame, width=40)
        name_entry.grid(row=0, column=1, sticky='ew', pady=5, padx=5)
        name_entry.focus_set()

        ttk.Label(form_frame, text='Vendor:', anchor='e', style='FormLabel.TLabel').grid(row=1, column=0, sticky='e', pady=5, padx=5)
        vendor_entry = ttk.Entry(form_frame, width=40)
        vendor_entry.grid(row=1, column=1, sticky='ew', pady=5, padx=5)

        # Factor entry fields
        factor_entries = {}
        keys = ['Score', 'Need', 'Criticality', 'Installed', 'DisasterRecovery', 'Safety', 'Security', 'Monetary', 'CustomerService']

        for idx, key in enumerate(keys, start=2):
            ttk.Label(form_frame, text=f'{key}:', anchor='e', style='FormLabel.TLabel').grid(row=idx, column=0, sticky='e', pady=5, padx=5)
            entry = ttk.Entry(form_frame, width=5)
            entry.grid(row=idx, column=1, sticky='w', pady=5, padx=5)
            factor_entries[key] = entry

        # Notes field
        ttk.Label(form_frame, text='Notes:', anchor='e', style='FormLabel.TLabel').grid(row=11, column=0, sticky='ne', pady=5, padx=5)
        notes_text = tk.Text(form_frame, height=4, width=40, wrap='word')
        notes_text.grid(row=11, column=1, sticky='w', pady=5, padx=5)

        # Empty Score Guide labelframe on the right
        guide_frame = ttk.Labelframe(main_container, text='Score Guide', padding=12, style='Guide.TLabelframe')
        guide_frame.grid(row=0, column=1, sticky='n', padx=(10, 0), pady=10)
        # Title + body to improve readability and match style
        baseline_msg = 'Entering a 0 for any factor will exclude that factor from the Risk calculation. Ratings should be from 1 (low) to 10 (high).'
        guide_title_label = ttk.Label(guide_frame, text='Note --- ', anchor='w', style='GuideNumber.TLabel')
        guide_title_label.grid(row=0, column=0, sticky='w', pady=(0, 2))
        # Subtle divider between title and body
        ttk.Separator(guide_frame, orient='horizontal').grid(row=1, column=0, sticky='ew', pady=(0, 6))
        guide_body_label = ttk.Label(guide_frame, text=baseline_msg, justify='left', anchor='nw', wraplength=260, style='GuideText.TLabel')
        guide_body_label.grid(row=2, column=0, sticky='nw')
        # Spacer to ensure visibility (empty box area)
        guide_spacer = ttk.Frame(guide_frame, width=280, height=220)
        guide_spacer.grid(row=3, column=0)
        guide_spacer.grid_propagate(False)

        # Dynamic guide: show guidance text for each field on focus/click
        if 'Score' in factor_entries:
            SCORE_GUIDE_TEXT = (
                "Guiding Question\n"
                "• How would you rate the overall performance and reliability of this system today?\n\n"
                "Scoring Help\n"
                "• 1–3: Poor/unstable performance or frequent issues\n"
                "• 4–6: Adequate for current needs but with notable limitations\n"
                "• 7–8: Strong performance/reliability with minor gaps\n"
                "• 9–10: Excellent/mission-grade performance and reliability"
            )
            def _show_score_guide(_e=None):
                try:
                    guide_title_label.config(text='Score')
                    guide_body_label.config(text=SCORE_GUIDE_TEXT)
                except Exception:
                    pass
            factor_entries['Score'].bind('<FocusIn>', _show_score_guide)
            factor_entries['Score'].bind('<Button-1>', _show_score_guide)

        if 'Need' in factor_entries:
            NEED_GUIDE_TEXT = (
                "Guiding Question\n"
                "• Is this system required to operate the business, or is it optional/nice-to-have?\n\n"
                "Scoring Help\n"
                "• 1–3: Optional; workarounds readily available\n"
                "• 4–6: Helpful but not strictly required; alternatives exist\n"
                "• 7–8: Required for multiple teams or core workflows\n"
                "• 9–10: Absolutely required to operate or meet obligations"
            )
            def _show_need_guide(_e=None):
                try:
                    guide_title_label.config(text='Need')
                    guide_body_label.config(text=NEED_GUIDE_TEXT)
                except Exception:
                    pass
            factor_entries['Need'].bind('<FocusIn>', _show_need_guide)
            factor_entries['Need'].bind('<Button-1>', _show_need_guide)

        if 'Criticality' in factor_entries:
            CRIT_GUIDE_TEXT = (
                "Guiding Question\n"
                "• What is the business impact if this system is unavailable or fails?\n\n"
                "Scoring Help\n"
                "• 1–3: Minor inconvenience; small/localized impact\n"
                "• 4–6: Moderate productivity loss; delays but manageable\n"
                "• 7–8: Major disruption to critical processes\n"
                "• 9–10: Immediate stoppage/safety/regulatory impact"
            )
            def _show_crit_guide(_e=None):
                try:
                    guide_title_label.config(text='Criticality')
                    guide_body_label.config(text=CRIT_GUIDE_TEXT)
                except Exception:
                    pass
            factor_entries['Criticality'].bind('<FocusIn>', _show_crit_guide)
            factor_entries['Criticality'].bind('<Button-1>', _show_crit_guide)

        if 'Installed' in factor_entries:
            INST_GUIDE_TEXT = (
                "Guiding Question\n"
                "• How entrenched and widely adopted is this system across the organization?\n\n"
                "Scoring Help\n"
                "• 1–3: Pilot/limited use; easy to replace\n"
                "• 4–6: Moderate adoption; switching cost exists but manageable\n"
                "• 7–8: Widely adopted; high dependency across teams\n"
                "• 9–10: Deeply entrenched; very high switching cost/risk"
            )
            def _show_inst_guide(_e=None):
                try:
                    guide_title_label.config(text='Installed')
                    guide_body_label.config(text=INST_GUIDE_TEXT)
                except Exception:
                    pass
            factor_entries['Installed'].bind('<FocusIn>', _show_inst_guide)
            factor_entries['Installed'].bind('<Button-1>', _show_inst_guide)

        if 'DisasterRecovery' in factor_entries:
            DR_GUIDE_TEXT = (
                "Guiding Question\n"
                "• How quickly must this system be recovered after an outage, and how much data loss is tolerable?\n\n"
                "Scoring Help\n"
                "• 1–3: Long RTO/RPO acceptable; daily backups sufficient\n"
                "• 4–6: Recovery within hours required; some data loss acceptable\n"
                "• 7–8: Near real-time recovery/replication expected\n"
                "• 9–10: Continuous availability; near-zero downtime/data loss"
            )
            def _show_dr_guide(_e=None):
                try:
                    guide_title_label.config(text='Disaster Recovery')
                    guide_body_label.config(text=DR_GUIDE_TEXT)
                except Exception:
                    pass
            factor_entries['DisasterRecovery'].bind('<FocusIn>', _show_dr_guide)
            factor_entries['DisasterRecovery'].bind('<Button-1>', _show_dr_guide)

        if 'Safety' in factor_entries:
            SAFETY_GUIDE_TEXT = (
                "Guiding Question\n"
                "• Could failure of this system impact human safety or regulatory health/safety obligations?\n\n"
                "Scoring Help\n"
                "• 1–2: No relation to safety\n"
                "• 3–4: Indirect/low safety relevance\n"
                "• 5–6: Some safety implication depending on context\n"
                "• 7–8: High safety implications in certain scenarios\n"
                "• 9–10: Life/safety critical"
            )
            def _show_safety_guide(_e=None):
                try:
                    guide_title_label.config(text='Safety')
                    guide_body_label.config(text=SAFETY_GUIDE_TEXT)
                except Exception:
                    pass
            factor_entries['Safety'].bind('<FocusIn>', _show_safety_guide)
            factor_entries['Safety'].bind('<Button-1>', _show_safety_guide)

        if 'Security' in factor_entries:
            SECURITY_GUIDE_TEXT = (
                "Guiding Question\n"
                "• What is the security exposure if this system is compromised or misused?\n\n"
                "Scoring Help\n"
                "• 1–3: Isolated/low-sensitivity data; limited access\n"
                "• 4–6: Internal data with moderate sensitivity\n"
                "• 7–8: Handles sensitive/regulated data or external integrations\n"
                "• 9–10: High-risk target; critical/regulated data and broad access"
            )
            def _show_security_guide(_e=None):
                try:
                    guide_title_label.config(text='Security')
                    guide_body_label.config(text=SECURITY_GUIDE_TEXT)
                except Exception:
                    pass
            factor_entries['Security'].bind('<FocusIn>', _show_security_guide)
            factor_entries['Security'].bind('<Button-1>', _show_security_guide)

        if 'Monetary' in factor_entries:
            MONETARY_GUIDE_TEXT = (
                "Guiding Question\n"
                "• What is the direct financial impact of downtime, failure, or poor performance?\n\n"
                "Scoring Help\n"
                "• 1–3: Minimal/indirect cost impact\n"
                "• 4–6: Moderate costs or revenue impact\n"
                "• 7–8: High revenue impact or material cost exposure\n"
                "• 9–10: Immediate material revenue loss/penalties"
            )
            def _show_monetary_guide(_e=None):
                try:
                    guide_title_label.config(text='Monetary')
                    guide_body_label.config(text=MONETARY_GUIDE_TEXT)
                except Exception:
                    pass
            factor_entries['Monetary'].bind('<FocusIn>', _show_monetary_guide)
            factor_entries['Monetary'].bind('<Button-1>', _show_monetary_guide)

        if 'CustomerService' in factor_entries:
            CS_GUIDE_TEXT = (
                "Guiding Question\n"
                "• How directly does this system affect the customer experience or service quality?\n\n"
                "Scoring Help\n"
                "• 1–3: Internal-only; no direct customer effect\n"
                "• 4–6: Indirect impact on customer interactions\n"
                "• 7–8: Customer-facing; noticeable impact on service quality\n"
                "• 9–10: Primary channel or critical to customer experience"
            )
            def _show_cs_guide(_e=None):
                try:
                    guide_title_label.config(text='Customer Service')
                    guide_body_label.config(text=CS_GUIDE_TEXT)
                except Exception:
                    pass
            factor_entries['CustomerService'].bind('<FocusIn>', _show_cs_guide)
            factor_entries['CustomerService'].bind('<Button-1>', _show_cs_guide)

        # Add separator
        ttk.Separator(main_container, orient='horizontal').grid(row=1, column=0, columnspan=2, sticky='ew', pady=(10, 5))

        # Button frame at bottom
        button_frame = ttk.Frame(main_container)
        button_frame.grid(row=2, column=0, columnspan=2, sticky='ew', pady=5)

        # Configure button frame grid
        button_frame.grid_columnconfigure(0, weight=1)  # Push buttons to the right

        # Submit handler
        def submit_integration():
            # Get values from form
            try:
                name = name_entry.get().strip()
                if not name:
                    messagebox.showwarning('Validation Error', 'Integration name is required')
                    return

                # Create a dictionary with appropriate types for each field
                fields: dict[str, object] = {
                    'name': name,
                    'vendor': vendor_entry.get().strip(),
                    'notes': notes_text.get('1.0', 'end-1c').strip(),
                }

                # Get factor values and create ratings dictionary for score calculation
                ratings = {}
                db_key_map = {
                    'Score': 'score',
                    'Need': 'need',
                    'Criticality': 'criticality',
                    'Installed': 'installed',
                    'DisasterRecovery': 'disaster_recovery',
                    'Safety': 'safety',
                    'Security': 'security',
                    'Monetary': 'monetary',
                    'CustomerService': 'customer_service',
                }
                for key in keys:
                    value = factor_entries[key].get().strip()
                    if value:
                        try:
                            int_value = int(value)
                        except ValueError:
                            messagebox.showwarning('Validation Error', f'{key} must be a number')
                            return
                    else:
                        int_value = 0
                    ratings[key] = int_value
                    fields[db_key_map.get(key, key.lower())] = int_value

                # Calculate risk score (weighted percentage 0–100); keep string type for legacy flow
                fields['risk_score'] = f"{self._calculate_integration_risk(ratings):.2f}"

                # Submit integration to database
                integration_id = database.add_system_integration(self.current_parent_system_id, fields)
                if integration_id:
                    # Update parent application's last_modified so main table timestamp reflects new integration
                    try:
                        if self.current_parent_system_id:
                            database.touch_application(self.current_parent_system_id)
                    except Exception:
                        pass
                    selected_cat = getattr(self, 'selected_category_name', None)
                    # If a category is currently selected in the main table, link this integration to that category
                    if selected_cat:
                        try:
                            conn_link = database.connect_db()
                            cur_link = conn_link.cursor()
                            # Look up category id by name (case-insensitive)
                            cur_link.execute('SELECT id FROM categories WHERE LOWER(TRIM(name)) = LOWER(TRIM(?))', (selected_cat,))
                            row_cat = cur_link.fetchone()
                            if row_cat and row_cat[0]:
                                cat_id = row_cat[0]
                                # Ensure integration_categories table exists then insert link
                                cur_link.execute('''CREATE TABLE IF NOT EXISTS integration_categories (
                                    integration_id INTEGER NOT NULL,
                                    category_id INTEGER NOT NULL,
                                    PRIMARY KEY (integration_id, category_id),
                                    FOREIGN KEY (integration_id) REFERENCES system_integrations(id) ON DELETE CASCADE,
                                    FOREIGN KEY (category_id) REFERENCES categories(id) ON DELETE CASCADE
                                )''')
                                cur_link.execute('INSERT OR IGNORE INTO integration_categories (integration_id, category_id) VALUES (?, ?)', (integration_id, cat_id))
                                conn_link.commit()
                            conn_link.close()
                        except Exception:
                            pass
                    messagebox.showinfo('Success', 'System integration added successfully')
                    dialog.destroy()
                    # Refresh the integrations table
                    self.refresh_integration_table()
                    # Auto-scroll & highlight new integration row if present
                    try:
                        new_iid = str(integration_id)
                        if new_iid in self.integrations_tree.get_children(''):
                            # Scroll into view and (optionally) select but do not force focus or open edit dialog
                            self.integrations_tree.see(new_iid)
                            try:
                                self.integrations_tree.selection_set(new_iid)
                            except Exception:
                                pass
                            # Create a temporary highlight tag
                            hl_tag = 'new_integration_highlight'
                            if not hasattr(self, '_new_integration_highlight_configured'):
                                try:
                                    self.integrations_tree.tag_configure(hl_tag, background='#99ccff')
                                    setattr(self, '_new_integration_highlight_configured', True)
                                except Exception:
                                    pass
                            # Apply highlight tag (keep existing risk color tag if any by re-setting tags)
                            try:
                                current_vals = self.integrations_tree.item(new_iid, 'values')
                                current_tags = list(self.integrations_tree.item(new_iid, 'tags') or [])
                                if hl_tag not in current_tags:
                                    current_tags.append(hl_tag)
                                self.integrations_tree.item(new_iid, values=current_vals, tags=tuple(current_tags))
                            except Exception:
                                pass
                            # Schedule removal of highlight after 2 seconds
                            def _remove_highlight():
                                try:
                                    cur_tags = list(self.integrations_tree.item(new_iid, 'tags') or [])
                                    if hl_tag in cur_tags:
                                        cur_tags.remove(hl_tag)
                                        self.integrations_tree.item(new_iid, tags=tuple(cur_tags))
                                except Exception:
                                    pass
                            self.after(2000, _remove_highlight)
                            # (Auto-open edit dialog removed per revert request)
                    except Exception:
                        pass
                    # Refresh main applications table to update averages / last_modified
                    try:
                        # Remember current composite iid (appId:category)
                        remember_iid = None
                        if hasattr(self, 'selected_app_id') and self.selected_app_id is not None:
                            remember_iid = f"{self.selected_app_id}:{selected_cat or ''}"
                        self.refresh_table()
                        # Reselect the previously selected row so context/details persist
                        if remember_iid and remember_iid in self.tree.get_children(''):
                            try:
                                self.tree.selection_set(remember_iid)
                                self.tree.focus(remember_iid)
                            except Exception:
                                pass
                    except Exception:
                        pass
                    # Recompute averages / details after refresh
                    try:
                        self.on_tree_select(None)
                    except Exception:
                        pass
                else:
                    messagebox.showerror('Error', 'Failed to submit system integration')
            except Exception as e:
                messagebox.showerror('Error', f'Failed to add system integration: {e}')

        # Create buttons using grid
        submit_btn = ttk.Button(
            button_frame,
            text='Add',
            command=submit_integration,
            style='Primary.TButton',
            width=15
        )
        submit_btn.grid(row=0, column=1, padx=5)

        cancel_btn = ttk.Button(
            button_frame,
            text='Cancel',
            command=dialog.destroy,
            style='Secondary.TButton',
            width=15
        )
        cancel_btn.grid(row=0, column=0, padx=5)
    
    def import_csv_dialog(self):
        """
        Open file dialog to select a single CSV file and import it.
        Expects the CSV to match the smoke-test single-file format (applications + integrations rows).
        """
        try:
            from tkinter import filedialog
            path = filedialog.askopenfilename(title='Select CSV file to import', filetypes=[('CSV files', '*.csv'), ('All files', '*.*')])
            if not path:
                return
            # Run import in background to avoid blocking the Tk mainloop
            import threading
            t = threading.Thread(target=lambda: self.import_csv_worker(path), daemon=True)
            t.start()
        except Exception as e:
            messagebox.showerror('Import Error', f'Failed to open CSV file: {e}')

    # Minimal progress window used during long-running imports
    class ProgressWindow(tk.Toplevel):
        def __init__(self, parent, title='Progress', message='Working...'):
            super().__init__(parent)
            self.title(title)
            self.transient(parent)
            self.grab_set()
            self.resizable(False, False)
            self.protocol('WM_DELETE_WINDOW', lambda: None)
            
            # Create main frame with padding
            self.frame = ttk.Frame(self, padding=10)
            self.frame.pack(fill='both', expand=True)
            
            # Message label above progress bar
            self.label = ttk.Label(self.frame, text=message)
            self.label.pack(fill='x', expand=True, pady=(0, 5))
            
            # Progress bar
            self.progress = ttk.Progressbar(self.frame, mode='determinate', length=300)
            self.progress.pack(fill='x', expand=True)
            
            # Row count label below progress bar
            self.row_label = ttk.Label(self.frame, text='Processing row: 0')
            self.row_label.pack(fill='x', expand=True, pady=(5, 0))
            # Keep window small and above
            try:
                self.attributes('-topmost', True)
            except Exception:
                pass

        def show(self):
            # Center over parent and force window to appear
            try:
                self.update_idletasks()
                parent = self.master
                x = parent.winfo_rootx() + (parent.winfo_width() - self.winfo_width()) // 2
                y = parent.winfo_rooty() + (parent.winfo_height() - self.winfo_height()) // 2
                self.geometry(f'+{x}+{y}')
                self.deiconify()
                self.lift()
                self.focus_force()
                self.update()
            except Exception:
                pass

        def update_message(self, text, row=None, total=None):
            try:
                self.label.config(text=text)
                if row is not None:
                    self.row_label.configure(text=f'Processing row: {row}')
                    if total:
                        # Update progress bar
                        progress = (row / total) * 100
                        self.progress['value'] = progress
                        self.row_label.configure(text=f'Processing row: {row} of {total}')
                self.lift()
                self.update_idletasks()
                self.update()
            except Exception:
                pass

        def close(self):
            try:
                self.grab_release()
            except Exception:
                pass
            try:
                self.destroy()
            except Exception:
                pass

    def import_csv_file(self, path, progress_callback=None):
        """
        Parse the CSV at `path` and insert/update applications, application-department links, and integrations.
        Multiple rows for the same application are allowed (multiple integrations). Missing numeric fields will be coerced or defaulted.
        """
        import csv
        import time
        created_apps = 0
        updated_apps = 0
        created_integrations = 0
        updated_integrations = 0
        errors = []

        # Expected header columns (case-insensitive match)
        expected_cols = [
            'division', 'name', 'vendor', 'business_unit', 'score', 'need', 'criticality', 'installed',
            'disasterrecovery', 'safety', 'security', 'monetary', 'customerservice', 'notes',
            'category',
            'integration_name', 'integration_vendor', 'integration_score', 'integration_need', 'integration_criticality',
            'integration_installed', 'integration_dr', 'integration_safety', 'integration_security', 'integration_monetary',
            'integration_customerservice', 'integration_risk', 'integration_last_modified'
        ]

        # Initialize connection as None for proper cleanup
        conn = None
        cur = None

        # Detect CSV delimiter (handles tab, comma, semicolon, pipe)
        def _detect_delimiter(pth):
            """Robust delimiter detection with heuristic and sniffer fallback."""
            candidates = [',', '\t', ';', '|']
            try:
                # Read a small initial chunk and the first non-empty line
                with open(pth, 'r', encoding='utf-8-sig') as f:
                    head = f.read(4096)
                    # Heuristic: count occurrences in the first line
                    first_line = head.splitlines()[0] if head else ''
                    counts = {d: first_line.count(d) for d in candidates}
                    # Choose delimiter with highest count in first line
                    best = max(counts.keys(), key=lambda d: counts[d]) if counts else ','
                    if counts.get(best, 0) > 0:
                        return best
                    # Fallback to csv.Sniffer across candidates
                    try:
                        sniffer = csv.Sniffer()
                        dialect = sniffer.sniff(head, delimiters=''.join(candidates))
                        return getattr(dialect, 'delimiter', ',')
                    except Exception:
                        pass
            except Exception:
                pass
            # Last resort default to tab if present in file, else comma
            try:
                with open(pth, 'r', encoding='utf-8-sig') as f:
                    body = f.read(2048)
                    if '\t' in body:
                        return '\t'
            except Exception:
                pass
            return ','

        delim = _detect_delimiter(path)

        # First count total rows
        total_rows = 0
        try:
            with open(path, newline='', encoding='utf-8-sig') as fh:
                total_rows = sum(1 for _ in csv.DictReader(fh, delimiter=delim))
        except Exception as e:
            return {'error': f'Failed to count rows: {e}'}

        # Initialize database schema first
        database.initialize_database()
        print("DEBUG: Database schema initialized")
        
        # Debug: Print out header fields
        with open(path, newline='', encoding='utf-8-sig') as fh:
            reader = csv.DictReader(fh, delimiter=delim)
            if reader.fieldnames:
                print("DEBUG: CSV Headers:", reader.fieldnames)
        
        # Open file and process rows
        try:
            with open(path, newline='', encoding='utf-8-sig') as fh:
                reader = csv.DictReader(fh, delimiter=delim)
                if reader.fieldnames is None:
                    return {'error': 'CSV file has no header row.'}
                # normalize header names - keep both normalized and original versions
                import re
                def _norm_key(s: str) -> str:
                    try:
                        return re.sub(r'[^a-z0-9]', '', str(s).strip().lower())
                    except Exception:
                        return ''
                header_map = {}
                for h in reader.fieldnames:
                    normalized = _norm_key(h)
                    header_map[normalized] = h
                    header_map[str(h).strip()] = h  # Also keep original version

                # check presence of minimal columns - accept many aliases for application name
                required_name_aliases = [
                    'name', 'app_name', 'appname',
                    'application_name', 'applicationname', 'application name',
                    'division', 'division_name', 'divisionname', 'division name',
                    'system', 'system_name', 'systemname', 'system name',
                    'app name', 'App Name', 'Application Name'
                ]
                if not any((alias in header_map or ''.join(filter(str.isalnum, alias.lower())) in header_map) for alias in required_name_aliases):
                    return {'error': 'CSV must include an Application/Division name column (e.g., "Application Name", "App Name", "Division", or "Name").'}

                # Create a new database connection for this thread with timeout and immediate mode
                import sqlite3
                # Use default (deferred) isolation to minimize locking; extend timeout
                conn = sqlite3.connect(database.DB_NAME, timeout=60.0)
                conn.execute('PRAGMA busy_timeout = 60000')  # Set busy timeout to 60 seconds
                cur = conn.cursor()

                # helper to find or create business unit with retries
                def ensure_business_unit(name):
                    if not name:
                        return None
                    try:
                        # Normalize name: trim and collapse internal whitespace, case-insensitive match for lookup
                        name_norm = ' '.join(str(name).split()).strip()
                        # First try to find existing business unit (case/trim-insensitive)
                        cur.execute('SELECT id FROM business_units WHERE LOWER(TRIM(name)) = LOWER(TRIM(?))', (name_norm,))
                        row = cur.fetchone()
                        if row:
                            return row[0]
                        # Create new business unit with retries
                        max_retries = 3
                        last_error = None
                        for retry in range(max_retries):
                            try:
                                print(f"DEBUG: Attempting to create business unit: {name_norm}")
                                cur.execute('INSERT INTO business_units (name, last_modified) VALUES (?, CURRENT_TIMESTAMP)', (name_norm,))
                                conn.commit()  # Commit immediately to prevent locks
                                new_id = cur.lastrowid
                                print(f"DEBUG: Created business unit {name_norm} with ID {new_id}")
                                return new_id
                            except sqlite3.OperationalError as e:
                                last_error = e
                                if 'database is locked' in str(e) and retry < max_retries - 1:
                                    time.sleep(0.5 * (retry + 1))  # Exponential backoff
                                    continue
                                print(f"DEBUG: Database lock error creating business unit {name}: {e}")
                                raise
                            except sqlite3.IntegrityError as e:
                                # In case of race condition where unit was created between select and insert
                                if 'UNIQUE constraint failed' in str(e):
                                    print(f"DEBUG: Business unit {name_norm} already exists (race condition)")
                                    cur.execute('SELECT id FROM business_units WHERE LOWER(TRIM(name)) = LOWER(TRIM(?))', (name_norm,))
                                    row = cur.fetchone()
                                    if row:
                                        return row[0]
                                raise
                        if last_error:
                            raise last_error
                    except Exception as e:
                        print(f"DEBUG: Failed to create/find business unit '{name}': {e}")
                        errors.append(f"Failed to create/find business unit '{name}': {e}")
                        return None

                # helper to parse int-like fields
                def parse_int(v, default=0):
                    try:
                        return int(float(str(v)))
                    except Exception:
                        return default

                # helper to parse float
                def parse_float(v, default=0.0):
                    try:
                        return float(v)
                    except Exception:
                        return default

                # cache for app name -> id, remember last non-empty BU(s) per app, and accumulate categories
                app_cache = {}
                app_bu_map = {}
                app_category_accum = {}  # app_name -> set of category names accumulated across rows

                # helper to flexibly read a field by multiple key options using normalized header mapping
                def get(*opts):
                    for opt in opts:
                        # Try exact match
                        if opt in header_map:
                            val = raw.get(header_map[opt])
                            if val is not None:
                                cv = str(val).strip()
                                return '' if cv.lower() == 'none' else cv
                        # Try normalized version
                        nk = _norm_key(opt)
                        if nk in header_map:
                            val = raw.get(header_map[nk])
                            if val is not None:
                                cv = str(val).strip()
                                return '' if cv.lower() == 'none' else cv
                    return ''

                # parse potentially multiple BU names from a field
                def parse_bu_names(val: str):
                    if val is None:
                        return []
                    # split on common delimiters
                    parts = re.split(r'[;,/|]', str(val))
                    names = []
                    for p in parts:
                        nm = ' '.join(p.split()).strip()
                        if nm:
                            names.append(nm)
                    return names

                                # parse multiple category names from Category field
                def parse_category_names(val: str):
                    if val is None:
                        return []
                    # First split on commas or semicolons between major categories
                    categories = re.split(r'\s*[;,]\s*', str(val))
                    out = []
                    for cat in categories:
                        # Then split on forward slash or vertical bar for subcategories
                        subcats = re.split(r'\s*[/|]\s*', cat)
                        for subcat in subcats:
                            nm = ' '.join(subcat.split()).strip()
                            if nm and nm.lower() != 'none':
                                out.append(nm)
                    return out

                # Initialize app name tracking
                last_app_name = None  # carry forward the last non-empty application/division name
                for rownum, raw in enumerate(reader, start=2):
                    try:
                        # Skip completely blank rows (all empty/whitespace)
                        try:
                            if not raw or all((v is None or str(v).strip() == '' for v in raw.values())):
                                continue
                        except Exception:
                            pass
                        # Map fields flexibly with normalized keys
                        app_name = get('name', 'app_name', 'app name', 'application_name', 'application name', 'Application Name', 'Division', 'division', 'division_name', 'system', 'system_name', 'System Name') or ''
                        # If this row omits the division/application name, carry forward the last one
                        if not app_name and last_app_name:
                            app_name = last_app_name
                        # Skip rows without any application name to avoid blank entries
                        if not app_name:
                            continue
                        vendor = get('vendor', 'Vendor') or ''
                        # Get categories from both main category field and any integration-specific categories
                        main_category = get('category', 'Category') or ''
                        int_category = get('integration_category', 'Category Type', 'CategoryType', 'System Type') or ''
                        
                        # Collect all categories
                        all_categories = []
                        if main_category:
                            all_categories.extend(parse_category_names(main_category))
                        if int_category:
                            all_categories.extend(parse_category_names(int_category))
                        
                        # Accumulate unique categories for this app
                        if all_categories:
                            acc = app_category_accum.setdefault(app_name, set())
                            for nm in all_categories:
                                nm_clean = nm.strip()
                                if nm_clean:
                                    acc.add(nm_clean)
                            print(f"DEBUG: Accumulated categories for {app_name}: {sorted(acc)}")
                        # Accept multiple possible BU header names, including 'Business Unit / Division'
                        bu_field_val = get('business_unit', 'businessunit', 'Business Unit', 'business_unit_division', 'businessunitdivision', 'Business Unit / Division', 'department', 'dept', 'bu') or ''
                        bu_names = parse_bu_names(bu_field_val)
                        if bu_names:
                            app_bu_map[app_name] = bu_names
                        notes = get('notes', 'Notes') or ''
                        # Remember last non-empty app name for subsequent rows
                        if app_name:
                            last_app_name = app_name

                        # Application rating fields
                        score = parse_int(get('score', 'Score'), 0)
                        need = parse_int(get('need', 'Need'), 0)
                        criticality = parse_int(get('criticality', 'Criticality'), 0)
                        installed = parse_int(get('installed', 'Installed'), 0)
                        disaster_recovery = parse_int(get('disasterrecovery', 'disaster_recovery', 'dr', 'Disaster Recovery'), 0)
                        safety = parse_int(get('safety', 'Safety'), 0)
                        security = parse_int(get('security', 'Security'), 0)
                        monetary = parse_int(get('monetary', 'Monetary'), 0)
                        customer_service = parse_int(get('customerservice', 'customer_service', 'Customer Service'), 0)

                        # Integration fields
                        int_name = get('integration_name', 'integration') or ''
                        int_vendor = get('integration_vendor') or ''
                        # parse integration score only if provided; keep None to indicate 'not provided'
                        int_score_raw = get('integration_score')
                        int_score = parse_int(int_score_raw, 0) if int_score_raw not in (None, '') else None
                        int_need = parse_int(get('integration_need'), 0)
                        int_criticality = parse_int(get('integration_criticality'), 0)
                        int_installed = parse_int(get('integration_installed'), 0)
                        int_dr = parse_int(get('integration_dr'), 0)
                        int_safety = parse_int(get('integration_safety'), 0)
                        int_security = parse_int(get('integration_security'), 0)
                        int_monetary = parse_int(get('integration_monetary'), 0)
                        int_customer_service = parse_int(get('integration_customerservice'), 0)
                        int_risk = None
                        # integration risk may be provided
                        ir = get('integration_risk')
                        int_risk = parse_float(ir, 0.0) if ir not in (None, '') else None
                        last_mod = get('integration_last_modified') or get('last_modified') or None

                        # Ensure application exists (by name). If multiple apps have same name, pick first.
                        # First get or create all categories for this app from accumulated set
                        try:
                            all_cats = sorted(app_category_accum.get(app_name, set()))
                            cat_ids = []
                            print(f"DEBUG: Processing accumulated categories for {app_name}: {all_cats}")
                            # Create all needed categories first
                            for nm in all_cats:
                                nm_clean = nm.strip()
                                if not nm_clean:
                                    continue
                                cur.execute('SELECT id FROM categories WHERE LOWER(TRIM(name)) = LOWER(TRIM(?))', (nm_clean,))
                                crow = cur.fetchone()
                                if crow:
                                    cat_ids.append(crow[0])
                                else:
                                    cur.execute('INSERT INTO categories (name, last_modified) VALUES (?, CURRENT_TIMESTAMP)', (nm_clean,))
                                    cat_ids.append(cur.lastrowid)
                            conn.commit()
                        except Exception as e:
                            print(f"DEBUG: Error ensuring categories exist: {e}")
                            cat_ids = []

                        # Now check if app exists
                        app_id = app_cache.get(app_name)
                        if not app_id:
                            cur.execute('SELECT id FROM applications WHERE division = ?', (app_name,))
                            r = cur.fetchone()
                            if r:
                                app_id = r[0]
                                app_cache[app_name] = app_id
                                # Update application fields with provided values
                                try:
                                    database.update_application(app_id, {
                                        'vendor': vendor,
                                        'score': score,
                                        'need': need,
                                        'criticality': criticality,
                                        'installed': installed,
                                        'disaster_recovery': disaster_recovery,
                                        'safety': safety,
                                        'security': security,
                                        'monetary': monetary,
                                        'customer_service': customer_service,
                                        'notes': notes
                                    })
                                    # Update categories many-to-many for existing app using local connection
                                    try:
                                        all_cats = sorted(app_category_accum.get(app_name, set()))
                                        cat_ids = []
                                        print(f"DEBUG: Updating categories for existing app {app_name}: {all_cats}")
                                        
                                        # First ensure all categories exist and get their IDs in a single transaction
                                        cat_ids = []
                                        for nm in all_cats:
                                            nm_clean = nm.strip()
                                            if not nm_clean:
                                                continue
                                            # Try to find existing category first
                                            cur.execute('SELECT id FROM categories WHERE LOWER(TRIM(name)) = LOWER(TRIM(?))', (nm_clean,))
                                            crow = cur.fetchone()
                                            if crow:
                                                cat_ids.append(crow[0])
                                            else:
                                                # Create new category
                                                cur.execute('INSERT INTO categories (name, last_modified) VALUES (?, CURRENT_TIMESTAMP)', (nm_clean,))
                                                cat_ids.append(cur.lastrowid)
                                        
                                        # Update category links in same transaction
                                        if cat_ids:
                                            print(f"DEBUG: Setting {len(cat_ids)} categories for existing app {app_name} (ID: {app_id})")
                                            # First remove old links
                                            cur.execute('DELETE FROM application_categories WHERE app_id = ?', (app_id,))
                                            # Add new links
                                            for cid in cat_ids:
                                                cur.execute('INSERT OR IGNORE INTO application_categories (app_id, category_id) VALUES (?, ?)', 
                                                          (app_id, cid))
                                            print(f"DEBUG: Successfully updated categories for app {app_name}")
                                            # Ensure category link changes persist even if no integration row commits later
                                            try:
                                                conn.commit()
                                            except Exception as e:
                                                print(f"DEBUG: Commit failed after updating categories for {app_name}: {e}")
                                    except Exception as e:
                                        print(f"DEBUG: Failed linking categories to existing app {app_id}: {e}")
                                        if isinstance(e, sqlite3.OperationalError) and 'database is locked' in str(e):
                                            print("DEBUG: Database lock detected, will retry on next pass")
                                    updated_apps += 1
                                except Exception:
                                    pass
                            else:
                                # Create new application via helper; it expects dept ids list and returns app ids
                                # Ensure business unit exists and establish link after creation
                                try:
                                    # Try to create application with retries
                                    max_retries = 3
                                    for retry in range(max_retries):
                                        try:
                                            # Use current row BU(s) or last seen BU(s) for this app if present
                                            eff_bu_names = bu_names if bu_names else app_bu_map.get(app_name, [])
                                            dept_ids = []
                                            for bn in eff_bu_names:
                                                bu_id = ensure_business_unit(bn)
                                                if bu_id and bu_id not in dept_ids:
                                                    dept_ids.append(bu_id)
                                            # Direct SQL insert with our connection
                                            # Compute application risk using the standard formula
                                            app_risk = max(0, (10 - (score or 0)) * (criticality or 0))
                                            from datetime import timezone as _tz
                                            now_iso = datetime.now(_tz.utc).isoformat()
                                            # Determine if legacy 'name' column exists; if so, populate it as well
                                            try:
                                                cur.execute("PRAGMA table_info(applications)")
                                                cols = [r[1] for r in cur.fetchall()]
                                            except Exception:
                                                cols = []
                                            if 'name' in cols:
                                                cur.execute('''
                                                    INSERT INTO applications 
                                                    (division, name, vendor, score, need, criticality, installed, disaster_recovery, 
                                                    safety, security, monetary, customer_service, notes, risk_score, last_modified, category_id)
                                                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                                ''', (app_name, app_name, vendor, score, need, criticality, installed, disaster_recovery,
                                                    safety, security, monetary, customer_service, notes, app_risk, now_iso,
                                                    None))
                                            else:
                                                cur.execute('''
                                                    INSERT INTO applications 
                                                    (division, vendor, score, need, criticality, installed, disaster_recovery, 
                                                    safety, security, monetary, customer_service, notes, risk_score, last_modified, category_id)
                                                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                                ''', (app_name, vendor, score, need, criticality, installed, disaster_recovery,
                                                    safety, security, monetary, customer_service, notes, app_risk, now_iso,
                                                    None))
                                            app_id = cur.lastrowid
                                            # Apply all accumulated categories for new app
                                            try:
                                                all_cats = sorted(app_category_accum.get(app_name, set()))
                                                for nm in all_cats:
                                                    nm_clean = nm.strip()
                                                    if not nm_clean:
                                                        continue
                                                    cur.execute('SELECT id FROM categories WHERE LOWER(TRIM(name)) = LOWER(TRIM(?))', (nm_clean,))
                                                    crow = cur.fetchone()
                                                    if crow:
                                                        cid = crow[0]
                                                    else:
                                                        cur.execute('INSERT INTO categories (name, last_modified) VALUES (?, CURRENT_TIMESTAMP)', (nm_clean,))
                                                        cid = cur.lastrowid
                                                    cur.execute('INSERT OR IGNORE INTO application_categories (app_id, category_id) VALUES (?, ?)', (app_id, cid))
                                                conn.commit()
                                            except Exception as e:
                                                print(f"DEBUG: Failed linking categories to new app {app_id}: {e}")
                                            
                                            # Create business unit links if needed
                                            if dept_ids:
                                                for dept_id in dept_ids:
                                                    if dept_id:  # Only if we got a valid department ID
                                                        try:
                                                            # Check if link already exists
                                                            cur.execute(
                                                                'SELECT 1 FROM application_business_units WHERE app_id = ? AND unit_id = ?',
                                                                (app_id, dept_id)
                                                            )
                                                            if not cur.fetchone():
                                                                cur.execute(
                                                                    'INSERT INTO application_business_units (app_id, unit_id) VALUES (?, ?)',
                                                                    (app_id, dept_id)
                                                                )
                                                        except Exception as e:
                                                            errors.append(f"Failed to link app {app_name} to business unit: {e}")
                                            
                                            # Apply accumulated categories for new app using local connection
                                            try:
                                                all_cats = sorted(app_category_accum.get(app_name, set()))
                                                cat_ids = []
                                                print(f"DEBUG: Setting initial categories for new app {app_name}: {all_cats}")
                                                
                                                # First ensure all categories exist and get their IDs
                                                for nm in all_cats:
                                                    nm_clean = nm.strip()
                                                    if not nm_clean:
                                                        continue
                                                    cur.execute('SELECT id FROM categories WHERE LOWER(TRIM(name)) = LOWER(TRIM(?))', (nm_clean,))
                                                    crow = cur.fetchone()
                                                    if crow:
                                                        cid = crow[0]
                                                    else:
                                                        cur.execute('INSERT INTO categories (name, last_modified) VALUES (?, CURRENT_TIMESTAMP)', (nm_clean,))
                                                        cid = cur.lastrowid
                                                        conn.commit()  # Commit new category immediately
                                                    cat_ids.append(cid)
                                                
                                                # Link categories in same transaction
                                                if app_id is not None and cat_ids:  # Type safety check
                                                    print(f"DEBUG: Setting categories for new app {app_name} (ID: {app_id}): {all_cats}")
                                                    # First remove any existing links
                                                    cur.execute('DELETE FROM application_categories WHERE app_id = ?', (app_id,))
                                                    # Add new links
                                                    for cid in cat_ids:
                                                        cur.execute('INSERT OR IGNORE INTO application_categories (app_id, category_id) VALUES (?, ?)', 
                                                                  (app_id, cid))
                                                    print(f"DEBUG: Successfully linked categories for new app {app_name}")
                                            except Exception as e:
                                                print(f"DEBUG: Failed linking categories to new app {app_id}: {e}")
                                            
                                            conn.commit()
                                            created_apps += 1
                                            app_cache[app_name] = app_id
                                            break
                                        except sqlite3.OperationalError as e:
                                            if 'database is locked' in str(e) and retry < max_retries - 1:
                                                time.sleep(0.5 * (retry + 1))  # Exponential backoff
                                                continue
                                            raise
                                except Exception as e:
                                    errors.append(f"Row {rownum}: Failed to create application '{app_name}': {e}")
                                    continue

                        # Ensure business unit link exists (use current BU(s) or last seen BU(s) for this app)
                        eff_bu_names = bu_names if bu_names else app_bu_map.get(app_name, [])
                        # Link business units if provided
                        if eff_bu_names and app_id:
                            try:
                                for bn in eff_bu_names:
                                    print(f"DEBUG: Ensuring business unit exists for app {app_name}: {bn}")
                                    bu_id = ensure_business_unit(bn)
                                    if bu_id:
                                        print(f"DEBUG: Linking app {app_name} to business unit {bn} (ID: {bu_id})")
                                        # Insert link if not exists
                                        cur.execute('SELECT 1 FROM application_business_units WHERE app_id = ? AND unit_id = ?', (app_id, bu_id))
                                        if not cur.fetchone():
                                            cur.execute('INSERT INTO application_business_units (app_id, unit_id) VALUES (?, ?)', (app_id, bu_id))
                                            conn.commit()  # Commit the link
                                            print(f"DEBUG: Successfully linked app {app_name} to business unit {bn}")
                            except Exception as e:
                                print(f"DEBUG: Failed to link business unit(s) to app {app_name}: {e}")
                                errors.append(f"Failed to link business unit(s) to app {app_name}: {e}")
                                pass

                        # Insert integration regardless of BU presence (as long as we have an app and a name)
                        if app_id and int_name:
                            print(f"DEBUG: Processing integration {int_name} for app {app_name} (ID: {app_id})")
                            try:
                                cat_ids_for_app = []  # ensure defined for later debug logging
                                # Calculate risk score using weighted percentage before inserting
                                ratings = {
                                    'Score': int_score or 0,
                                    'Need': int_need or 0,
                                    'Criticality': int_criticality or 0,
                                    'Installed': int_installed or 0,
                                    'DisasterRecovery': int_dr or 0,
                                    'Safety': int_safety or 0,
                                    'Security': int_security or 0,
                                    'Monetary': int_monetary or 0,
                                    'CustomerService': int_customer_service or 0,
                                }
                                risk_score = float(self._calculate_integration_risk(ratings))

                                print(f"DEBUG: Inserting integration with risk score {risk_score}")
                                # Insert new integration
                                cur.execute('''
                                    INSERT INTO system_integrations 
                                    (parent_app_id, name, vendor, score, need,
                                     criticality, installed, disaster_recovery,
                                     safety, security, monetary, customer_service,
                                     notes, risk_score, last_modified)
                                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                                ''', (app_id, int_name, int_vendor, int_score or 0,
                                     int_need or 0, int_criticality or 0, int_installed or 0,
                                     int_dr or 0, int_safety or 0, int_security or 0,
                                     int_monetary or 0, int_customer_service or 0,
                                     '', risk_score, last_mod or datetime.now().isoformat()))
                                integration_id = cur.lastrowid
                                created_integrations += 1

                                # Determine row-specific categories (only those explicitly present on this row)
                                row_categories = []
                                if main_category:
                                    row_categories.extend(parse_category_names(main_category))
                                if int_category:
                                    row_categories.extend(parse_category_names(int_category))
                                # Normalize & deduplicate row categories
                                row_categories_norm = []
                                seen_cat = set()
                                for rc in row_categories:
                                    rc_clean = rc.strip()
                                    if not rc_clean:
                                        continue
                                    key = rc_clean.lower()
                                    if key not in seen_cat:
                                        seen_cat.add(key)
                                        row_categories_norm.append(rc_clean)

                                # Link integration ONLY to row-specific categories
                                try:
                                    # Ensure join table exists
                                    cur.execute('''CREATE TABLE IF NOT EXISTS integration_categories (
                                        integration_id INTEGER NOT NULL,
                                        category_id INTEGER NOT NULL,
                                        PRIMARY KEY (integration_id, category_id),
                                        FOREIGN KEY (integration_id) REFERENCES system_integrations(id) ON DELETE CASCADE,
                                        FOREIGN KEY (category_id) REFERENCES categories(id) ON DELETE CASCADE
                                    )''')
                                    linked_cat_ids = []
                                    if row_categories_norm:
                                        for cat_name in row_categories_norm:
                                            try:
                                                cur.execute('SELECT id FROM categories WHERE LOWER(TRIM(name)) = LOWER(TRIM(?))', (cat_name,))
                                                c_row = cur.fetchone()
                                                if c_row:
                                                    cid = c_row[0]
                                                else:
                                                    # Create category if somehow not present yet
                                                    cur.execute('INSERT INTO categories (name, last_modified) VALUES (?, CURRENT_TIMESTAMP)', (cat_name,))
                                                    cid = cur.lastrowid
                                                linked_cat_ids.append(cid)
                                                cur.execute('INSERT OR IGNORE INTO integration_categories (integration_id, category_id) VALUES (?, ?)', (integration_id, cid))
                                            except Exception as ic_e:
                                                print(f"DEBUG: Failed linking integration {integration_id} to category '{cat_name}': {ic_e}")
                                    else:
                                        print(f"DEBUG: No row-specific categories for integration {integration_id} ({int_name}); leaving unlinked.")
                                except Exception as link_e:
                                    print(f"DEBUG: Failed to create/link integration categories: {link_e}")
                                conn.commit()
                                print(f"DEBUG: Successfully created integration {int_name} with ID {integration_id} and linked to row categories {row_categories_norm}")
                            except Exception as e:
                                print(f"DEBUG: Failed to create integration: {e}")
                                errors.append(f"Row {rownum}: Failed to create integration '{int_name}': {e}")
                                # Don't continue here - let the row processing complete
                                    
                    except Exception as e:
                        errors.append(f"Row {rownum}: {e}")
                        continue  # Only continue on row-level errors
                        
                    # report progress every 25 rows
                    try:
                        if progress_callback and rownum % 25 == 0:
                            progress_callback(rownum, total_rows)
                    except Exception:
                        pass

                # Cleanup any blank applications created inadvertently
                try:
                    # Final reconciliation: ensure every application's category links reflect the full accumulated set
                    # This guards against any mid-import failures or partial commits.
                    for app_name, cat_set in app_category_accum.items():
                        try:
                            cur.execute('SELECT id FROM applications WHERE division = ?', (app_name,))
                            r = cur.fetchone()
                            if not r:
                                continue
                            app_id_final = r[0]
                            all_cats = sorted(cat_set)
                            # Resolve / create category ids
                            cat_ids_final = []
                            for nm in all_cats:
                                nm_clean = nm.strip()
                                if not nm_clean:
                                    continue
                                cur.execute('SELECT id FROM categories WHERE LOWER(TRIM(name)) = LOWER(TRIM(?))', (nm_clean,))
                                crow = cur.fetchone()
                                if crow:
                                    cat_ids_final.append(crow[0])
                                else:
                                    cur.execute('INSERT INTO categories (name, last_modified) VALUES (?, CURRENT_TIMESTAMP)', (nm_clean,))
                                    cat_ids_final.append(cur.lastrowid)
                            # Replace links only if we have categories
                            if cat_ids_final:
                                cur.execute('DELETE FROM application_categories WHERE app_id = ?', (app_id_final,))
                                for cid in cat_ids_final:
                                    cur.execute('INSERT OR IGNORE INTO application_categories (app_id, category_id) VALUES (?, ?)', (app_id_final, cid))
                                print(f"DEBUG: Reconciled categories for {app_name}: {all_cats}")
                        except Exception as rec_e:
                            print(f"DEBUG: Reconciliation failed for {app_name}: {rec_e}")
                    conn.commit()
                except Exception as e:
                    print(f"DEBUG: Final category reconciliation error: {e}")
                try:
                    if conn and cur:
                        cur.execute("DELETE FROM applications WHERE division IS NULL OR TRIM(division) = ''")
                        conn.commit()
                except Exception:
                    pass

                # commit once after processing
                try:
                    if conn:
                        conn.commit()
                except Exception as e:
                    errors.append(f"Failed to commit changes: {e}")
                finally:
                    try:
                        if conn:
                            conn.close()
                    except Exception:
                        pass

                # return summary
                return {
                    'created_apps': created_apps,
                    'updated_apps': updated_apps,
                    'created_integrations': created_integrations,
                    'updated_integrations': updated_integrations,
                    'errors': errors
                }
        except Exception as e:
            # Ensure connection is closed on error
            try:
                if conn:
                    conn.close()
            except Exception:
                pass
            return {'error': str(e)}

    def import_csv_worker(self, path):
        """Worker run on background thread; schedules UI updates on main thread when finished."""
        # create a progress window on main thread, then run import and close
        # store the temporary window on the instance so nested functions can access it
        def refresh_after_import():
            """Schedule refresh operations on the main thread"""
            # Ensure any pending changes are committed before refreshing UI
            conn = None
            try:
                conn = database.connect_db()
                conn.commit()
            except Exception as e:
                print(f"DEBUG: Final commit before refresh failed: {e}")
            finally:
                if conn:
                    conn.close()
                    
            self.refresh_table()  # Refresh the main application table
            self.department_listbox.delete(0, 'end')  # Clear and refresh department listbox
            for dept_id, dept_name in self.get_departments():
                self.department_listbox.insert('end', dept_name)
            # Refresh divisions
            try:
                self.refresh_division_listbox()
            except Exception as e:
                print(f"DEBUG: Error refreshing division listbox: {e}")
            # Refresh categories so new CSV categories appear immediately
            try:
                self.populate_category_listbox()
            except Exception as e:
                print(f"DEBUG: Error refreshing category listbox: {e}")
            # Update submit button state in case selections changed/are now available
            try:
                self.update_submit_button_state()
            except Exception:
                pass
            
        try:
            self._import_progress_win = None
        except Exception:
            pass

        def show_progress():
            try:
                self._import_progress_win = self.ProgressWindow(self, title='Importing CSV', message='Starting import...')
                self._import_progress_win.show()
            except Exception:
                self._import_progress_win = None
                
        def on_import_complete():
            """Called when import is finished"""
            try:
                refresh_after_import()
            except Exception as e:
                print(f"DEBUG: Error in refresh after import: {e}")
            finally:
                if hasattr(self, '_import_progress_win') and self._import_progress_win:
                    try:
                        self._import_progress_win.close()
                    except Exception:
                        pass

        def update_progress(rows_processed, total_rows):
            try:
                win = getattr(self, '_import_progress_win', None)
                if win is not None:
                    try:
                        win.update_message(f'Importing data...', rows_processed, total_rows)
                    except Exception:
                        pass
            except Exception:
                pass

        def close_progress():
            try:
                win = getattr(self, '_import_progress_win', None)
                if win is not None:
                    try:
                        win.close()
                    except Exception:
                        pass
            except Exception:
                pass

        # show the progress window on the main thread and ensure it's visible
        try:
            self.after(1, show_progress)
            self.update_idletasks()
        except Exception:
            pass

        # run import, passing progress callback
        try:
            result = self.import_csv_file(path, progress_callback=update_progress)
        except Exception as e:
            result = {'error': str(e)}

        # close progress window on main thread
        try:
            self.after(50, close_progress)
        except Exception:
            pass

        def finish():
            # finish runs on main thread via self.after
            try:
                if isinstance(result, dict) and result.get('error'):
                    messagebox.showerror('Import Error', result.get('error'))
                    return
                created_apps = result.get('created_apps', 0)
                updated_apps = result.get('updated_apps', 0)
                created_integrations = result.get('created_integrations', 0)
                updated_integrations = result.get('updated_integrations', 0)
                errors = result.get('errors', [])
                summary = f"Import complete. Apps created: {created_apps}, Apps updated: {updated_apps}, Integrations created: {created_integrations}, Integrations updated: {updated_integrations}."
                if errors:
                    summary += f"\nErrors: {len(errors)} (see console for details)."
                    for e in errors:
                        print('IMPORT ERROR:', e)
                messagebox.showinfo('Import Finished', summary)
            finally:
                # Always try to refresh UI and close progress window
                try:
                    # Close progress window first
                    if hasattr(self, '_import_progress_win') and self._import_progress_win:
                        self._import_progress_win.close()
                except Exception:
                    pass
                try:
                    # Update all UI elements
                    refresh_after_import()  # This refreshes table and department list
                    self.refresh_integration_table()  # Refresh integrations if any are showing
                except Exception as e:
                    print(f"DEBUG: Error refreshing UI after import: {e}")

        # schedule finish on main thread
        try:
            self.after(100, finish)
        except Exception:
            # if after isn't available, call directly (best-effort)
            finish()
    
    def sort_integration_table(self, col, reverse=False):
        """
        Sort the integrations table by the specified column
        """
        try:
            # Store the currently selected items to restore selection after sort
            selected_ids = [int(item) for item in self.integrations_tree.selection()]
            
            # Get data from the tree
            data = []
            for item in self.integrations_tree.get_children():
                values = self.integrations_tree.item(item, 'values')
                item_id = int(item)
                data.append((values, item_id))
            
            # Function to get the key for sorting
            def get_sort_key(item):
                values, _ = item
                col_index = list(self.integrations_tree['columns']).index(col)
                if 0 <= col_index < len(values):
                    value = values[col_index]
                    return self._convert_sort_value(value)
                return ""
            
            # Sort data
            data.sort(key=get_sort_key, reverse=reverse)
            
            # Clear and repopulate the tree with sorted data
            self.integrations_tree.delete(*self.integrations_tree.get_children())
            for values, item_id in data:
                self.integrations_tree.insert('', 'end', iid=str(item_id), values=values)
            
            # Set the sort direction for next click
            self.integrations_tree.heading(col, 
                              command=lambda: self.sort_integration_table(col, not reverse))
            
            # Restore selection if there was one
            if selected_ids:
                for item in self.integrations_tree.get_children():
                    if int(item) in selected_ids:
                        self.integrations_tree.selection_add(item)
                        self.integrations_tree.see(item)  # Make the first selected item visible
                        break
                
        except Exception as e:
            messagebox.showerror('Error', f'Failed to sort integrations table: {e}')
    
    def on_integration_select(self, event):
        """
        Handle integration selection in the integrations table
        """
        sel = self.integrations_tree.selection()
        if not sel:
            return
            
        # Get the selected integration details
        item = sel[0]
        try:
            integration_id = int(item)
            integration = database.get_system_integration(integration_id)
            
            if not integration:
                return
                
            # Display integration details in the details text widget
            self.details_text.configure(state='normal')
            self.details_text.delete('1.0', 'end')
            
            # Format the details nicely with safe conversion
            # Risk score as float with one decimal
            risk_score = 0.0
            try:
                if integration[14] is not None:  # risk_score is at index 14
                    risk_score = float(integration[14])
            except (ValueError, TypeError):
                risk_score = 0.0

            # Determine color tag and band label
            band = database.dr_priority_band(risk_score)
            if risk_score is None or risk_score <= 0:
                risk_tag = 'risk_na'
            else:
                color_name = get_risk_color(risk_score)
                risk_tag = f"risk_{color_name}" if color_name else 'risk_na'

            # Convert last modified to Eastern Time, show without fractional seconds
            last_mod_display = 'N/A'
            try:
                last_raw = integration[15]
                if last_raw:
                    from datetime import datetime, timezone
                    from zoneinfo import ZoneInfo
                    dt = datetime.fromisoformat(last_raw)
                    if dt.tzinfo is None:
                        dt = dt.replace(tzinfo=timezone.utc)
                    et_dt = dt.astimezone(ZoneInfo('America/New_York'))
                    last_mod_display = et_dt.strftime('%Y-%m-%d %H:%M:%S ET')
            except Exception:
                # Fallback to raw value if parsing or tz conversion fails
                last_mod_display = integration[15] or 'N/A'

            # Build details with color-highlighted risk score and ET timestamp
            self.details_text.insert('end', f"Integration: {integration[2]}\n")
            self.details_text.insert('end', f"Vendor: {integration[3] or 'N/A'}\n")
            self.details_text.insert('end', "Risk Score: ")
            if risk_score is None or risk_score <= 0:
                self.details_text.insert('end', 'N/A', risk_tag)
            else:
                self.details_text.insert('end', f"{risk_score:.1f}", risk_tag)
            self.details_text.insert('end', f" ({band})\n")
            self.details_text.insert('end', f"Last Modified: {last_mod_display}\n")
            self.details_text.insert('end', f"Notes: {integration[13] or 'N/A'}")
            self.details_text.configure(state='disabled')
            
        except Exception as e:
            messagebox.showerror('Error', f'Failed to load integration details: {e}')
    
    def on_system_double_click(self, event):
        """
        Handle double-click on a system in the main table
        """
        sel = self.tree.selection()
        if not sel:
            return
            
        item = sel[0]
        try:
            # Get the system ID and set it as current parent
            try:
                self.current_parent_system_id = int(item.split(':', 1)[0])
            except ValueError:
                raise ValueError(f"Invalid system id format: {item}")
            
            # Load integrations for this system
            self.refresh_integration_table()
            
            # Show a message that integrations are available
            system_name = self.tree.item(item, 'values')[0]
            messagebox.showinfo('System Selected', 
                               f'System "{system_name}" selected. You can now view or add integrations.')
            
        except Exception as e:
            messagebox.showerror('Error', f'Failed to load system integrations: {e}')
    
    def delete_selected_integration(self):
        """Delete the selected integration from the integrations tree"""
        # Check if an integration is selected
        sel = self.integrations_tree.selection()
        if not sel:
            messagebox.showinfo('Delete Integration', 'Please select an integration to delete.')
            return
            
        # Get the selected integration ID
        item = sel[0]
        try:
            integration_id = int(item)
            integration = database.get_system_integration(integration_id)
            
            if not integration:
                messagebox.showinfo("Delete Integration", "Could not find the selected integration.")
                return
                
            # Determine integration name for confirmation
            try:
                int_name = integration['name'] if hasattr(integration, 'keys') else integration[2]
            except Exception:
                int_name = integration[2] if len(integration) > 2 else 'Unknown'

            # Confirm deletion
            if messagebox.askyesno('Confirm Delete', f'Are you sure you want to delete the integration "{int_name}"? This cannot be undone.'):
                # Delete the integration from the database
                database.delete_system_integration(integration_id)
                
                # Refresh the integrations table
                self.refresh_integration_table()
                
                messagebox.showinfo('Success', 'Integration deleted successfully')
        except Exception as e:
            messagebox.showerror('Error', f'Failed to delete integration: {e}')
    
    def refresh_integration_table(self, parent_id=None):
        """Refresh the integrations table with current data"""
        if parent_id is None:
            parent_id = self.current_parent_system_id
        
        # Clear existing entries
        self.integrations_tree.delete(*self.integrations_tree.get_children())
        
        if parent_id is None:
            return
            
        try:
            selected_cat = getattr(self, 'selected_category_name', None)
            conn = database.connect_db()
            cur = conn.cursor()
            integrations = []
            try:
                cur.execute('''CREATE TABLE IF NOT EXISTS integration_categories (
                    integration_id INTEGER NOT NULL,
                    category_id INTEGER NOT NULL,
                    PRIMARY KEY (integration_id, category_id),
                    FOREIGN KEY (integration_id) REFERENCES system_integrations(id) ON DELETE CASCADE,
                    FOREIGN KEY (category_id) REFERENCES categories(id) ON DELETE CASCADE
                )''')
                if selected_cat:  # Filtered by selected category
                    cur.execute('''
                        SELECT si.* FROM system_integrations si
                        JOIN integration_categories ic ON si.id = ic.integration_id
                        JOIN categories c ON ic.category_id = c.id
                        WHERE si.parent_app_id = ? AND LOWER(TRIM(c.name)) = LOWER(TRIM(?))
                        ORDER BY si.name
                    ''', (parent_id, selected_cat))
                else:  # No category selected -> show all integrations for the parent
                    cur.execute('''
                        SELECT si.* FROM system_integrations si
                        WHERE si.parent_app_id = ?
                        ORDER BY si.name
                    ''', (parent_id,))
                integrations = cur.fetchall()
            except Exception as fe:
                print(f"DEBUG: Integration query failed: {fe}")
            conn.close()
            
            for row in integrations:
                if not row:
                    continue

                # Prefer named access, fallback to tuple indices
                def _rget(r, key, idx, default=None):
                    try:
                        if hasattr(r, 'keys'):
                            val = r[key]
                            return val if val is not None else default
                    except Exception:
                        pass
                    try:
                        return r[idx]
                    except Exception:
                        return default

                name = str(_rget(row, 'name', 2, ''))
                vendor = str(_rget(row, 'vendor', 3, ''))

                # Extract ratings in the correct order
                rating_keys = ['score', 'need', 'criticality', 'installed', 'disaster_recovery', 'safety', 'security', 'monetary', 'customer_service']
                ratings = []
                for i, rk in enumerate(rating_keys, start=4):
                    try:
                        val = None
                        if hasattr(row, 'keys') and rk in row.keys():
                            val = row[rk]
                        else:
                            val = row[i]
                        ratings.append(int(val) if val is not None else 0)
                    except Exception:
                        ratings.append(0)

                # Format risk_text
                risk_text = "N/A"
                risk_score = 0.0
                try:
                    if hasattr(row, 'keys') and 'risk_score' in row.keys():
                        if row['risk_score'] is not None:
                            risk_score = float(row['risk_score'])
                            risk_text = f"{risk_score:.1f} ({database.dr_priority_band(risk_score)})"
                    else:
                        if row[14] is not None:
                            risk_score = float(row[14])
                            risk_text = f"{risk_score:.1f} ({database.dr_priority_band(risk_score)})"
                except Exception:
                    pass

                # Format last_modified
                last_mod = ""
                lm = None
                try:
                    if hasattr(row, 'keys') and 'last_modified' in row.keys():
                        lm = row['last_modified']
                    else:
                        lm = row[15] if len(row) > 15 else None
                    if lm:
                        last_mod = datetime.fromisoformat(str(lm)).strftime('%Y-%m-%d %H:%M')
                except Exception:
                    try:
                        if lm:
                            last_mod = str(lm)
                    except Exception:
                        last_mod = ''

                # Build values in exact column order expected by the integrations_tree
                # New column order: Name, Vendor, Risk, Score, Criticality, Need, Installed, DR, Safety, Security, Monetary, CustomerService, Last Modified
                score = ratings[0] if len(ratings) > 0 else 0
                need = ratings[1] if len(ratings) > 1 else 0
                criticality = ratings[2] if len(ratings) > 2 else 0
                installed = ratings[3] if len(ratings) > 3 else 0
                dr = ratings[4] if len(ratings) > 4 else 0
                safety = ratings[5] if len(ratings) > 5 else 0
                security = ratings[6] if len(ratings) > 6 else 0
                monetary = ratings[7] if len(ratings) > 7 else 0
                cust_service = ratings[8] if len(ratings) > 8 else 0

                values = [
                    name, vendor, risk_text, score, criticality, need,
                    installed, dr, safety, security, monetary, cust_service,
                    last_mod
                ]

                color = get_risk_color(risk_score)
                if color:
                    self.integrations_tree.insert('', 'end', iid=str(_rget(row, 'id', 0)), values=values, tags=(color,))
                else:
                    self.integrations_tree.insert('', 'end', iid=str(_rget(row, 'id', 0)), values=values)
                
            # Configure the color tags for the tree
            self.integrations_tree.tag_configure('red', background='#ffcccc')
            self.integrations_tree.tag_configure('yellow', background='#fff2cc')
            self.integrations_tree.tag_configure('green', background='#ccffcc')
        
        except Exception as e:
            messagebox.showerror('Error', f'Failed to refresh integrations table: {e}')
            
    def on_integration_double_click(self, event):
        """Handle double-click event on an integration to edit it, using the same UI layout as add_system_integration."""
        sel = self.integrations_tree.selection()
        if not sel:
            return

        item = sel[0]
        try:
            integration_id = int(item)
            integration = database.get_system_integration(integration_id)

            if not integration:
                messagebox.showinfo('Edit Integration', 'Could not find the selected integration.')
                return

            # --- Dialog setup to mirror add_system_integration ---
            dialog = tk.Toplevel(self)
            try:
                int_title = integration['name'] if hasattr(integration, 'keys') else integration[2]
            except Exception:
                int_title = integration[2] if len(integration) > 2 else 'Edit Integration'
            dialog.title(f'Edit Integration: {int_title}')
            dialog.geometry('650x700')
            dialog.minsize(650, 700)
            dialog.resizable(False, False)
            dialog.transient(self)
            dialog.grab_set()

            # Main container with 2 columns: form | score guide
            main_container = ttk.Frame(dialog)
            main_container.grid(row=0, column=0, sticky='nsew', padx=10, pady=10)
            dialog.grid_rowconfigure(0, weight=1)
            dialog.grid_columnconfigure(0, weight=1)
            main_container.grid_rowconfigure(0, weight=1)
            main_container.grid_columnconfigure(0, weight=1)  # form grows
            # Slightly wider guide column for better readability
            main_container.grid_columnconfigure(1, weight=0, minsize=280)  # guide fixed

            # Form area
            form_frame = ttk.Frame(main_container, padding=15)
            form_frame.grid(row=0, column=0, sticky='nsew')
            form_frame.grid_columnconfigure(1, weight=1)

            # Name & Vendor
            ttk.Label(form_frame, text='Integration Name:', anchor='e', style='FormLabel.TLabel').grid(row=0, column=0, sticky='e', pady=5, padx=5)
            name_entry = ttk.Entry(form_frame, width=40)
            name_entry.grid(row=0, column=1, sticky='ew', pady=5, padx=5)
            try:
                name_entry.insert(0, integration['name'] or '')
            except Exception:
                name_entry.insert(0, integration[2] or '')
            name_entry.focus_set()

            ttk.Label(form_frame, text='Vendor:', anchor='e', style='FormLabel.TLabel').grid(row=1, column=0, sticky='e', pady=5, padx=5)
            vendor_entry = ttk.Entry(form_frame, width=40)
            vendor_entry.grid(row=1, column=1, sticky='ew', pady=5, padx=5)
            try:
                vendor_entry.insert(0, integration['vendor'] or '')
            except Exception:
                vendor_entry.insert(0, integration[3] or '')

            # Factors
            factor_entries: dict[str, ttk.Entry] = {}
            keys = ['Score', 'Need', 'Criticality', 'Installed', 'DisasterRecovery', 'Safety', 'Security', 'Monetary', 'CustomerService']
            factor_indices = {  # indices for tuple-style access
                'Score': 4, 'Need': 5, 'Criticality': 6, 'Installed': 7, 'DisasterRecovery': 8,
                'Safety': 9, 'Security': 10, 'Monetary': 11, 'CustomerService': 12
            }
            db_key_map_prefill = {
                'Score': 'score',
                'Need': 'need',
                'Criticality': 'criticality',
                'Installed': 'installed',
                'DisasterRecovery': 'disaster_recovery',
                'Safety': 'safety',
                'Security': 'security',
                'Monetary': 'monetary',
                'CustomerService': 'customer_service',
            }
            for idx, key in enumerate(keys, start=2):
                ttk.Label(form_frame, text=f'{key}:', anchor='e', style='FormLabel.TLabel').grid(row=idx, column=0, sticky='e', pady=5, padx=5)
                entry = ttk.Entry(form_frame, width=5)
                entry.grid(row=idx, column=1, sticky='w', pady=5, padx=5)
                # Pre-fill
                try:
                    inserted = False
                    if hasattr(integration, 'keys'):
                        # Prefer snake_case db key when available
                        dbk = db_key_map_prefill.get(key, key.lower())
                        if dbk in integration and integration[dbk] is not None:
                            entry.insert(0, str(integration[dbk]))
                            inserted = True
                        elif key.lower() in integration and integration[key.lower()] is not None:
                            entry.insert(0, str(integration[key.lower()]))
                            inserted = True
                    if not inserted:
                        # Try tuple index fallback
                        if integration[factor_indices[key]] is not None:
                            entry.insert(0, str(integration[factor_indices[key]]))
                except Exception:
                    pass
                factor_entries[key] = entry

            # Notes
            ttk.Label(form_frame, text='Notes:', anchor='e', style='FormLabel.TLabel').grid(row=11, column=0, sticky='ne', pady=5, padx=5)
            notes_text = tk.Text(form_frame, height=4, width=40, wrap='word')
            notes_text.grid(row=11, column=1, sticky='w', pady=5, padx=5)
            try:
                existing_notes = integration['notes'] if hasattr(integration, 'keys') else integration[13]
            except Exception:
                existing_notes = integration[13] if len(integration) > 13 else ''
            if existing_notes:
                notes_text.insert('1.0', existing_notes)

            # Empty Score Guide labelframe on the right
            guide_frame = ttk.Labelframe(main_container, text='Score Guide', padding=12, style='Guide.TLabelframe')
            guide_frame.grid(row=0, column=1, sticky='n', padx=(10, 0), pady=10)
            # Title + body to improve readability and match style
            baseline_msg = 'Entering a 0 for any factor will exclude that factor from the Risk calculation. Ratings should be from 1 (low) to 10 (high).'
            guide_title_label = ttk.Label(guide_frame, text='Note --- ', anchor='w', style='GuideNumber.TLabel')
            guide_title_label.grid(row=0, column=0, sticky='w', pady=(0, 2))
            ttk.Separator(guide_frame, orient='horizontal').grid(row=1, column=0, sticky='ew', pady=(0, 6))
            guide_body_label = ttk.Label(guide_frame, text=baseline_msg, justify='left', anchor='nw', wraplength=260, style='GuideText.TLabel')
            guide_body_label.grid(row=2, column=0, sticky='nw')
            # Spacer to ensure visibility (empty box)
            guide_spacer = ttk.Frame(guide_frame, width=280, height=220)
            guide_spacer.grid(row=3, column=0)
            guide_spacer.grid_propagate(False)

            # Dynamic guide: show guidance text for each field on focus/click
            if 'Score' in factor_entries:
                SCORE_GUIDE_TEXT = (
                    "Guiding Question\n"
                    "• How would you rate the overall performance and reliability of this system today?\n\n"
                    "Scoring Help\n"
                    "• 1–3: Poor/unstable performance or frequent issues\n"
                    "• 4–6: Adequate for current needs but with notable limitations\n"
                    "• 7–8: Strong performance/reliability with minor gaps\n"
                    "• 9–10: Excellent/mission-grade performance and reliability"
                )
                def _show_score_guide(_e=None):
                    try:
                        guide_title_label.config(text='Score')
                        guide_body_label.config(text=SCORE_GUIDE_TEXT)
                    except Exception:
                        pass
                factor_entries['Score'].bind('<FocusIn>', _show_score_guide)
                factor_entries['Score'].bind('<Button-1>', _show_score_guide)

            if 'Need' in factor_entries:
                NEED_GUIDE_TEXT = (
                    "Guiding Question\n"
                    "• Is this system required to operate the business, or is it optional/nice-to-have?\n\n"
                    "Scoring Help\n"
                    "• 1–3: Optional; workarounds readily available\n"
                    "• 4–6: Helpful but not strictly required; alternatives exist\n"
                    "• 7–8: Required for multiple teams or core workflows\n"
                    "• 9–10: Absolutely required to operate or meet obligations"
                )
                def _show_need_guide(_e=None):
                    try:
                        guide_title_label.config(text='Need')
                        guide_body_label.config(text=NEED_GUIDE_TEXT)
                    except Exception:
                        pass
                factor_entries['Need'].bind('<FocusIn>', _show_need_guide)
                factor_entries['Need'].bind('<Button-1>', _show_need_guide)

            if 'Criticality' in factor_entries:
                CRIT_GUIDE_TEXT = (
                    "Guiding Question\n"
                    "• What is the business impact if this system is unavailable or fails?\n\n"
                    "Scoring Help\n"
                    "• 1–3: Minor inconvenience; small/localized impact\n"
                    "• 4–6: Moderate productivity loss; delays but manageable\n"
                    "• 7–8: Major disruption to critical processes\n"
                    "• 9–10: Immediate stoppage/safety/regulatory impact"
                )
                def _show_crit_guide(_e=None):
                    try:
                        guide_title_label.config(text='Criticality')
                        guide_body_label.config(text=CRIT_GUIDE_TEXT)
                    except Exception:
                        pass
                factor_entries['Criticality'].bind('<FocusIn>', _show_crit_guide)
                factor_entries['Criticality'].bind('<Button-1>', _show_crit_guide)

            if 'Installed' in factor_entries:
                INST_GUIDE_TEXT = (
                    "Guiding Question\n"
                    "• How entrenched and widely adopted is this system across the organization?\n\n"
                    "Scoring Help\n"
                    "• 1–3: Pilot/limited use; easy to replace\n"
                    "• 4–6: Moderate adoption; switching cost exists but manageable\n"
                    "• 7–8: Widely adopted; high dependency across teams\n"
                    "• 9–10: Deeply entrenched; very high switching cost/risk"
                )
                def _show_inst_guide(_e=None):
                    try:
                        guide_title_label.config(text='Installed')
                        guide_body_label.config(text=INST_GUIDE_TEXT)
                    except Exception:
                        pass
                factor_entries['Installed'].bind('<FocusIn>', _show_inst_guide)
                factor_entries['Installed'].bind('<Button-1>', _show_inst_guide)

            if 'DisasterRecovery' in factor_entries:
                DR_GUIDE_TEXT = (
                    "Guiding Question\n"
                    "• How quickly must this system be recovered after an outage, and how much data loss is tolerable?\n\n"
                    "Scoring Help\n"
                    "• 1–3: Long RTO/RPO acceptable; daily backups sufficient\n"
                    "• 4–6: Recovery within hours required; some data loss acceptable\n"
                    "• 7–8: Near real-time recovery/replication expected\n"
                    "• 9–10: Continuous availability; near-zero downtime/data loss"
                )
                def _show_dr_guide(_e=None):
                    try:
                        guide_title_label.config(text='Disaster Recovery')
                        guide_body_label.config(text=DR_GUIDE_TEXT)
                    except Exception:
                        pass
                factor_entries['DisasterRecovery'].bind('<FocusIn>', _show_dr_guide)
                factor_entries['DisasterRecovery'].bind('<Button-1>', _show_dr_guide)

            if 'Safety' in factor_entries:
                SAFETY_GUIDE_TEXT = (
                    "Guiding Question\n"
                    "• Could failure of this system impact human safety or regulatory health/safety obligations?\n\n"
                    "Scoring Help\n"
                    "• 1–2: No relation to safety\n"
                    "• 3–4: Indirect/low safety relevance\n"
                    "• 5–6: Some safety implication depending on context\n"
                    "• 7–8: High safety implications in certain scenarios\n"
                    "• 9–10: Life/safety critical"
                )
                def _show_safety_guide(_e=None):
                    try:
                        guide_title_label.config(text='Safety')
                        guide_body_label.config(text=SAFETY_GUIDE_TEXT)
                    except Exception:
                        pass
                factor_entries['Safety'].bind('<FocusIn>', _show_safety_guide)
                factor_entries['Safety'].bind('<Button-1>', _show_safety_guide)

            if 'Security' in factor_entries:
                SECURITY_GUIDE_TEXT = (
                    "Guiding Question\n"
                    "• What is the security exposure if this system is compromised or misused?\n\n"
                    "Scoring Help\n"
                    "• 1–3: Isolated/low-sensitivity data; limited access\n"
                    "• 4–6: Internal data with moderate sensitivity\n"
                    "• 7–8: Handles sensitive/regulated data or external integrations\n"
                    "• 9–10: High-risk target; critical/regulated data and broad access"
                )
                def _show_security_guide(_e=None):
                    try:
                        guide_title_label.config(text='Security')
                        guide_body_label.config(text=SECURITY_GUIDE_TEXT)
                    except Exception:
                        pass
                factor_entries['Security'].bind('<FocusIn>', _show_security_guide)
                factor_entries['Security'].bind('<Button-1>', _show_security_guide)

            if 'Monetary' in factor_entries:
                MONETARY_GUIDE_TEXT = (
                    "Guiding Question\n"
                    "• What is the direct financial impact of downtime, failure, or poor performance?\n\n"
                    "Scoring Help\n"
                    "• 1–3: Minimal/indirect cost impact\n"
                    "• 4–6: Moderate costs or revenue impact\n"
                    "• 7–8: High revenue impact or material cost exposure\n"
                    "• 9–10: Immediate material revenue loss/penalties"
                )
                def _show_monetary_guide(_e=None):
                    try:
                        guide_title_label.config(text='Monetary')
                        guide_body_label.config(text=MONETARY_GUIDE_TEXT)
                    except Exception:
                        pass
                factor_entries['Monetary'].bind('<FocusIn>', _show_monetary_guide)
                factor_entries['Monetary'].bind('<Button-1>', _show_monetary_guide)

            if 'CustomerService' in factor_entries:
                CS_GUIDE_TEXT = (
                    "Guiding Question\n"
                    "• How directly does this system affect the customer experience or service quality?\n\n"
                    "Scoring Help\n"
                    "• 1–3: Internal-only; no direct customer effect\n"
                    "• 4–6: Indirect impact on customer interactions\n"
                    "• 7–8: Customer-facing; noticeable impact on service quality\n"
                    "• 9–10: Primary channel or critical to customer experience"
                )
                def _show_cs_guide(_e=None):
                    try:
                        guide_title_label.config(text='Customer Service')
                        guide_body_label.config(text=CS_GUIDE_TEXT)
                    except Exception:
                        pass
                factor_entries['CustomerService'].bind('<FocusIn>', _show_cs_guide)
                factor_entries['CustomerService'].bind('<Button-1>', _show_cs_guide)
            # Separator under content
            ttk.Separator(main_container, orient='horizontal').grid(row=1, column=0, columnspan=2, sticky='ew', pady=(10, 5))

            # Buttons row
            button_frame = ttk.Frame(main_container)
            button_frame.grid(row=2, column=0, columnspan=2, sticky='ew', pady=5)
            button_frame.grid_columnconfigure(0, weight=1)

            def update_integration():
                try:
                    name = name_entry.get().strip()
                    if not name:
                        messagebox.showwarning('Validation Error', 'Integration name is required')
                        return

                    fields: dict[str, object] = {}
                    fields['name'] = name
                    fields['vendor'] = vendor_entry.get().strip()
                    fields['notes'] = notes_text.get('1.0', 'end-1c').strip()

                    ratings: dict[str, int] = {}
                    db_key_map = {
                        'Score': 'score',
                        'Need': 'need',
                        'Criticality': 'criticality',
                        'Installed': 'installed',
                        'DisasterRecovery': 'disaster_recovery',
                        'Safety': 'safety',
                        'Security': 'security',
                        'Monetary': 'monetary',
                        'CustomerService': 'customer_service',
                    }
                    for key in keys:
                        try:
                            value = factor_entries[key].get().strip()
                            if value:
                                iv = int(value)
                                ratings[key] = iv
                                fields[db_key_map.get(key, key.lower())] = iv
                            else:
                                ratings[key] = 0
                                fields[db_key_map.get(key, key.lower())] = 0
                        except ValueError:
                            messagebox.showwarning('Validation Error', f'{key} must be a number')
                            return
                    # Weighted risk percentage 0–100
                    fields['risk_score'] = float(self._calculate_integration_risk(ratings))

                    success = database.update_system_integration(integration_id, fields)
                    if success:
                        messagebox.showinfo('Success', 'Integration updated successfully')
                        dialog.destroy()
                        self.refresh_integration_table()
                    else:
                        messagebox.showerror('Error', 'Failed to update integration')
                except Exception as e:
                    messagebox.showerror('Error', f'Failed to update integration: {e}')

            def delete_integration():
                if messagebox.askyesno('Confirm Delete', 'Are you sure you want to delete this integration? This cannot be undone.'):
                    try:
                        database.delete_system_integration(integration_id)
                        messagebox.showinfo('Success', 'Integration deleted successfully')
                        dialog.destroy()
                        self.refresh_integration_table()
                    except Exception as e:
                        messagebox.showerror('Error', f'Failed to delete integration: {e}')

            update_btn = ttk.Button(button_frame, text='Update', command=update_integration, style='Primary.TButton', width=15)
            delete_btn = ttk.Button(button_frame, text='Delete', command=delete_integration, style='Danger.TButton', width=15)
            cancel_btn = ttk.Button(button_frame, text='Cancel', command=dialog.destroy, style='Secondary.TButton', width=15)
            # place buttons: cancel on left, update on right like add dialog
            cancel_btn.grid(row=0, column=0, padx=5)
            update_btn.grid(row=0, column=1, padx=5)
            delete_btn.grid(row=0, column=2, padx=5)

            dialog.bind('<Return>', lambda _e: update_integration())
            dialog.bind('<Escape>', lambda _e: dialog.destroy())

        except Exception as e:
            messagebox.showerror('Error', f'Failed to load integration for editing: {e}')
            
if __name__ == '__main__':
    app = AppTracker()
    app.mainloop()
